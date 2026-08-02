"""Формирование Excel-отчёта и визуализации по найденным GMV-аномалиям.

Модуль не содержит бизнес-логики отбора аномалий: он только представляет уже
рассчитанный результат. Excel-контракт состоит из девяти листов и защищён
регрессионным тестом ``test_pipeline_excel_sheet_contract``. Дополнительно
строится необязательный DAG-граф сегментов по листу «Анализ аномалий».
"""

from __future__ import annotations

import json
import math
from pathlib import Path
from textwrap import wrap
from typing import Dict, List, Optional, Sequence, Tuple

import pandas as pd
from openpyxl.styles import PatternFill

from .anomaly_scoring import _safe_float
from .config import AnomalyThresholds, MANAGER_METRIC_PCT_COLUMNS
# FIXED: Общий парсер ключа сегмента вместо приватного имени из set_packing.
from .segment_keys import parse_segment_key_parts


def _format_rub(value: float) -> str:
    """Отформатировать рублёвое значение без дробной части.

    Args:
        value: Числовое значение.

    Returns:
        Строка с разделителем групп разрядов.

    Raises:
        ValueError: Не выбрасывается.

    Examples:
        >>> _format_rub(1234567.89)
        '1 234 568'
    """

    return f"{value:,.0f}".replace(",", " ")


# ADDED: Компактная подпись ребёнка без уже показанных родительских срезов.
def _dominant_child_slice_suffix(parent_segment: object, child_segment: object) -> str:
    """Вернуть только срезы ребёнка, которых нет у родительского сегмента.

    Args:
        parent_segment: Человекочитаемый ключ родительского сегмента.
        child_segment: Человекочитаемый ключ доминирующего ребёнка.

    Returns:
        Значения добавленных срезов через ``*`` либо пустую строку, если
        дочерний сегмент не расширяет родительский.

    Raises:
        ValueError: Не выбрасывается.

    Examples:
        >>> _dominant_child_slice_suffix(
        ...     'business=SMB × payment=QR',
        ...     'business=SMB × payment=QR × country=РФ × product=FULLPAYMENT',
        ... )
        'РФ*FULLPAYMENT'
    """

    parent_parts = set(parse_segment_key_parts(parent_segment))
    child_parts = parse_segment_key_parts(child_segment)
    return "*".join(
        value
        for part, value in child_parts
        if (part, value) not in parent_parts
    )


def _wrap_tree_detail_line(value: object, width: int = 34) -> List[str]:
    """Разбить подпись узла на строки, помещающиеся в карточку графа.

    Args:
        value: Текст подписи.
        width: Максимальное число символов в строке.

    Returns:
        Непустые строки текста, разбитые по заданной ширине.

    Raises:
        ValueError: Если ``width`` меньше единицы.

    Examples:
        >>> _wrap_tree_detail_line('РФ*FULLPAYMENT', width=8)
        ['РФ*FULLP', 'AYMENT']
    """

    if width < 1:
        raise ValueError("width должен быть не меньше единицы")
    text = str(value).strip()
    return wrap(text, width=width, break_long_words=True, break_on_hyphens=False)


def _format_pct(value: object) -> object:
    """Отформатировать относительное изменение как процент для менеджерского вывода.

    Args:
        value: Числовая доля, например 0.15 для 15%.

    Returns:
        Строка с процентом или пустая строка, если процент нельзя корректно посчитать.

    Raises:
        ValueError: Не выбрасывается.

    Examples:
        >>> _format_pct(0.1234)
        '12.34%'
    """

    if value is None or pd.isna(value):
        return ""
    numeric_value = float(value)
    if math.isnan(numeric_value) or math.isinf(numeric_value):
        return ""
    return f"{numeric_value * 100:.2f}%"


def build_manager_summary(
    final_df: pd.DataFrame,
    thresholds: AnomalyThresholds,
    current_total_gmv: float,
    candidates: Optional[pd.DataFrame] = None,
) -> pd.DataFrame:
    """Сформировать менеджерский вывод по необычным сегментам.

    Args:
        final_df: Итоговые выбранные аномалии.
        thresholds: Пороги алгоритма.
        current_total_gmv: Total GMV текущей недели.
        candidates: Диагностика всех кандидатов для добавления структурных изменений.

    Returns:
        Таблица менеджерского вывода.

    Raises:
        ValueError: Если ``max_manager_facts`` меньше единицы.

    Examples:
        >>> # manager_df = build_manager_summary(final_df, AnomalyThresholds(), 1000)
    """

    # ADDED: Техническое предусловие обращения к первой строке top-N.
    if thresholds.max_manager_facts < 1:
        raise ValueError("max_manager_facts должен быть не меньше 1")

    def metric_pct_output(row: Optional[pd.Series] = None) -> Dict[str, object]:
        """Вернуть отформатированные WoW-проценты для менеджерской строки.

        Args:
            row: Строка выбранного сегмента или None для служебных строк.

        Returns:
            Словарь с колонками процентов по GMV и операционным метрикам.

        Raises:
            ValueError: Не выбрасывается.

        Examples:
            >>> metric_pct_output(None)["GMV WoW %"]
            ''
        """

        if row is None:
            return {display_col: "" for display_col in MANAGER_METRIC_PCT_COLUMNS.values()}
        return {
            display_col: _format_pct(row.get(source_col))
            for source_col, display_col in MANAGER_METRIC_PCT_COLUMNS.items()
        }

    structure_state_labels = {
        "новый сегмент": "новый",
        "возобновившийся сегмент": "возобновившийся",
        "исчезнувший сегмент": "исчезнувший",
    }

    def structure_change_label(row: pd.Series) -> str:
        """ADDED: Вернуть короткий тип структурного изменения.

        Args:
            row: Строка кандидата или выбранной аномалии.

        Returns:
            Одно из значений `новый`, `возобновившийся`, `исчезнувший` или пустая строка.

        Raises:
            ValueError: Не выбрасывается.

        Examples:
            >>> structure_change_label(pd.Series({'state': 'новый сегмент'}))
            'новый'
        """

        return structure_state_labels.get(str(row.get("state", "")).strip(), "")

    def structure_change_interpretation(row: pd.Series) -> str:
        """ADDED: Сформировать интерпретацию структурного изменения для менеджера.

        Args:
            row: Строка кандидата или выбранной аномалии.

        Returns:
            Короткий тип изменения.

        Raises:
            ValueError: Не выбрасывается.

        Examples:
        >>> structure_change_interpretation(pd.Series({'state': 'новый сегмент', 'gmv_current': 10, 'gmv_previous': 0, 'wow_delta_gmv': 10}))
            'новый'
        """

        return structure_change_label(row)

    def structure_change_rows() -> List[Dict[str, object]]:
        """ADDED: Собрать дополнительные строки менеджерского вывода по структурным изменениям.

        Args:
            Нет аргументов.

        Returns:
            Список строк для вкладки `01_Менеджерский_вывод`.

        Raises:
            ValueError: Не выбрасывается.

        Examples:
            >>> structure_change_rows()
            []
        """

        if candidates is None or candidates.empty or "state" not in candidates.columns:
            return []
        state_order = {"новый сегмент": 0, "возобновившийся сегмент": 1, "исчезнувший сегмент": 2}
        structure_df = candidates[
            candidates["state"].astype(str).isin(structure_state_labels)
            & candidates["slice_depth"].astype(int).gt(0)
        ].copy()
        if structure_df.empty:
            return []
        structure_df["state_order"] = structure_df["state"].astype(str).map(state_order).fillna(99).astype(int)
        structure_df["abs_delta_sort"] = structure_df["wow_delta_gmv"].astype(float).abs()
        structure_df = structure_df.sort_values(
            ["state_order", "abs_delta_sort", "slice_depth", "segment_key"],
            ascending=[True, False, False, True],
            kind="stable",
        )
        rows_out: List[Dict[str, object]] = []
        for _, row in structure_df.iterrows():
            rows_out.append(
                {
                    "раздел": "Изменение структуры",
                    "тип": "Изменение структуры",
                    "сегмент": str(row["segment_key"]),
                    "Delta GMV": _format_rub(float(row["wow_delta_gmv"])),
                    **metric_pct_output(row),
                    "z_score": round(float(row["robust_z"]), 2),
                    "интерпретация": structure_change_interpretation(row),
                }
            )
        return rows_out

    rows: List[Dict[str, object]] = [
        {
            "раздел": "Заголовок",
            "тип": "",
            "сегмент": "Менеджерский вывод по необычным сегментам GMV",
            "Delta GMV": "",
            **metric_pct_output(None),
            "z_score": "",
            "интерпретация": f"Total GMV текущей недели: {_format_rub(current_total_gmv)}.",
        }
    ]

    if final_df.empty:
        rows.append(
            {
                "раздел": "Краткий вывод",
                "тип": "",
                "сегмент": "",
                "Delta GMV": "",
                **metric_pct_output(None),
                "z_score": "",
                "интерпретация": "Материальные необычные сегменты по заданным порогам не найдены.",
            }
        )
    else:
        top = final_df.head(thresholds.max_manager_facts)
        main = top.iloc[0]
        rows.append(
            {
                "раздел": "Краткий вывод",
                "тип": "оптимальная аномалия",
                "сегмент": str(main["segment_key"]),
                "Delta GMV": _format_rub(float(main["wow_delta_gmv"])),
                **metric_pct_output(main),
                "z_score": round(float(main["robust_z"]), 2),
                "интерпретация": (
                    "GMV сегмента "
                    f"{'вырос' if float(main['wow_delta_gmv']) > 0 else 'снизился'} "
                    "относительно предыдущего периода"
                ),
            }
        )

        for _, row in top.iterrows():
            direction = (
                "вырос"
                if float(row["wow_delta_gmv"]) > 0
                else "снизился"
            )
            interpretation = (
                f"GMV сегмента {direction} относительно предыдущего периода"
            )
            rows.append(
                {
                    "раздел": "Таблица факторов",
                    "тип": "оптимальная аномалия",
                    "сегмент": str(row["segment_key"]),
                    "Delta GMV": _format_rub(float(row["wow_delta_gmv"])),
                    **metric_pct_output(row),
                    "z_score": round(float(row["robust_z"]), 2),
                    "интерпретация": interpretation,
                }
            )

    rows.extend(structure_change_rows())

    return pd.DataFrame(rows)


def build_history_for_selected(panel_df: pd.DataFrame, final_df: pd.DataFrame, total_by_date: pd.Series) -> pd.DataFrame:
    """Подготовить историю недель для выбранных аномалий.

    Args:
        panel_df: Полная недельная панель.
        final_df: Итоговые выбранные аномалии.
        total_by_date: Total GMV по неделям.

    Returns:
        Таблица истории выбранных сегментов.

    Raises:
        ValueError: Не выбрасывается.

    Examples:
        >>> # history = build_history_for_selected(panel, final_df, total)
    """

    if final_df.empty:
        return pd.DataFrame()

    selected_ids = final_df["segment_id"].astype(str).tolist()
    history = panel_df[panel_df["segment_id"].astype(str).isin(selected_ids)].copy()
    history["total_gmv"] = history["cal_date"].map(total_by_date.to_dict()).astype(float)
    history["share_in_total"] = history["gmv"] / history["total_gmv"]
    history = history.sort_values(["segment_key", "cal_date"])
    history["share_delta"] = history.groupby("segment_id")["share_in_total"].diff()
    return history[
        [
            "segment_id",
            "segment_key",
            "slice_depth",
            "cal_date",
            "gmv",
            "total_gmv",
            "share_in_total",
            "share_delta",
            "row_missing_in_source",
        ]
    ].reset_index(drop=True)


def build_missing_zero_report(panel_df: pd.DataFrame, dates: Sequence[int], current_cal_date: int) -> pd.DataFrame:
    """Сформировать отчёт по пропускам и нулевому GMV.

    Args:
        panel_df: Полная недельная панель.
        dates: Список недель.
        current_cal_date: Текущая неделя.

    Returns:
        Таблица сегментов с пропусками или нулевыми неделями.

    Raises:
        ValueError: Не выбрасывается.

    Examples:
        >>> # missing_df = build_missing_zero_report(panel, dates, dates[-1])
    """

    previous_cal_date = int(dates[list(dates).index(int(current_cal_date)) - 1])
    grouped = panel_df.groupby(["segment_id", "segment_key", "segment_level", "slice_depth"], as_index=False).agg(
        missing_week_count=("row_missing_in_source", "sum"),
        zero_week_count=("gmv", lambda s: int((s == 0).sum())),
        nonzero_week_count=("gmv", lambda s: int((s > 0).sum())),
    )
    current = panel_df[panel_df["cal_date"].astype(int) == int(current_cal_date)][["segment_id", "gmv", "row_missing_in_source"]]
    previous = panel_df[panel_df["cal_date"].astype(int) == previous_cal_date][["segment_id", "gmv", "row_missing_in_source"]]
    current = current.rename(columns={"gmv": "gmv_current", "row_missing_in_source": "current_row_missing"})
    previous = previous.rename(columns={"gmv": "gmv_previous", "row_missing_in_source": "previous_row_missing"})
    report = grouped.merge(previous, how="left", on="segment_id").merge(current, how="left", on="segment_id")
    report = report[(report["missing_week_count"] > 0) | (report["zero_week_count"] > 0)].copy()
    return report.sort_values(["missing_week_count", "zero_week_count", "segment_key"], ascending=[False, False, True]).reset_index(drop=True)


def build_control_table(
    history_df: pd.DataFrame,
    panel_df: pd.DataFrame,
    candidates: pd.DataFrame,
    final_df: pd.DataFrame,
    coverage: Dict[str, frozenset[str]],
    dates: Sequence[int],
    current_cal_date: int,
    total_by_date: pd.Series,
) -> pd.DataFrame:
    """Сформировать контрольные показатели результата.

    Args:
        history_df: Исходная очищенная таблица.
        panel_df: Полная недельная панель.
        candidates: Диагностика кандидатов.
        final_df: Итоговые выбранные аномалии.
        coverage: Покрытие кандидатов атомами.
        dates: Список недель.
        current_cal_date: Текущая неделя.
        total_by_date: Total GMV по неделям.

    Returns:
        Таблица контрольных показателей.

    Raises:
        ValueError: Не выбрасывается.

    Examples:
        >>> # control = build_control_table(history_df, panel, candidates, final_df, coverage, dates, dates[-1], total)
    """

    selected_atoms: List[str] = []
    if not final_df.empty and "selected_atomic_descendants" in final_df.columns:
        for value in final_df["selected_atomic_descendants"].fillna("").astype(str):
            selected_atoms.extend(
                atom_id.strip()
                for atom_id in value.split(" || ")
                if atom_id.strip()
            )
    else:
        for segment_id in final_df["segment_id"].astype(str).tolist() if not final_df.empty else []:
            selected_atoms.extend(list(coverage.get(segment_id, frozenset())))
    selected_atom_unique_count = len(set(selected_atoms))
    double_count_violation_count = len(selected_atoms) - selected_atom_unique_count
    max_depth = int(candidates["slice_depth"].max()) if not candidates.empty else 0
    atomic_count = int((candidates["slice_depth"].astype(int) == max_depth).sum()) if not candidates.empty else 0

    rows = [
        ("source_rows", len(history_df)),
        ("panel_rows", len(panel_df)),
        ("week_count", len(dates)),
        ("previous_cal_date", int(dates[list(dates).index(int(current_cal_date)) - 1])),
        ("current_cal_date", int(current_cal_date)),
        ("current_total_gmv", float(total_by_date.loc[current_cal_date])),
        ("candidate_count", len(candidates)),
        ("eligible_candidate_count", int(candidates["is_eligible"].astype(bool).sum()) if "is_eligible" in candidates else 0),
        ("selected_count", len(final_df)),
        (
            "set_packing_unresolved_count",
            int((candidates["is_eligible"].astype(bool) & ~candidates["is_resolved"].astype(bool)).sum())
            if {"is_eligible", "is_resolved"}.issubset(candidates.columns)
            else 0,
        ),
        (
            "set_packing_not_proven_count",
            int(candidates.get("set_packing_status", pd.Series("", index=candidates.index)).astype(str).eq("SET_PACKING_NOT_PROVEN").sum())
            if not candidates.empty and "set_packing_status" in candidates
            else 0,
        ),
        ("atomic_count", atomic_count),
        ("selected_atomic_unique_count", selected_atom_unique_count),
        ("double_count_violation_count", double_count_violation_count),
        (
            "set_packing_global_status",
            str(candidates.get("set_packing_global_status", pd.Series([""], index=[0])).dropna().astype(str).replace("", pd.NA).dropna().iloc[0])
            if "set_packing_global_status" in candidates and not candidates["set_packing_global_status"].dropna().astype(str).replace("", pd.NA).dropna().empty
            else "",
        ),
        (
            "set_packing_component_count",
            int(candidates.get("set_packing_component_id", pd.Series("", index=candidates.index)).astype(str).replace("", pd.NA).dropna().nunique())
            if not candidates.empty and "set_packing_component_id" in candidates
            else 0,
        ),
        (
            "set_packing_non_optimal_component_count",
            int(
                candidates.loc[
                    candidates.get("set_packing_component_id", pd.Series("", index=candidates.index)).astype(str).ne("")
                    & candidates.get("set_packing_solver_status", pd.Series("", index=candidates.index)).astype(str).ne("OPTIMAL"),
                    "set_packing_component_id",
                ].astype(str).nunique()
            )
            if not candidates.empty and {"set_packing_component_id", "set_packing_solver_status"}.issubset(candidates.columns)
            else 0,
        ),
        (
            "set_packing_objective_value",
            float(final_df["selection_score"].astype(float).sum()) if not final_df.empty and "selection_score" in final_df else 0.0,
        ),
        (
            "set_packing_conflict_pair_count",
            int(candidates["conflict_count"].astype(int).sum() // 2) if "conflict_count" in candidates else 0,
        ),
        ("filled_missing_rows", int(panel_df["row_missing_in_source"].sum())),
    ]
    return pd.DataFrame(rows, columns=["показатель", "значение"])


def build_anomaly_analysis_sheet(candidates: pd.DataFrame, final_df: pd.DataFrame, thresholds: AnomalyThresholds) -> pd.DataFrame:
    """Сформировать лист анализа аномалий, прошедших первичный фильтр.

    Args:
        candidates: Диагностика кандидатов после оптимизационного отбора.
        final_df: Итоговые выбранные аномалии.
        thresholds: Пороги алгоритма, включая лимит фактов менеджерского вывода.

    Returns:
        Таблица аномалий, прошедших первичный фильтр, с причиной отсутствия в менеджерском выводе.

    Raises:
        ValueError: Не выбрасывается.

    Examples:
        >>> # analysis_df = build_anomaly_analysis_sheet(candidates, final_df, AnomalyThresholds())
    """

    columns = [
        "сегмент",
        "глубина",
        "z_scope",
        "z_scale_source",
        "z_uses_sigma_floor",
        "base_anomaly_score",
        "hierarchy_eligible_descendant_count",
        "hierarchy_group_count",
        "hierarchy_best_group_size",
        "hierarchy_best_group_ids_json",
        "hierarchy_best_group_segment_keys",
        "hierarchy_best_group_score",
        "hierarchy_direction_unity",
        "hierarchy_dominant_share",
        "hierarchy_balance_max",
        "hierarchy_balance_effective",
        "hierarchy_balance",
        "hierarchy_coherence",
        "hierarchy_coherence_adjustment",
        "hierarchy_single_child_capture",
        "hierarchy_single_child_direction_match",
        "hierarchy_single_child_uncapped_score",
        "hierarchy_dominance_cap_score",
        "hierarchy_dominance_rule_matches",
        "hierarchy_dominance_cap_applied",
        "hierarchy_dominant_child_segment",
        "hierarchy_score_factor",
        "anomaly_score",
        "выбран",
        "разрешён",
        "set_packing_status",
        "set_packing_reason",
        "set_packing_global_status",
        "set_packing_component_id",
        "set_packing_solver",
        "set_packing_solver_status",
        "set_packing_abs_gap",
        "set_packing_rel_gap",
        "set_packing_component_score_min",
        "set_packing_component_score_max",
        "set_packing_component_score_dynamic_range",
        "conflict_count",
        "conflict_segment_keys",
        "selected_atomic_count",
        "номер добавления в менеджерский вывод",
        "причина не попадания в менеджерский вывод",
        "Delta GMV",
    ]
    if candidates.empty or "passes_initial_anomaly_filter" not in candidates.columns:
        return pd.DataFrame(columns=columns)

    initial_anomaly_mask = (
        candidates["passes_initial_anomaly_filter"].eq(True)
        & candidates["wow_delta_gmv"].astype(float).abs().ge(
            thresholds.min_anomaly_abs
        )
    )
    analysis_mask = initial_anomaly_mask
    analysis = candidates[analysis_mask].copy()
    if analysis.empty:
        return pd.DataFrame(columns=columns)

    # ADDED: В граф передаётся человекочитаемый ключ ребёнка, а не его ID.
    segment_key_by_id = (
        candidates.assign(segment_id=candidates["segment_id"].astype(str))
        .set_index("segment_id")["segment_key"]
        .astype(str)
        .to_dict()
    )

    def dominance_rule_matches(row: pd.Series) -> bool:
        """Проверить выполнение правила доминирующего ребёнка.

        Args:
            row: Строка кандидата из диагностики.

        Returns:
            ``True``, если единственный ребёнок объясняет достаточную долю
            движения родителя и направление изменения совпадает.

        Raises:
            ValueError: Не выбрасывается.

        Examples:
            >>> dominance_rule_matches(pd.Series({'hierarchy_best_group_size': 0}))
            False
        """

        best_group_size = _safe_float(
            row.get("hierarchy_best_group_size"),
            math.nan,
        )
        capture = _safe_float(
            row.get("hierarchy_single_child_capture"),
            math.nan,
        )
        direction_value = row.get(
            "hierarchy_single_child_direction_match",
            False,
        )
        # FIXED: После round-trip через Excel boolean читается как ``1.0``.
        direction_match = (
            str(direction_value).strip().lower() in {"true", "1", "yes"}
            or _safe_float(direction_value, 0.0) == 1.0
        )
        return (
            best_group_size == 1.0
            and math.isfinite(capture)
            and capture >= float(thresholds.dominant_child_capture_threshold)
            and direction_match
        )

    def dominant_child_segment(row: pd.Series) -> str:
        """Вернуть ключ ребёнка, объясняющего родительский сегмент.

        Args:
            row: Строка кандидата из диагностики.

        Returns:
            Человекочитаемый ключ доминирующего ребёнка или пустая строка.

        Raises:
            ValueError: Не выбрасывается.

        Examples:
            >>> dominant_child_segment(pd.Series({'hierarchy_best_group_size': 0}))
            ''
        """

        if not dominance_rule_matches(row):
            return ""
        # FIXED: Технические ID читаются из однозначного JSON-массива.
        try:
            group_ids = json.loads(
                str(row.get("hierarchy_best_group_ids_json", "[]"))
            )
        except (TypeError, ValueError, json.JSONDecodeError):
            return ""
        if not isinstance(group_ids, list) or len(group_ids) != 1:
            return ""
        child_id = str(group_ids[0]).strip()
        return str(segment_key_by_id.get(child_id, "")) if child_id else ""

    analysis["hierarchy_dominant_child_segment"] = analysis.apply(
        dominant_child_segment,
        axis=1,
    )
    # ADDED: Отделяем выполнение правила от факта урезания score родителя.
    analysis["hierarchy_dominance_rule_matches"] = analysis.apply(
        dominance_rule_matches,
        axis=1,
    )
    # ADDED: В отчёт выводится только добавка ``lambda * (coherence - 0.5)``.
    best_group_sizes = pd.to_numeric(
        analysis.get("hierarchy_best_group_size", pd.Series(0, index=analysis.index)),
        errors="coerce",
    )
    coherence = pd.to_numeric(
        analysis.get("hierarchy_coherence", pd.Series(math.nan, index=analysis.index)),
        errors="coerce",
    )
    analysis["hierarchy_coherence_adjustment"] = math.nan
    multi_child_mask = best_group_sizes.gt(1) & coherence.notna()
    analysis.loc[multi_child_mask, "hierarchy_coherence_adjustment"] = (
        float(thresholds.aggregation_bonus_lambda)
        * (coherence.loc[multi_child_mask] - 0.5)
    )

    selected_ids = set(final_df["segment_id"].astype(str).tolist()) if not final_df.empty else set()
    manager_ids = (
        set(final_df.head(thresholds.max_manager_facts)["segment_id"].astype(str).tolist())
        if not final_df.empty and "segment_id" in final_df.columns
        else set()
    )
    manager_order_by_id = (
        {
            str(segment_id): order
            for order, segment_id in enumerate(
                final_df.head(thresholds.max_manager_facts)["segment_id"].astype(str).tolist(),
                start=1,
            )
        }
        if not final_df.empty and "segment_id" in final_df.columns
        else {}
    )

    def manager_exclusion_reason(row: pd.Series) -> str:
        """Вернуть причину отсутствия кандидата в менеджерском выводе.

        Args:
            row: Строка кандидата из диагностики.

        Returns:
            Текст причины для листа анализа аномалий.

        Raises:
            ValueError: Не выбрасывается.

        Examples:
            >>> manager_exclusion_reason(pd.Series({'segment_id': 'x'}))
            'не выбран в итоговый набор аномалий'
        """

        segment_id = str(row.get("segment_id", ""))
        if segment_id in manager_ids:
            return "попал в итоговый менеджерский вывод"
        if segment_id in selected_ids:
            return f"выбран в итоговый набор, но не попал в топ-{thresholds.max_manager_facts} менеджерского вывода"
        selection_reason = str(row.get("selection_exclusion_reason", "")).strip()
        if selection_reason:
            return selection_reason
        classifier_reason = str(row.get("reason", "")).strip()
        return classifier_reason if classifier_reason else "не выбран в итоговый набор аномалий"

    analysis["причина не попадания в менеджерский вывод"] = analysis.apply(manager_exclusion_reason, axis=1)
    export = pd.DataFrame(
        {
            "сегмент": analysis["segment_key"].astype(str),
            "глубина": analysis["slice_depth"].astype(int),
            "z_scope": analysis["robust_z"].astype(float).round(2),
            "z_scale_source": analysis.get(
                "z_scale_source", pd.Series("", index=analysis.index)
            ).astype(str),
            "z_uses_sigma_floor": analysis.get(
                "z_uses_sigma_floor", pd.Series(False, index=analysis.index)
            ).astype(bool),
            "base_anomaly_score": pd.to_numeric(
                analysis.get("base_anomaly_score", pd.Series(math.nan, index=analysis.index)),
                errors="coerce",
            ),
            "hierarchy_eligible_descendant_count": pd.to_numeric(
                analysis.get(
                    "hierarchy_eligible_descendant_count",
                    pd.Series(0, index=analysis.index),
                ),
                errors="coerce",
            ).fillna(0).astype(int),
            "hierarchy_group_count": pd.to_numeric(
                analysis.get("hierarchy_group_count", pd.Series(0, index=analysis.index)),
                errors="coerce",
            ).fillna(0).astype(int),
            "hierarchy_best_group_size": pd.to_numeric(
                analysis.get("hierarchy_best_group_size", pd.Series(0, index=analysis.index)),
                errors="coerce",
            ).fillna(0).astype(int),
            "hierarchy_best_group_ids_json": analysis.get(
                "hierarchy_best_group_ids_json", pd.Series("[]", index=analysis.index)
            ).astype(str),
            "hierarchy_best_group_segment_keys": analysis.get(
                "hierarchy_best_group_segment_keys",
                pd.Series("[]", index=analysis.index),
            ).astype(str),
            "hierarchy_best_group_score": pd.to_numeric(
                analysis.get("hierarchy_best_group_score", pd.Series(0.0, index=analysis.index)),
                errors="coerce",
            ),
            "hierarchy_direction_unity": pd.to_numeric(
                analysis.get("hierarchy_direction_unity", pd.Series(math.nan, index=analysis.index)),
                errors="coerce",
            ),
            "hierarchy_dominant_share": pd.to_numeric(
                analysis.get("hierarchy_dominant_share", pd.Series(math.nan, index=analysis.index)),
                errors="coerce",
            ),
            "hierarchy_balance_max": pd.to_numeric(
                analysis.get("hierarchy_balance_max", pd.Series(math.nan, index=analysis.index)),
                errors="coerce",
            ),
            "hierarchy_balance_effective": pd.to_numeric(
                analysis.get("hierarchy_balance_effective", pd.Series(math.nan, index=analysis.index)),
                errors="coerce",
            ),
            "hierarchy_balance": pd.to_numeric(
                analysis.get("hierarchy_balance", pd.Series(math.nan, index=analysis.index)),
                errors="coerce",
            ),
            "hierarchy_coherence": pd.to_numeric(
                analysis.get("hierarchy_coherence", pd.Series(math.nan, index=analysis.index)),
                errors="coerce",
            ),
            "hierarchy_coherence_adjustment": pd.to_numeric(
                analysis.get(
                    "hierarchy_coherence_adjustment",
                    pd.Series(math.nan, index=analysis.index),
                ),
                errors="coerce",
            ),
            "hierarchy_single_child_capture": pd.to_numeric(
                analysis.get(
                    "hierarchy_single_child_capture",
                    pd.Series(math.nan, index=analysis.index),
                ),
                errors="coerce",
            ),
            "hierarchy_single_child_direction_match": analysis.get(
                "hierarchy_single_child_direction_match",
                pd.Series(pd.NA, index=analysis.index, dtype="boolean"),
            ).astype("boolean"),
            "hierarchy_single_child_uncapped_score": pd.to_numeric(
                analysis.get(
                    "hierarchy_single_child_uncapped_score",
                    pd.Series(math.nan, index=analysis.index),
                ),
                errors="coerce",
            ),
            "hierarchy_dominance_cap_score": pd.to_numeric(
                analysis.get(
                    "hierarchy_dominance_cap_score",
                    pd.Series(math.nan, index=analysis.index),
                ),
                errors="coerce",
            ),
            "hierarchy_dominance_rule_matches": analysis[
                "hierarchy_dominance_rule_matches"
            ].astype(bool),
            "hierarchy_dominance_cap_applied": analysis.get(
                "hierarchy_dominance_cap_applied",
                pd.Series(False, index=analysis.index),
            ).astype(bool),
            "hierarchy_dominant_child_segment": analysis[
                "hierarchy_dominant_child_segment"
            ],
            "hierarchy_score_factor": pd.to_numeric(
                analysis.get("hierarchy_score_factor", pd.Series(1.0, index=analysis.index)),
                errors="coerce",
            ),
            "anomaly_score": analysis["anomaly_score"].astype(float),
            "выбран": analysis.get(
                "selected", pd.Series(False, index=analysis.index)
            ).astype(bool),
            "разрешён": analysis.get(
                "is_resolved", pd.Series(False, index=analysis.index)
            ).astype(bool),
            "set_packing_status": analysis.get(
                "set_packing_status", pd.Series("", index=analysis.index)
            ).astype(str),
            "set_packing_reason": analysis.get(
                "set_packing_reason", pd.Series("", index=analysis.index)
            ).astype(str),
            "set_packing_global_status": analysis.get(
                "set_packing_global_status", pd.Series("", index=analysis.index)
            ).astype(str),
            "set_packing_component_id": analysis.get(
                "set_packing_component_id", pd.Series("", index=analysis.index)
            ).astype(str),
            "set_packing_solver": analysis.get(
                "set_packing_solver", pd.Series("", index=analysis.index)
            ).astype(str),
            "set_packing_solver_status": analysis.get(
                "set_packing_solver_status", pd.Series("", index=analysis.index)
            ).astype(str),
            "set_packing_abs_gap": pd.to_numeric(
                analysis.get("set_packing_abs_gap", pd.Series(math.nan, index=analysis.index)),
                errors="coerce",
            ),
            "set_packing_rel_gap": pd.to_numeric(
                analysis.get("set_packing_rel_gap", pd.Series(math.nan, index=analysis.index)),
                errors="coerce",
            ),
            "set_packing_component_score_min": pd.to_numeric(
                analysis.get(
                    "set_packing_component_score_min",
                    pd.Series(math.nan, index=analysis.index),
                ),
                errors="coerce",
            ),
            "set_packing_component_score_max": pd.to_numeric(
                analysis.get(
                    "set_packing_component_score_max",
                    pd.Series(math.nan, index=analysis.index),
                ),
                errors="coerce",
            ),
            "set_packing_component_score_dynamic_range": pd.to_numeric(
                analysis.get(
                    "set_packing_component_score_dynamic_range",
                    pd.Series(math.nan, index=analysis.index),
                ),
                errors="coerce",
            ),
            "conflict_count": pd.to_numeric(
                analysis.get("conflict_count", pd.Series(0, index=analysis.index)),
                errors="coerce",
            ).fillna(0).astype(int),
            "conflict_segment_keys": analysis.get(
                "conflict_segment_keys", pd.Series("", index=analysis.index)
            ).astype(str),
            "selected_atomic_count": pd.to_numeric(
                analysis.get("selected_atomic_count", pd.Series(0, index=analysis.index)),
                errors="coerce",
            ).fillna(0).astype(int),
            "номер добавления в менеджерский вывод": analysis["segment_id"].astype(str).map(manager_order_by_id).astype("Int64"),
            "причина не попадания в менеджерский вывод": analysis["причина не попадания в менеджерский вывод"],
            "Delta GMV": analysis["wow_delta_gmv"].astype(float).round(0).astype("Int64"),
        }
    )
    return export.sort_values(
        by=["глубина", "anomaly_score", "Delta GMV", "сегмент"],
        ascending=[False, False, False, True],
    ).reset_index(drop=True)


def highlight_manager_rows_on_anomaly_analysis(worksheet) -> None:
    """Выделить зелёным строки листа анализа аномалий, попавшие в менеджерский вывод.

    Args:
        worksheet: Лист openpyxl `Анализ аномалий`.

    Returns:
        None.

    Raises:
        AttributeError: Если объект не похож на лист openpyxl.

    Examples:
        >>> # highlight_manager_rows_on_anomaly_analysis(worksheet)
    """

    if worksheet.max_row < 2:
        return

    target_header = "причина не попадания в менеджерский вывод"
    target_value = "попал в итоговый менеджерский вывод"
    green_fill = PatternFill(fill_type="solid", fgColor="C6EFCE")
    header_to_col = {
        str(cell.value): cell.column
        for cell in worksheet[1]
        if cell.value is not None
    }
    reason_col = header_to_col.get(target_header)
    if reason_col is None:
        return

    for row_idx in range(2, worksheet.max_row + 1):
        if worksheet.cell(row=row_idx, column=reason_col).value != target_value:
            continue
        for col_idx in range(1, worksheet.max_column + 1):
            worksheet.cell(row=row_idx, column=col_idx).fill = green_fill


def _order_tree_layers_by_shared_children(
    node_ids_by_depth: Dict[int, List[str]],
    edges: Sequence[Tuple[str, str]],
    records: Dict[str, Dict[str, object]],
    sweep_count: int = 5,
) -> Dict[int, List[str]]:
    """Упорядочить уровни дерева с группировкой родителей общих детей.

    Сначала выполняется top-down barycenter-проход для сближения детей с
    родителями. Затем bottom-up проход объединяет в непрерывные компоненты
    родителей, связанных хотя бы с одним общим ребёнком. Последний проход
    всегда bottom-up, поэтому требование близости общих родителей имеет
    приоритет над косметическим выравниванием детей.

    Args:
        node_ids_by_depth: Узлы, сгруппированные по глубине.
        edges: Рёбра ``(parent_id, child_id)`` соседних глубин.
        records: Атрибуты узлов, включая человекочитаемый сегмент.
        sweep_count: Число двунаправленных проходов минимизации пересечений.

    Returns:
        Упорядоченные идентификаторы узлов для каждой глубины.

    Raises:
        ValueError: Если sweep_count меньше единицы.

    Examples:
        >>> records = {'p1': {'сегмент': 'p1'}, 'p2': {'сегмент': 'p2'}, 'c': {'сегмент': 'c'}}
        >>> _order_tree_layers_by_shared_children({1: ['p2', 'p1'], 2: ['c']}, [('p1', 'c'), ('p2', 'c')], records)[1]
        ['p1', 'p2']
    """

    if sweep_count < 1:
        raise ValueError("sweep_count должен быть не меньше единицы")

    depth_values = sorted(node_ids_by_depth)
    ordered = {
        depth: sorted(
            node_ids,
            key=lambda node_id: str(records[node_id]["сегмент"]),
        )
        for depth, node_ids in node_ids_by_depth.items()
    }
    children_by_parent: Dict[str, List[str]] = {
        node_id: []
        for node_ids in node_ids_by_depth.values()
        for node_id in node_ids
    }
    parents_by_child: Dict[str, List[str]] = {
        node_id: []
        for node_ids in node_ids_by_depth.values()
        for node_id in node_ids
    }
    for parent_id, child_id in edges:
        children_by_parent[parent_id].append(child_id)
        parents_by_child[child_id].append(parent_id)

    def top_down_sweep() -> None:
        """Сблизить детей со средними позициями их родителей.

        Args:
            Аргументы отсутствуют.

        Returns:
            None.

        Raises:
            ValueError: Не выбрасывается.

        Examples:
            >>> # top_down_sweep()
        """

        for depth in depth_values[1:]:
            parent_depth = int(depth) - 1
            if parent_depth not in ordered:
                continue
            parent_positions = {
                node_id: position
                for position, node_id in enumerate(ordered[parent_depth])
            }
            sort_keys: Dict[str, Tuple[float, Tuple[int, ...], str]] = {}
            for node_id in ordered[int(depth)]:
                positions = sorted(
                    parent_positions[parent_id]
                    for parent_id in parents_by_child[node_id]
                    if parent_id in parent_positions
                )
                barycenter = (
                    sum(positions) / len(positions)
                    if positions
                    else float("inf")
                )
                sort_keys[node_id] = (
                    barycenter,
                    tuple(positions),
                    str(records[node_id]["сегмент"]),
                )
            ordered[int(depth)] = sorted(
                ordered[int(depth)],
                key=lambda node_id: sort_keys[node_id],
            )

    def bottom_up_shared_child_sweep() -> None:
        """Сделать компоненты родителей общих детей непрерывными.

        Args:
            Аргументы отсутствуют.

        Returns:
            None.

        Raises:
            ValueError: Не выбрасывается.

        Examples:
            >>> # bottom_up_shared_child_sweep()
        """

        for depth in reversed(depth_values[:-1]):
            child_depth = int(depth) + 1
            if child_depth not in ordered:
                continue
            child_positions = {
                node_id: position
                for position, node_id in enumerate(ordered[child_depth])
            }
            parent_ids = list(ordered[int(depth)])
            parent_id_set = set(parent_ids)
            shared_child_neighbors = {
                parent_id: set()
                for parent_id in parent_ids
            }
            for child_id in ordered[child_depth]:
                linked_parents = [
                    parent_id
                    for parent_id in parents_by_child[child_id]
                    if parent_id in parent_id_set
                ]
                for parent_id in linked_parents:
                    shared_child_neighbors[parent_id].update(
                        other_parent_id
                        for other_parent_id in linked_parents
                        if other_parent_id != parent_id
                    )

            components: List[List[str]] = []
            remaining = set(parent_ids)
            for seed_id in parent_ids:
                if seed_id not in remaining:
                    continue
                stack = [seed_id]
                component: List[str] = []
                remaining.remove(seed_id)
                while stack:
                    parent_id = stack.pop()
                    component.append(parent_id)
                    for neighbor_id in sorted(shared_child_neighbors[parent_id]):
                        if neighbor_id in remaining:
                            remaining.remove(neighbor_id)
                            stack.append(neighbor_id)
                components.append(component)

            parent_sort_keys: Dict[str, Tuple[float, Tuple[int, ...], str]] = {}
            for parent_id in parent_ids:
                positions = sorted(
                    child_positions[child_id]
                    for child_id in children_by_parent[parent_id]
                    if child_id in child_positions
                )
                barycenter = (
                    sum(positions) / len(positions)
                    if positions
                    else float("inf")
                )
                parent_sort_keys[parent_id] = (
                    barycenter,
                    tuple(positions),
                    str(records[parent_id]["сегмент"]),
                )

            component_rows: List[Tuple[Tuple[float, float, str], List[str]]] = []
            for component in components:
                component_order = sorted(
                    component,
                    key=lambda node_id: parent_sort_keys[node_id],
                )
                finite_barycenters = [
                    parent_sort_keys[node_id][0]
                    for node_id in component_order
                    if not math.isinf(parent_sort_keys[node_id][0])
                ]
                component_key = (
                    min(finite_barycenters) if finite_barycenters else float("inf"),
                    (
                        sum(finite_barycenters) / len(finite_barycenters)
                        if finite_barycenters
                        else float("inf")
                    ),
                    min(str(records[node_id]["сегмент"]) for node_id in component_order),
                )
                component_rows.append((component_key, component_order))

            component_rows.sort(key=lambda row: row[0])
            ordered[int(depth)] = [
                node_id
                for _, component in component_rows
                for node_id in component
            ]

    def refine_shared_child_adjacency() -> None:
        """Локально минимизировать разрывы между родителями общего ребёнка.

        Перестановки оцениваются лексикографически: сначала уменьшается число
        разорванных родительских групп, затем суммарный и максимальный разрыв.

        Args:
            Аргументы отсутствуют.

        Returns:
            None.

        Raises:
            ValueError: Не выбрасывается.

        Examples:
            >>> # refine_shared_child_adjacency()
        """

        for depth in depth_values[:-1]:
            child_depth = int(depth) + 1
            if child_depth not in ordered:
                continue
            parent_groups = [
                [
                    parent_id
                    for parent_id in parents_by_child[child_id]
                    if parent_id in ordered[int(depth)]
                ]
                for child_id in ordered[child_depth]
            ]
            parent_groups = [
                parent_ids
                for parent_ids in parent_groups
                if len(parent_ids) > 1
            ]
            if not parent_groups:
                continue

            current_order = list(ordered[int(depth)])

            def gap_objective(candidate_order: Sequence[str]) -> Tuple[int, int, int]:
                """Оценить разрывы родительских групп для одного уровня.

                Args:
                    candidate_order: Проверяемый порядок родителей.

                Returns:
                    Число разорванных групп, сумма разрывов и максимальный разрыв.

                Raises:
                    KeyError: Если родитель отсутствует в порядке.

                Examples:
                    >>> # gap_objective(['p1', 'p2'])
                """

                positions = {
                    node_id: position
                    for position, node_id in enumerate(candidate_order)
                }
                gaps = [
                    max(positions[parent_id] for parent_id in parent_ids)
                    - min(positions[parent_id] for parent_id in parent_ids)
                    + 1
                    - len(parent_ids)
                    for parent_ids in parent_groups
                ]
                return (
                    sum(gap > 0 for gap in gaps),
                    sum(gaps),
                    max(gaps, default=0),
                )

            while True:
                best_objective = gap_objective(current_order)
                best_order: Optional[List[str]] = None
                for left_index in range(len(current_order)):
                    for right_index in range(left_index + 1, len(current_order)):
                        candidate_order = list(current_order)
                        candidate_order[left_index], candidate_order[right_index] = (
                            candidate_order[right_index],
                            candidate_order[left_index],
                        )
                        candidate_objective = gap_objective(candidate_order)
                        if candidate_objective < best_objective:
                            best_objective = candidate_objective
                            best_order = candidate_order
                if best_order is None:
                    break
                current_order = best_order
            ordered[int(depth)] = current_order

    for _ in range(sweep_count):
        top_down_sweep()
        bottom_up_shared_child_sweep()
    bottom_up_shared_child_sweep()
    refine_shared_child_adjacency()
    return ordered


def _tree_edge_port_offset(edge_index: int, edge_count: int, node_width: float) -> float:
    """Рассчитать отдельный горизонтальный порт ребра на границе узла.

    Args:
        edge_index: Индекс ребра среди рёбер узла, начиная с нуля.
        edge_count: Общее число рёбер на соответствующей стороне узла.
        node_width: Ширина узла в координатах графика.

    Returns:
        Смещение порта относительно центра узла.

    Raises:
        ValueError: Если индекс или количество рёбер некорректны.

    Examples:
        >>> round(_tree_edge_port_offset(0, 2, 3.0), 2)
        -0.54
        >>> round(_tree_edge_port_offset(1, 2, 3.0), 2)
        0.54
    """

    if edge_count < 1 or edge_index < 0 or edge_index >= edge_count:
        raise ValueError("Некорректный индекс порта ребра")
    usable_width = node_width * 0.72
    return usable_width * ((edge_index + 0.5) / edge_count - 0.5)


def build_anomaly_tree_from_excel(
    report_path: str | Path,
    output_path: str | Path | None = None,
    sheet_name: str = "Анализ аномалий",
    dpi: int = 160,
) -> Path:
    """Построить дерево аномальных сегментов по листу итогового Excel-файла.

    Узлы группируются по глубине. Ребро проводится только между соседними
    глубинами ``d`` и ``d + 1``, если признаки родителя являются строгим
    подмножеством признаков ребёнка. Один ребёнок может иметь несколько
    родителей, поэтому визуализация фактически является ориентированным DAG.
    Родители общих детей образуют непрерывные группы. Каждому ребру выделяются
    отдельные порты и отдельная трасса; цвет ребра кодирует дочерний сегмент.

    Args:
        report_path: Путь к Excel-отчёту с листом анализа аномалий.
        output_path: Путь к PNG, SVG или PDF. Если None, рядом с отчётом
            создаётся файл ``<имя_отчёта>_tree.png``.
        sheet_name: Имя листа-источника.
        dpi: Разрешение растрового изображения.

    Returns:
        Путь к созданному файлу визуализации.

    Raises:
        FileNotFoundError: Если Excel-отчёт отсутствует.
        ImportError: Если matplotlib недоступен.
        ValueError: Если лист пуст, не содержит обязательных колонок или
            указан неподдерживаемый формат результата.
        OSError: Если изображение невозможно записать.

    Examples:
        >>> # build_anomaly_tree_from_excel('gmv_anomaly_report_1.xlsx', 'gmv_tree.png')
    """

    try:
        import matplotlib

        matplotlib.use("Agg")
        import matplotlib.pyplot as plt
        from matplotlib.lines import Line2D
        from matplotlib.path import Path as MatplotlibPath
        from matplotlib.patches import FancyArrowPatch, FancyBboxPatch, Patch, Rectangle
    except ImportError as exc:
        raise ImportError("Для построения дерева требуется matplotlib") from exc

    report = Path(report_path)
    if not report.exists():
        raise FileNotFoundError(f"Excel-отчёт не найден: {report}")

    tree_path = (
        Path(output_path)
        if output_path is not None
        else report.with_name(f"{report.stem}_tree.png")
    )
    if not tree_path.suffix:
        tree_path = tree_path.with_suffix(".png")
    if tree_path.suffix.lower() not in {".png", ".svg", ".pdf"}:
        raise ValueError("Дерево можно сохранить только в PNG, SVG или PDF")
    if dpi <= 0:
        raise ValueError("dpi должен быть положительным целым числом")

    analysis = pd.read_excel(report, sheet_name=sheet_name)
    required_columns = {"сегмент", "глубина", "z_scope", "anomaly_score", "Delta GMV"}
    missing_columns = sorted(required_columns - set(analysis.columns))
    if missing_columns:
        raise ValueError(
            f"На листе {sheet_name!r} отсутствуют обязательные колонки: {missing_columns}"
        )

    nodes = analysis.copy()
    # FIXED: Граф больше не использует диагностические причины set packing.
    optional_tree_columns = [
        # FIXED: Поле ниже используется при нормализации, но отсутствует в
        # минимальном контракте листа «Анализ аномалий».
        "conflict_count",
        "hierarchy_best_group_size",
        "hierarchy_best_group_segment_keys",
        "hierarchy_coherence_adjustment",
        "hierarchy_dominance_rule_matches",
        "hierarchy_dominance_cap_applied",
        "hierarchy_dominant_child_segment",
        "номер добавления в менеджерский вывод",
    ]
    for column in optional_tree_columns:
        if column not in nodes.columns:
            nodes[column] = ""
    nodes["сегмент"] = nodes["сегмент"].astype(str).str.strip()
    nodes["глубина"] = pd.to_numeric(nodes["глубина"], errors="coerce")
    nodes["z_scope"] = pd.to_numeric(nodes["z_scope"], errors="coerce")
    nodes["anomaly_score"] = pd.to_numeric(nodes["anomaly_score"], errors="coerce")
    nodes["Delta GMV"] = pd.to_numeric(nodes["Delta GMV"], errors="coerce")
    nodes["номер добавления в менеджерский вывод"] = pd.to_numeric(
        nodes["номер добавления в менеджерский вывод"],
        errors="coerce",
    )
    nodes["conflict_count"] = pd.to_numeric(nodes["conflict_count"], errors="coerce").fillna(0).astype(int)
    nodes = nodes[
        nodes["глубина"].notna()
        & nodes["глубина"].gt(0)
        & nodes["сегмент"].ne("")
        & nodes["сегмент"].ne("nan")
    ].copy()
    if nodes.empty:
        raise ValueError(f"Лист {sheet_name!r} не содержит сегментов глубины выше нуля")

    nodes["глубина"] = nodes["глубина"].astype(int)
    nodes = nodes.drop_duplicates(subset=["глубина", "сегмент"], keep="first")
    nodes = nodes.sort_values(["глубина", "сегмент"], kind="stable").reset_index(drop=True)
    nodes["node_id"] = [
        f"depth={int(depth)}|segment={segment}"
        for depth, segment in zip(nodes["глубина"], nodes["сегмент"])
    ]
    nodes["parts"] = nodes["сегмент"].map(parse_segment_key_parts)
    nodes["feature_set"] = nodes["parts"].map(frozenset)

    def clean_tree_value(value: object) -> str:
        """Вернуть безопасный текст для подписи узла дерева.

        Args:
            value: Исходное значение из Excel.

        Returns:
            Строка без `nan` и лишних пробелов.

        Raises:
            ValueError: Не выбрасывается.

        Examples:
            >>> clean_tree_value(float('nan'))
            ''
        """

        if value is None or pd.isna(value):
            return ""
        text = str(value).strip()
        return "" if text.lower() in {"", "nan", "<na>", "none"} else text

    def build_tree_detail_lines(row: pd.Series) -> List[str]:
        """Собрать детали иерархической корректировки score для узла.

        Args:
            row: Строка узла из листа анализа аномалий.

        Returns:
            Строки с объясняющими детьми либо пустой список.

        Raises:
            ValueError: Не выбрасывается.

        Examples:
            >>> build_tree_detail_lines(pd.Series({'hierarchy_best_group_size': 0}))
            []
        """

        rule_matches = clean_tree_value(
            row.get("hierarchy_dominance_rule_matches")
        )
        # FIXED: Excel читает boolean-колонки как ``1.0``/``0.0``.
        rule_matches_flag = (
            rule_matches.lower() in {"true", "1", "yes"}
            or _safe_float(rule_matches, 0.0) == 1.0
        )
        if rule_matches_flag:
            child_segment = clean_tree_value(
                row.get("hierarchy_dominant_child_segment")
            )
            child_suffix = _dominant_child_slice_suffix(
                row.get("сегмент", ""),
                child_segment,
            )
            cap_applied = clean_tree_value(
                row.get("hierarchy_dominance_cap_applied")
            )
            cap_applied_flag = (
                cap_applied.lower() in {"true", "1", "yes"}
                or _safe_float(cap_applied, 0.0) == 1.0
            )
            label = (
                "Доминирующий ребенок cut:"
                if cap_applied_flag
                else "Доминирующий ребенок:"
            )
            return [label, *_wrap_tree_detail_line(child_suffix)] if child_suffix else []

        adjustment = _safe_float(
            row.get("hierarchy_coherence_adjustment"),
            math.nan,
        )
        group_segments_text = clean_tree_value(
            row.get("hierarchy_best_group_segment_keys")
        )
        try:
            group_segments = json.loads(group_segments_text)
        except (TypeError, ValueError, json.JSONDecodeError):
            group_segments = []
        if not math.isfinite(adjustment) or not isinstance(group_segments, list):
            return []

        lines = [f"Δ score factor: {adjustment:+.3f}"]
        for child_segment in group_segments:
            child_suffix = _dominant_child_slice_suffix(
                row.get("сегмент", ""),
                child_segment,
            )
            if child_suffix:
                lines.extend(_wrap_tree_detail_line(child_suffix))
        return lines if len(lines) > 1 else []

    nodes["is_manager_output"] = nodes["номер добавления в менеджерский вывод"].notna()
    nodes["tree_detail_lines"] = nodes.apply(build_tree_detail_lines, axis=1)
    # ADDED: Перечёркиваем ровно те карточки, в которых показан статус
    # доминирующего ребёнка, включая вариант с применённым cap.
    dominant_child_status_labels = {
        "Доминирующий ребенок:",
        "Доминирующий ребенок cut:",
    }
    nodes["has_dominant_child"] = nodes["tree_detail_lines"].map(
        lambda lines: bool(lines)
        and str(lines[0]).strip() in dominant_child_status_labels
    )

    records = {
        str(row["node_id"]): row.to_dict()
        for _, row in nodes.iterrows()
    }
    depth_values = sorted(nodes["глубина"].unique().tolist())
    node_ids_by_depth = {
        int(depth): nodes.loc[nodes["глубина"].eq(depth), "node_id"].astype(str).tolist()
        for depth in depth_values
    }

    edges: List[Tuple[str, str]] = []
    children_by_parent: Dict[str, List[str]] = {
        node_id: [] for node_id in records
    }
    parents_by_child: Dict[str, List[str]] = {
        node_id: [] for node_id in records
    }
    for depth in depth_values:
        parent_ids = node_ids_by_depth[int(depth)]
        child_ids = node_ids_by_depth.get(int(depth) + 1, [])
        for parent_id in parent_ids:
            parent_features = records[parent_id]["feature_set"]
            for child_id in child_ids:
                if parent_features < records[child_id]["feature_set"]:
                    edges.append((parent_id, child_id))
                    children_by_parent[parent_id].append(child_id)
                    parents_by_child[child_id].append(parent_id)

    ordered_by_depth = _order_tree_layers_by_shared_children(
        node_ids_by_depth,
        edges,
        records,
    )

    node_width = 3.15
    horizontal_gap = 0.22
    cluster_padding_x = 0.28
    cluster_padding_y = 0.24
    max_nodes_in_layer = max(len(node_ids) for node_ids in ordered_by_depth.values())
    total_width = max_nodes_in_layer * node_width + max(0, max_nodes_in_layer - 1) * horizontal_gap
    node_heights = {
        depth: 1.24
        + 0.24
        * max(
            len(records[node_id]["parts"])
            for node_id in node_ids
        )
        + 0.18
        * max(
            len(records[node_id].get("tree_detail_lines", []))
            for node_id in node_ids
        )
        for depth, node_ids in ordered_by_depth.items()
    }
    edges_by_parent_depth = {
        int(depth): [
            edge
            for edge in edges
            if int(records[edge[0]]["глубина"]) == int(depth)
        ]
        for depth in depth_values
    }
    gap_after_depth = {
        int(depth): max(
            1.05,
            min(3.0, 0.07 * len(edges_by_parent_depth[int(depth)]) + 0.72),
        )
        for depth in depth_values
    }
    total_height = sum(
        node_heights[depth] + 2 * cluster_padding_y
        for depth in depth_values
    ) + sum(
        gap_after_depth[int(depth)]
        for depth in depth_values[:-1]
    )

    positions: Dict[str, Tuple[float, float]] = {}
    layer_bounds: Dict[int, Tuple[float, float, float, float]] = {}
    current_top = total_height
    for depth_index, depth in enumerate(depth_values):
        node_ids = ordered_by_depth[int(depth)]
        node_height = node_heights[int(depth)]
        cluster_height = node_height + 2 * cluster_padding_y
        center_y = current_top - cluster_height / 2
        row_width = len(node_ids) * node_width + max(0, len(node_ids) - 1) * horizontal_gap
        row_left = (total_width - row_width) / 2
        for position, node_id in enumerate(node_ids):
            center_x = row_left + node_width / 2 + position * (node_width + horizontal_gap)
            positions[node_id] = (center_x, center_y)
        layer_bounds[int(depth)] = (
            row_left - cluster_padding_x,
            center_y - cluster_height / 2,
            row_width + 2 * cluster_padding_x,
            cluster_height,
        )
        current_top -= cluster_height
        if depth_index < len(depth_values) - 1:
            current_top -= gap_after_depth[int(depth)]

    figure_width = max(14.0, min(42.0, total_width * 0.78))
    figure_height = max(8.0, total_height * 1.15 + 2.0)
    figure, axis = plt.subplots(figsize=(figure_width, figure_height))
    axis.set_xlim(-0.55, total_width + 0.55)
    axis.set_ylim(-1.15, total_height + 1.15)
    axis.axis("off")

    for depth in depth_values:
        left, bottom, width, height = layer_bounds[int(depth)]
        axis.add_patch(
            FancyBboxPatch(
                (left, bottom),
                width,
                height,
                boxstyle="round,pad=0.04,rounding_size=0.22",
                linewidth=0.9,
                edgecolor="#c7daf7",
                facecolor="#f8fbff",
                zorder=0,
            )
        )
        axis.text(
            left + 0.14,
            bottom + height - 0.12,
            f"Сегменты глубины {int(depth)}",
            color="#315a9f",
            fontsize=9,
            ha="left",
            va="top",
            zorder=4,
        )

    sorted_children_by_parent = {
        parent_id: sorted(
            child_ids,
            key=lambda child_id: (positions[child_id][0], child_id),
        )
        for parent_id, child_ids in children_by_parent.items()
    }
    sorted_parents_by_child = {
        child_id: sorted(
            parent_ids,
            key=lambda parent_id: (positions[parent_id][0], parent_id),
        )
        for child_id, parent_ids in parents_by_child.items()
    }
    edge_colors = [
        "#2563eb",
        "#7c3aed",
        "#0891b2",
        "#db2777",
        "#4f46e5",
        "#0f766e",
        "#c2410c",
        "#0369a1",
    ]
    ordered_child_ids = sorted(
        {child_id for _, child_id in edges},
        key=lambda child_id: (
            int(records[child_id]["глубина"]),
            positions[child_id][0],
            child_id,
        ),
    )
    edge_color_by_child = {
        child_id: edge_colors[index % len(edge_colors)]
        for index, child_id in enumerate(ordered_child_ids)
    }

    for parent_depth in depth_values[:-1]:
        layer_edges = sorted(
            edges_by_parent_depth[int(parent_depth)],
            key=lambda edge: (
                positions[edge[0]][0],
                positions[edge[1]][0],
                edge[0],
                edge[1],
            ),
        )
        if not layer_edges:
            continue
        lane_count = len(layer_edges)
        for lane_index, (parent_id, child_id) in enumerate(layer_edges):
            parent_x, parent_y = positions[parent_id]
            child_x, child_y = positions[child_id]
            outgoing_ids = sorted_children_by_parent[parent_id]
            incoming_ids = sorted_parents_by_child[child_id]
            start_x = parent_x + _tree_edge_port_offset(
                outgoing_ids.index(child_id),
                len(outgoing_ids),
                node_width,
            )
            end_x = child_x + _tree_edge_port_offset(
                incoming_ids.index(parent_id),
                len(incoming_ids),
                node_width,
            )
            start_y = parent_y - node_heights[int(records[parent_id]["глубина"])] / 2
            end_y = child_y + node_heights[int(records[child_id]["глубина"])] / 2
            lane_top = start_y - 0.13
            lane_bottom = end_y + 0.13
            lane_y = lane_top - (lane_top - lane_bottom) * (lane_index + 1) / (lane_count + 1)
            path = MatplotlibPath(
                [
                    (start_x, start_y),
                    (start_x, lane_y),
                    (end_x, lane_y),
                    (end_x, end_y),
                ],
                [
                    MatplotlibPath.MOVETO,
                    MatplotlibPath.LINETO,
                    MatplotlibPath.LINETO,
                    MatplotlibPath.LINETO,
                ],
            )
            axis.add_patch(
                FancyArrowPatch(
                    path=path,
                    arrowstyle="-",
                    linewidth=2.2,
                    color="white",
                    alpha=0.96,
                    zorder=1,
                )
            )
            axis.add_patch(
                FancyArrowPatch(
                    path=path,
                    arrowstyle="-|>",
                    mutation_scale=6.5,
                    linewidth=0.82,
                    color=edge_color_by_child[child_id],
                    alpha=0.88,
                    zorder=1.1,
                )
            )

    connected_node_ids = {
        node_id
        for edge in edges
        for node_id in edge
    }
    sequence = 1
    for depth in depth_values:
        for node_id in ordered_by_depth[int(depth)]:
            node = records[node_id]
            center_x, center_y = positions[node_id]
            node_height = node_heights[int(depth)]
            delta_gmv = _safe_float(node["Delta GMV"])
            positive = delta_gmv >= 0
            facecolor = "#eff9f1" if positive else "#fff0f0"
            edgecolor = "#afd8b8" if positive else "#f0b2b2"
            accent = "#49a15d" if positive else "#df4d4d"
            isolated = node_id not in connected_node_ids
            if isolated:
                edgecolor = "#8b9bb2"
            is_manager_output = bool(node.get("is_manager_output", False))
            node_edgecolor = "#172554" if is_manager_output else edgecolor
            node_linewidth = 2.85 if is_manager_output else 1.0

            left = center_x - node_width / 2
            bottom = center_y - node_height / 2
            axis.add_patch(
                Rectangle(
                    (left, bottom),
                    node_width,
                    node_height,
                    linewidth=node_linewidth,
                    linestyle="-" if is_manager_output else ("--" if isolated else "-"),
                    edgecolor=node_edgecolor,
                    facecolor=facecolor,
                    zorder=2,
                )
            )
            badge_width = 0.52
            badge_height = 0.27
            badge_left = left + 0.13
            badge_bottom = bottom + node_height - 0.39
            axis.add_patch(
                Rectangle(
                    (badge_left, badge_bottom),
                    badge_width,
                    badge_height,
                    linewidth=0,
                    facecolor=accent,
                    zorder=3,
                )
            )
            axis.text(
                badge_left + badge_width / 2,
                badge_bottom + badge_height / 2,
                str(sequence),
                color="white",
                fontsize=6.5,
                fontweight="bold",
                ha="center",
                va="center",
                zorder=4,
            )
            axis.text(
                left + 0.83,
                badge_bottom + badge_height / 2,
                f"Глубина {int(depth)}",
                color="#263342",
                fontsize=6.2,
                fontweight="bold",
                ha="left",
                va="center",
                zorder=4,
            )

            feature_lines = [
                f"{dimension}={value}"
                for dimension, value in node["parts"]
            ]
            body_top = badge_bottom - 0.11
            axis.text(
                left + 0.14,
                body_top,
                "\n".join(feature_lines),
                color="#222222",
                fontsize=5.8,
                linespacing=1.35,
                ha="left",
                va="top",
                zorder=4,
            )
            detail_lines = [
                str(line)
                for line in node.get("tree_detail_lines", [])
                if str(line).strip()
            ]
            if detail_lines:
                detail_top = body_top - 0.215 * max(1, len(feature_lines)) - 0.07
                axis.text(
                    left + 0.14,
                    detail_top,
                    "\n".join(detail_lines),
                    color="#334155",
                    fontsize=4.75,
                    linespacing=1.22,
                    ha="left",
                    va="top",
                    zorder=4,
                )
            delta_y = bottom + 0.35
            axis.text(
                left + 0.14,
                delta_y,
                f"ΔGMV: {'+' if delta_gmv > 0 else ''}{_format_rub(delta_gmv)}",
                color="#171717",
                fontsize=6.3,
                fontweight="bold",
                ha="left",
                va="bottom",
                zorder=4,
            )
            z_value = _safe_float(node["z_scope"], math.nan)
            score_value = _safe_float(node["anomaly_score"], math.nan)
            z_text = "nan" if math.isnan(z_value) else f"{z_value:.2f}"
            score_text = "nan" if math.isnan(score_value) else f"{score_value:.3f}"
            axis.text(
                left + 0.14,
                bottom + 0.12,
                f"z={z_text} · score={score_text}",
                color="#7f8b99",
                fontsize=5.3,
                ha="left",
                va="bottom",
                zorder=4,
            )
            if bool(node.get("has_dominant_child", False)):
                # ADDED: Полупрозрачный крест поверх всей карточки визуально
                # помечает сегмент, аномалию которого объясняет один ребёнок.
                cross_inset = 0.05
                cross_style = {
                    "color": "#5f6764",
                    "linewidth": 7.0,
                    "alpha": 0.42,
                    "solid_capstyle": "round",
                    "zorder": 4.5,
                }
                axis.add_line(
                    Line2D(
                        [left + cross_inset, left + node_width - cross_inset],
                        [bottom + cross_inset, bottom + node_height - cross_inset],
                        **cross_style,
                    )
                )
                axis.add_line(
                    Line2D(
                        [left + cross_inset, left + node_width - cross_inset],
                        [bottom + node_height - cross_inset, bottom + cross_inset],
                        **cross_style,
                    )
                )
            sequence += 1

    figure.suptitle(
        "Граф",
        fontsize=15,
        y=0.985,
    )
    legend_handles = [
        Patch(facecolor="#eff9f1", edgecolor="#afd8b8", label="Положительный ΔGMV"),
        Patch(facecolor="#fff0f0", edgecolor="#f0b2b2", label="Отрицательный ΔGMV"),
        Patch(facecolor="white", edgecolor="#172554", linewidth=2.4, label="Попал в менеджерский вывод"),
        Patch(facecolor="white", edgecolor="#8b9bb2", linestyle="--", label="Изолированный сегмент"),
        Line2D([0], [0], color="#2563eb", linewidth=1.4, label="Одинаковый цвет рёбер — один ребёнок"),
    ]
    figure.legend(
        handles=legend_handles,
        loc="lower left",
        bbox_to_anchor=(0.015, 0.01),
        frameon=False,
        ncol=5,
        fontsize=8,
    )
    tree_path.parent.mkdir(parents=True, exist_ok=True)
    try:
        figure.savefig(tree_path, dpi=dpi, bbox_inches="tight", facecolor="white")
    finally:
        plt.close(figure)
    return tree_path.resolve()


def write_anomaly_excel(
    output_path: str | Path,
    thresholds: AnomalyThresholds,
    dim_cols: Sequence[str],
    history_df: pd.DataFrame,
    panel_df: pd.DataFrame,
    candidates: pd.DataFrame,
    final_df: pd.DataFrame,
    manager_df: pd.DataFrame,
    total_by_date: pd.Series,
    dates: Sequence[int],
    current_cal_date: int,
    coverage: Dict[str, frozenset[str]],
    optimization_decision_log: pd.DataFrame,
) -> None:
    """Записать результат поиска аномалий в Excel.

    Args:
        output_path: Путь к итоговому Excel-файлу.
        thresholds: Пороги алгоритма.
        dim_cols: Список признаков.
        history_df: Исходная очищенная таблица.
        panel_df: Полная недельная панель.
        candidates: Диагностика кандидатов.
        final_df: Итоговые выбранные аномалии.
        manager_df: Готовая таблица вкладки `01_Менеджерский_вывод`.
        total_by_date: Total GMV по неделям.
        dates: Список недель.
        current_cal_date: Текущая неделя.
        coverage: Покрытие кандидатов атомами.
        optimization_decision_log: Журнал построения и решения Set Packing.

    Returns:
        None.

    Raises:
        OSError: Если Excel-файл невозможно записать.

    Examples:
        >>> # write_anomaly_excel('gmv_anomaly_report.xlsx', thresholds, dims, history_df, panel, candidates, final_df, manager_df, total, dates, dates[-1], coverage, log)
    """

    output_path = Path(output_path)
    params = pd.DataFrame(
        [
            ("признаки", " × ".join(dim_cols)),
            ("min_anomaly_abs", thresholds.min_anomaly_abs),
            ("min_z_score", thresholds.min_z_score),
            ("min_materiality_share", thresholds.min_materiality_share),
            ("sigma_floor", thresholds.sigma_floor),
            ("lifecycle_z_score", thresholds.lifecycle_z_score),
            (
                "hierarchy_reconciliation_abs_tolerance",
                thresholds.hierarchy_reconciliation_abs_tolerance,
            ),
            ("aggregation_bonus_lambda", thresholds.aggregation_bonus_lambda),
            ("single_child_factor", thresholds.single_child_factor),
            (
                "dominant_child_capture_threshold",
                thresholds.dominant_child_capture_threshold,
            ),
            (
                "dominant_child_score_margin",
                thresholds.dominant_child_score_margin,
            ),
            ("set_packing_gap_tolerance", thresholds.set_packing_gap_tolerance),
            ("max_exact_fallback_size", thresholds.max_exact_fallback_size),
            ("current_cal_date", int(current_cal_date)),
        ],
        columns=["показатель", "значение"],
    )
    control = build_control_table(history_df, panel_df, candidates, final_df, coverage, dates, current_cal_date, total_by_date)
    params.insert(0, "раздел", "Параметры")
    control.insert(0, "раздел", "Контроль")
    params_and_control = pd.concat([params, control], ignore_index=True)

    anomaly_analysis = build_anomaly_analysis_sheet(candidates, final_df, thresholds)
    history_top = build_history_for_selected(panel_df, final_df, total_by_date)
    missing_zero = build_missing_zero_report(panel_df, dates, current_cal_date)

    final_cols = [
        "rank",
        "output_block",
        "segment_key",
        "slice_depth",
        "gmv_current",
        "gmv_previous",
        "wow_delta_gmv",
        "relative_wow",
        "share_current",
        "share_delta_current",
        "baseline_relative_growth",
        "robust_z",
        "abs_robust_z",
        "z_scale_source",
        "z_uses_sigma_floor",
        "materiality_share",
        "gross_atomic_movement",
        "base_anomaly_score",
        "hierarchy_eligible_descendant_count",
        "hierarchy_group_count",
        "hierarchy_best_group_size",
        "hierarchy_best_group_ids_json",
        "hierarchy_best_group_segment_keys",
        "hierarchy_best_group_score",
        "hierarchy_direction_unity",
        "hierarchy_dominant_share",
        "hierarchy_balance_max",
        "hierarchy_balance_effective",
        "hierarchy_balance",
        "hierarchy_coherence",
        "hierarchy_coherence_adjustment",
        "hierarchy_single_child_capture",
        "hierarchy_single_child_direction_match",
        "hierarchy_single_child_uncapped_score",
        "hierarchy_dominance_cap_score",
        "hierarchy_dominance_cap_applied",
        "hierarchy_score_factor",
        "anomaly_score",
        "selection_score",
        "reliability_factor",
        "history_nonzero_weeks",
        "state",
        "reason",
        "covered_atomic_count",
        "selected",
        "set_packing_global_status",
        "set_packing_component_id",
        "set_packing_solver",
        "set_packing_solver_status",
        "set_packing_reason",
        "set_packing_objective_value",
        "set_packing_best_bound",
        "set_packing_abs_gap",
        "set_packing_rel_gap",
        "set_packing_solve_time_sec",
        "set_packing_variable_count",
        "set_packing_constraint_count",
        "set_packing_component_score_min",
        "set_packing_component_score_max",
        "set_packing_component_score_dynamic_range",
        "conflict_count",
        "conflict_segment_ids",
        "conflict_segment_keys",
        "atomic_coverage_source",
        "atomic_coverage_validation_status",
        "is_resolved",
        "original_atomic_count",
        "selected_atomic_count",
        "original_atomic_descendants",
        "selected_atomic_descendants",
    ]
    final_export = final_df[[col for col in final_cols if col in final_df.columns]].copy()

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        params_and_control.to_excel(writer, sheet_name="00_Параметры_и_контроль", index=False)
        manager_df.to_excel(writer, sheet_name="01_Менеджерский_вывод", index=False)
        final_export.to_excel(writer, sheet_name="02_Итог_аномалий", index=False)
        anomaly_analysis.to_excel(writer, sheet_name="Анализ аномалий", index=False)
        history_top.to_excel(writer, sheet_name="03_История_top", index=False)
        candidates.to_excel(writer, sheet_name="04_Диагностика_кандидатов", index=False)
        missing_zero.to_excel(writer, sheet_name="05_Пропуски_и_нули", index=False)
        control.to_excel(writer, sheet_name="06_Контроль", index=False)
        optimization_decision_log.to_excel(writer, sheet_name="07_Журнал_set_packing", index=False)

        highlight_manager_rows_on_anomaly_analysis(writer.sheets["Анализ аномалий"])

        for worksheet in writer.sheets.values():
            worksheet.freeze_panes = "A2"
            worksheet.auto_filter.ref = worksheet.dimensions
            for col_cells in worksheet.columns:
                max_length = 0
                col_letter = col_cells[0].column_letter
                for cell in col_cells[:200]:
                    value = "" if cell.value is None else str(cell.value)
                    max_length = max(max_length, len(value))
                worksheet.column_dimensions[col_letter].width = min(max(max_length + 2, 10), 48)
