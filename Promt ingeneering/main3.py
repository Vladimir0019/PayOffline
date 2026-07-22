"""
Поиск необычных сегментов GMV на истории нескольких недель.

Модуль считает статистическую необычность отдельно для каждого среза, а затем
выбирает компактный набор непересекающихся управленческих сегментов.
"""



# Исправить, доработать в будущем:
# 2) отсутствующие недели заменяются на 0, добавляется флаг на пропуск. Но это может влиять на расчет медианы изменения. Рассмотреть в будущем

from __future__ import annotations

import argparse
from itertools import combinations
import math
import re
import time
from dataclasses import dataclass
from pathlib import Path
from typing import Dict, List, Optional, Sequence, Tuple

import pandas as pd
from openpyxl.styles import PatternFill

from main2 import (
    build_segment_key_and_level,
    candidate_covers_atomic,
    normalize_dim_value,
    segment_id_from_row,
)


DEFAULT_INPUT_PATH = Path(__file__).with_name("payoffline_pulse_hier_4_13w_2207.xlsx")
DEFAULT_OUTPUT_PATH = Path(__file__).with_name("gmv_anomaly_report_2_2207.xlsx")
DEFAULT_TREE_OUTPUT_PATH = Path(__file__).with_name("Граф_2207.png")

# ADDED: Для исторического anomaly-файла исключаем все технические и метрические колонки.
ANOMALY_TECH_COLUMNS = {
    "period",
    "cal_date",
    "calendar_date",
    "slice_depth",
    "gmv",
    "tx",
    "au",
    "am",
    "aov",
    "tpm",
    "freq",
    "share_in_total_gmv",
    "segment_id",
    "segment_key",
    "segment_level",
}

# ADDED: Операционные метрики, для которых в менеджерском выводе нужен WoW-процент.
METRIC_COLUMNS = ["tx", "au", "am", "aov", "tpm", "freq"]
MANAGER_METRIC_PCT_COLUMNS = {
    "relative_wow": "GMV WoW %",
    "tx_wow_pct": "TX WoW %",
    "au_wow_pct": "AU WoW %",
    "am_wow_pct": "AM WoW %",
    "aov_wow_pct": "AOV WoW %",
    "tpm_wow_pct": "TPM WoW %",
    "freq_wow_pct": "Freq WoW %",
}


@dataclass(frozen=True)
class AnomalyThresholds:
    """Пороги алгоритма поиска необычных сегментов.

    Args:
        min_anomaly_abs: Минимальная абсолютная величина аномального вклада в рублях.
        min_z_score: Минимальный robust z-score для обычного сегмента.
        min_materiality_share: Минимальная доля изменения GMV в gross movement атомарного слоя.
        sigma_floor: Нижняя граница масштаба колебаний доли.
        z_cap: Верхняя граница z-score, участвующего в score.
        set_packing_gap_tolerance: Максимальный относительный MIP gap для признания решения оптимальным.
        max_exact_fallback_size: Максимальный размер компоненты для собственного exact fallback.
        max_manager_facts: Максимальное число фактов в менеджерском выводе.

    Returns:
        Экземпляр с порогами.

    Raises:
        ValueError: Не выбрасывается.

    Examples:
        >>> AnomalyThresholds().min_z_score
        2.0
    """

    min_anomaly_abs: float = 200_000.0
    min_z_score: float = 2.0
    min_materiality_share: float = 0.00001
    sigma_floor: float = 0.00001
    z_cap: float = 6.0
    set_packing_gap_tolerance: float = 1e-9
    max_exact_fallback_size: int = 25
    max_manager_facts: int = 10


def infer_anomaly_dimension_columns(df: pd.DataFrame, explicit_dims: Optional[Sequence[str]] = None) -> List[str]:
    """Определить бизнес-признаки для anomaly-анализа.

    Args:
        df: Входная таблица.
        explicit_dims: Явно заданные признаки. Если переданы, используются они.

    Returns:
        Список бизнес-признаков.

    Raises:
        ValueError: Если явно заданный признак отсутствует.

    Examples:
        >>> infer_anomaly_dimension_columns(pd.DataFrame(columns=['period', 'gmv', 'geo', 'tx']))
        ['geo']
    """

    if explicit_dims:
        missing = [col for col in explicit_dims if col not in df.columns]
        if missing:
            raise ValueError(f"В таблице отсутствуют явно заданные признаки: {missing}")
        return list(explicit_dims)
    return [col for col in df.columns if col not in ANOMALY_TECH_COLUMNS]


def _format_rub(value: float) -> str:
    """Отформатировать рублёвое значение без дробной части.

    Args:
        value: Числовое значение.

    Returns:
        Строка с разделителем групп разрядов.

    Raises:
        ValueError: Не выбрасывается.

    Examples:
        >>> _format_rub(1234567.89)
        '1 234 568'
    """

    return f"{value:,.0f}".replace(",", " ")


def _format_pct(value: object) -> object:
    """Отформатировать относительное изменение как процент для менеджерского вывода.

    Args:
        value: Числовая доля, например 0.15 для 15%.

    Returns:
        Строка с процентом или пустая строка, если процент нельзя корректно посчитать.

    Raises:
        ValueError: Не выбрасывается.

    Examples:
        >>> _format_pct(0.1234)
        '12.34%'
    """

    if value is None or pd.isna(value):
        return ""
    numeric_value = float(value)
    if math.isnan(numeric_value) or math.isinf(numeric_value):
        return ""
    return f"{numeric_value * 100:.2f}%"


def _safe_float(value: object, default: float = 0.0) -> float:
    """Преобразовать значение в float с безопасным fallback.

    Args:
        value: Исходное значение.
        default: Значение, возвращаемое при ошибке.

    Returns:
        Число float.

    Raises:
        ValueError: Не выбрасывается.

    Examples:
        >>> _safe_float("10")
        10.0
    """

    try:
        result = float(value)
    except (TypeError, ValueError):
        return default
    if math.isnan(result) or math.isinf(result):
        return default
    return result


def _segment_key_parts(segment_key: object) -> List[Tuple[str, str]]:
    """Разобрать ключ сегмента на пары «признак — значение».

    Args:
        segment_key: Человекочитаемый ключ сегмента.

    Returns:
        Список пар `(dimension, value)` в порядке исходного ключа.

    Raises:
        ValueError: Не выбрасывается.

    Examples:
        >>> _segment_key_parts("products=FULLPAYMENT × merchants_type=SMB")
        [('products', 'FULLPAYMENT'), ('merchants_type', 'SMB')]
    """

    parts: List[Tuple[str, str]] = []
    for raw_part in re.split(r"\s+(?:x|\u00d7|\u0413\u2014)\s+", str(segment_key)):
        part = raw_part.strip()
        if "=" not in part:
            continue
        dimension, value = part.split("=", 1)
        parts.append((dimension.strip(), value.strip()))
    return parts


def load_history_table(
    input_path: str | Path,
    sheet_name: int | str = 0,
    period: Optional[str] = "1W",
    dim_cols: Optional[Sequence[str]] = None,
) -> Tuple[pd.DataFrame, List[str], List[int]]:
    """Загрузить историческую таблицу срезов GMV.

    Args:
        input_path: Путь к Excel- или CSV-файлу.
        sheet_name: Имя или номер листа Excel.
        period: Значение периода для фильтрации. Если None, фильтр не применяется.
        dim_cols: Явно заданные признаки. Если None, признаки определяются автоматически.

    Returns:
        Кортеж: очищенная таблица, список признаков, список недель cal_date из total-слоя.

    Raises:
        ValueError: Если входной файл не поддерживается, нет обязательных колонок или total-слой некорректен.

    Examples:
        >>> # df, dims, dates = load_history_table('payoffline_pulse_hier_4_13w.xlsx')
    """

    input_path = Path(input_path)
    if input_path.suffix.lower() in {".xlsx", ".xls"}:
        df = pd.read_excel(input_path, sheet_name=sheet_name)
    elif input_path.suffix.lower() == ".csv":
        df = pd.read_csv(input_path)
    else:
        raise ValueError("Поддерживаются только .xlsx, .xls и .csv")

    required = {"cal_date", "slice_depth", "gmv"}
    missing_required = sorted(required - set(df.columns))
    if missing_required:
        raise ValueError(f"Не хватает обязательных колонок: {missing_required}")

    # ADDED: Удаляем служебную строку типов из Excel-выгрузки, если она присутствует.
    if "period" in df.columns:
        df = df[df["period"].astype(str) != "string"].copy()
        if period is not None:
            df = df[df["period"].astype(str) == str(period)].copy()

    df["cal_date"] = pd.to_numeric(df["cal_date"], errors="coerce")
    df["slice_depth"] = pd.to_numeric(df["slice_depth"], errors="coerce")
    df["gmv"] = pd.to_numeric(df["gmv"], errors="coerce")
    df = df.dropna(subset=["cal_date", "slice_depth", "gmv"]).copy()
    df["cal_date"] = df["cal_date"].astype(int)
    df["slice_depth"] = df["slice_depth"].astype(int)
    df["gmv"] = df["gmv"].astype(float)
    # ADDED: Метрики не являются признаками; приводим доступные колонки к числам для расчёта WoW-процентов.
    for metric in METRIC_COLUMNS:
        if metric in df.columns:
            df[metric] = pd.to_numeric(df[metric], errors="coerce")

    dims = infer_anomaly_dimension_columns(df, dim_cols)
    for col in dims:
        df[col] = df[col].map(normalize_dim_value)

    df["segment_id"] = df.apply(lambda row: segment_id_from_row(row, dims), axis=1)
    key_level_depth = df.apply(lambda row: build_segment_key_and_level(row, dims), axis=1)
    df["segment_key"] = [item[0] for item in key_level_depth]
    df["segment_level"] = [item[1] for item in key_level_depth]

    total_df = df[df["slice_depth"] == 0].copy()
    if total_df.empty:
        raise ValueError("Не найден total-слой slice_depth = 0")
    total_by_date = total_df.groupby("cal_date", as_index=False)["gmv"].sum()
    dates = sorted(total_by_date["cal_date"].astype(int).tolist())
    if len(dates) < 4:
        raise ValueError("Для robust z-score нужно минимум 4 недели в total-слое")

    date_diffs = pd.Series(dates).diff().dropna().astype(int).tolist()
    if date_diffs and any(diff != 7 for diff in date_diffs):
        raise ValueError("В total-слое есть пропуск календарной недели; без total GMV неделю нельзя восстановить")
    if (total_by_date["gmv"].astype(float) <= 0).any():
        raise ValueError("Total GMV должен быть положительным на каждой неделе")

    return df.reset_index(drop=True), dims, dates


def build_full_week_grid(history_df: pd.DataFrame, dim_cols: Sequence[str], dates: Sequence[int]) -> pd.DataFrame:
    """Построить полную сетку segment x week с GMV = 0 для отсутствующих строк.

    Args:
        history_df: Очищенная историческая таблица.
        dim_cols: Список признаков.
        dates: Полный список недель из total-слоя.

    Returns:
        Панель сегментов по неделям.

    Raises:
        ValueError: Если нет сегментов.

    Examples:
        >>> # panel = build_full_week_grid(history_df, ['geo'], [1, 8, 15])
    """

    meta_cols = ["segment_id", "segment_key", "segment_level", "slice_depth", *dim_cols]
    meta = history_df[meta_cols].drop_duplicates("segment_id").copy()
    if meta.empty:
        raise ValueError("Не найдены сегменты для построения недельной сетки")

    # ADDED: Сохраняем операционные метрики в недельной панели. Вход считается предагрегированным;
    # для счётчиков берём сумму, для ratio-метрик берём первое значение строки среза.
    available_metric_cols = [metric for metric in METRIC_COLUMNS if metric in history_df.columns]
    metric_agg = {
        metric: (metric, "sum" if metric in {"tx", "au", "am"} else "first")
        for metric in available_metric_cols
    }
    weekly = (
        history_df.groupby(["segment_id", "cal_date"], as_index=False)
        .agg(gmv=("gmv", "sum"), source_row_count=("gmv", "size"), **metric_agg)
    )
    full_index = pd.MultiIndex.from_product(
        [meta["segment_id"].tolist(), list(dates)],
        names=["segment_id", "cal_date"],
    )
    panel = full_index.to_frame(index=False).merge(weekly, how="left", on=["segment_id", "cal_date"])
    panel["gmv"] = panel["gmv"].fillna(0.0).astype(float)
    for metric in available_metric_cols:
        panel[metric] = panel[metric].fillna(0.0).astype(float)
    panel["source_row_count"] = panel["source_row_count"].fillna(0).astype(int)
    panel = panel.merge(meta, how="left", on="segment_id")
    panel["row_missing_in_source"] = panel["source_row_count"] == 0
    return panel.sort_values(["segment_id", "cal_date"]).reset_index(drop=True)


def _history_reliability(nonzero_weeks: int) -> float:
    """Оценить надёжность истории сегмента.

    Args:
        nonzero_weeks: Количество ненулевых исторических недель до текущей.

    Returns:
        Коэффициент надёжности истории.

    Raises:
        ValueError: Не выбрасывается.

    Examples:
        >>> _history_reliability(8)
        1.0
    """

    if nonzero_weeks >= 8:
        return 1.0
    if nonzero_weeks >= 4:
        return 0.7
    if nonzero_weeks >= 1:
        return 0.4
    return 0.4


def calculate_segment_anomaly(
    segment_panel: pd.DataFrame,
    total_by_date: pd.Series,
    dates: Sequence[int],
    current_cal_date: int,
    thresholds: AnomalyThresholds,
) -> Dict[str, object]:
    """Посчитать метрики необычности одного сегмента.

    Args:
        segment_panel: Недельная панель одного сегмента.
        total_by_date: Total GMV по неделям.
        dates: Полный список недель.
        current_cal_date: Анализируемая текущая неделя.
        thresholds: Пороги алгоритма.

    Returns:
        Словарь метрик сегмента.

    Raises:
        ValueError: Если текущая неделя первая в истории.

    Examples:
        >>> # metrics = calculate_segment_anomaly(panel_one_segment, total, dates, dates[-1], AnomalyThresholds())
    """

    current_idx = list(dates).index(int(current_cal_date))
    if current_idx == 0:
        raise ValueError("Текущая неделя не может быть первой неделей истории")
    previous_cal_date = int(dates[current_idx - 1])

    segment_panel = segment_panel.sort_values("cal_date")
    gmv = segment_panel.set_index("cal_date")["gmv"].astype(float).reindex(dates, fill_value=0.0)
    total = total_by_date.reindex(dates).astype(float)
    share = gmv / total
    share_delta = share.diff()
    gmv_delta = gmv.diff()
    relative_growth = gmv_delta / gmv.shift(1)
    relative_growth = relative_growth.where(gmv.shift(1) > 0)

    # FIXED: z-score считаем по истории относительных приростов GMV сегмента, а не по изменению доли.
    history_growth = relative_growth.reindex(dates[1:current_idx]).dropna()
    baseline_growth = _safe_float(history_growth.median(), 0.0) if not history_growth.empty else 0.0
    mad = _safe_float((history_growth - baseline_growth).abs().median(), 0.0) if not history_growth.empty else 0.0
    sigma = max(1.4826 * mad, thresholds.sigma_floor)

    previous_share = _safe_float(share.loc[previous_cal_date], 0.0)
    current_gmv = _safe_float(gmv.loc[current_cal_date], 0.0)
    previous_gmv = _safe_float(gmv.loc[previous_cal_date], 0.0)
    wow_delta_gmv = current_gmv - previous_gmv

    history_nonzero_weeks = int((gmv.reindex(dates[:current_idx]) > 0).sum())
    reliability = _history_reliability(history_nonzero_weeks)

    if current_gmv > 0 and history_nonzero_weeks == 0:
        state = "новый сегмент"
    elif current_gmv > 0 and previous_gmv == 0 and history_nonzero_weeks > 0:
        state = "возобновившийся сегмент"
    elif current_gmv == 0 and previous_gmv > 0:
        state = "исчезнувший сегмент"
    else:
        state = "обычный"

    relative_wow = None
    if previous_gmv > 0:
        relative_wow = wow_delta_gmv / previous_gmv

    # ADDED: WoW-проценты по операционным метрикам для менеджерского вывода.
    metric_values: Dict[str, object] = {}
    for metric in METRIC_COLUMNS:
        if metric not in segment_panel.columns:
            continue
        metric_series = segment_panel.set_index("cal_date")[metric].astype(float).reindex(dates, fill_value=0.0)
        previous_metric = _safe_float(metric_series.loc[previous_cal_date], 0.0)
        current_metric = _safe_float(metric_series.loc[current_cal_date], 0.0)
        metric_values[f"{metric}_previous"] = previous_metric
        metric_values[f"{metric}_current"] = current_metric
        metric_values[f"{metric}_wow_pct"] = None if previous_metric == 0 else (current_metric - previous_metric) / previous_metric

    if relative_wow is None:
        robust_z = thresholds.z_cap if state != "обычный" and wow_delta_gmv != 0 else 0.0
    else:
        robust_z = (relative_wow - baseline_growth) / sigma
    # Ограничение на z снято специально
    # robust_z_capped = max(-thresholds.z_cap, min(thresholds.z_cap, robust_z))
    # abs_z_capped = min(abs(robust_z), thresholds.z_cap)
    robust_z_capped = robust_z
    abs_z_capped = abs(robust_z)

    return {
        "current_cal_date": int(current_cal_date),
        "previous_cal_date": int(previous_cal_date),
        "gmv_current": current_gmv,
        "gmv_previous": previous_gmv,
        "wow_delta_gmv": wow_delta_gmv,
        "relative_wow": relative_wow,
        "share_current": _safe_float(share.loc[current_cal_date], 0.0),
        "share_previous": previous_share,
        "share_delta_current": _safe_float(share_delta.loc[current_cal_date], 0.0),
        "baseline_relative_growth": baseline_growth,
        "mad_relative_growth": mad,
        "sigma_relative_growth": sigma,
        "robust_z": robust_z,
        "robust_z_capped": robust_z_capped,
        "abs_z_capped": abs_z_capped,
        "abnormal_gmv": wow_delta_gmv,
        "abs_abnormal_gmv": abs(wow_delta_gmv),
        "history_nonzero_weeks": history_nonzero_weeks,
        "history_points": int(len(history_growth)),
        "reliability_factor": reliability,
        "state": state,
        **metric_values,
    }


def build_anomaly_candidates(
    panel_df: pd.DataFrame,
    dim_cols: Sequence[str],
    dates: Sequence[int],
    thresholds: AnomalyThresholds,
    current_cal_date: Optional[int] = None,
) -> Tuple[pd.DataFrame, pd.Series]:
    """Посчитать аномальность независимо для каждого среза.

    Args:
        panel_df: Полная недельная панель segment x week.
        dim_cols: Список признаков.
        dates: Список недель.
        thresholds: Пороги алгоритма.
        current_cal_date: Текущая неделя. Если None, берётся последняя.

    Returns:
        Кортеж: таблица кандидатов и total GMV по неделям.

    Raises:
        ValueError: Если текущая неделя отсутствует.

    Examples:
        >>> # candidates, total = build_anomaly_candidates(panel, dims, dates, AnomalyThresholds())
    """

    current = int(current_cal_date) if current_cal_date is not None else int(dates[-1])
    if current not in dates:
        raise ValueError(f"Текущая неделя {current} отсутствует в total-слое")

    total_panel = panel_df[panel_df["slice_depth"].astype(int) == 0].copy()
    total_by_date = total_panel.groupby("cal_date")["gmv"].sum().reindex(dates).astype(float)
    if total_by_date.isna().any() or (total_by_date <= 0).any():
        raise ValueError("Total GMV должен присутствовать и быть положительным на каждой неделе")

    rows: List[Dict[str, object]] = []
    meta_cols = ["segment_id", "segment_key", "segment_level", "slice_depth", *dim_cols]
    for segment_id, segment_panel in panel_df.groupby("segment_id", sort=False):
        meta = segment_panel.iloc[0][meta_cols].to_dict()
        metrics = calculate_segment_anomaly(segment_panel, total_by_date, dates, current, thresholds)
        rows.append({**meta, **metrics})

    candidates = pd.DataFrame(rows)
    candidates["slice_depth"] = candidates["slice_depth"].astype(int)
    # ADDED: Материальность считаем как долю изменения сегмента в gross movement атомарного слоя.
    max_depth = int(candidates["slice_depth"].max())
    atomic_candidates = candidates[candidates["slice_depth"] == max_depth].copy()
    gross_atomic_movement = float(atomic_candidates["wow_delta_gmv"].astype(float).abs().sum())
    candidates["gross_atomic_movement"] = gross_atomic_movement
    candidates["materiality_share"] = 0.0
    if gross_atomic_movement > 0:
        candidates["materiality_share"] = candidates["wow_delta_gmv"].astype(float).abs() / gross_atomic_movement
    # ADDED: Предварительный фильтр аномальности. Неаномальные или нематериальные сегменты
    # остаются в диагностике, но не участвуют в оптимизационном выборе.
    candidates["passes_initial_anomaly_filter"] = (
        (candidates["slice_depth"].astype(int) > 0)
        & (candidates["abs_z_capped"].astype(float) >= thresholds.min_z_score)
        & (candidates["materiality_share"].astype(float) >= thresholds.min_materiality_share)
        & (candidates["wow_delta_gmv"].astype(float).abs() >= thresholds.min_anomaly_abs)
    )
    return candidates.sort_values(["slice_depth", "segment_key"]).reset_index(drop=True), total_by_date


def build_atomic_coverage(candidates: pd.DataFrame, dim_cols: Sequence[str]) -> Dict[str, frozenset[str]]:
    """Построить покрытие каждого кандидата атомарными сегментами.

    Args:
        candidates: Таблица кандидатов.
        dim_cols: Список признаков.

    Returns:
        Словарь segment_id -> frozenset атомарных segment_id.

    Raises:
        ValueError: Если атомарный слой пустой.

    Examples:
        >>> # coverage = build_atomic_coverage(candidates, ['geo', 'products'])
    """

    max_depth = int(candidates["slice_depth"].max())
    atomic_df = candidates[candidates["slice_depth"] == max_depth].copy()
    if atomic_df.empty:
        raise ValueError("Атомарный слой пустой")

    coverage: Dict[str, frozenset[str]] = {}
    for _, candidate in candidates.iterrows():
        atoms = []
        for _, atom in atomic_df.iterrows():
            if candidate_covers_atomic(candidate, atom, dim_cols):
                atoms.append(str(atom["segment_id"]))
        coverage[str(candidate["segment_id"])] = frozenset(atoms)
    return coverage


def apply_local_depth_penalty(
    candidates: pd.DataFrame,
    coverage: Dict[str, frozenset[str]],
    depth_factor: float = 0.9,
) -> pd.DataFrame:
    """ADDED: Рассчитать локальный штраф за глубину относительно ветки сегмента.

    Args:
        candidates: Таблица всех кандидатов до оптимизационного отбора.
        coverage: Атомарное покрытие `segment_id -> set[atomic_segment_id]`, построенное по всем сегментам.
        depth_factor: Коэффициент штрафа за один уровень до максимальной eligible-глубины ветки.

    Returns:
        Копия `candidates` с колонками `base_anomaly_score`, `local_max_eligible_depth`,
        `local_depth_gap`, `eligible_descendant_count`, `depth_score_weight`, `anomaly_score`.

    Raises:
        ValueError: Если не хватает обязательных колонок, depth_factor некорректен,
            или eligible-сегмент имеет отсутствующее/пустое покрытие либо некорректный base score.

    Examples:
        >>> df = pd.DataFrame([
        ...     {'segment_id': 'p', 'slice_depth': 1, 'passes_initial_anomaly_filter': True,
        ...      'abs_z_capped': 2.0, 'materiality_share': 0.5, 'reliability_factor': 1.0},
        ... ])
        >>> apply_local_depth_penalty(df, {'p': frozenset({'a'})})['depth_score_weight'].iloc[0]
        1.0
    """

    required_columns = {
        "segment_id",
        "slice_depth",
        "passes_initial_anomaly_filter",
        "abs_z_capped",
        "materiality_share",
        "reliability_factor",
    }
    missing_columns = sorted(required_columns - set(candidates.columns))
    if missing_columns:
        raise ValueError(f"Для apply_local_depth_penalty не хватает колонок: {missing_columns}")
    if not 0.0 < float(depth_factor) <= 1.0:
        raise ValueError("depth_factor должен быть в интервале (0, 1]")

    result = candidates.copy()
    result["segment_id"] = result["segment_id"].astype(str)
    result["slice_depth"] = result["slice_depth"].astype(int)
    result["base_anomaly_score"] = (
        pd.to_numeric(result["abs_z_capped"], errors="coerce")
        * pd.to_numeric(result["materiality_share"], errors="coerce")
        * pd.to_numeric(result["reliability_factor"], errors="coerce")
    )
    result["local_max_eligible_depth"] = result["slice_depth"].astype(int)
    result["local_depth_gap"] = 0
    result["eligible_descendant_count"] = 0
    result["depth_score_weight"] = 1.0
    result["anomaly_score"] = result["base_anomaly_score"] * result["depth_score_weight"]

    normalized_coverage = {
        str(segment_id): frozenset(str(atom_id) for atom_id in (atoms or frozenset()))
        for segment_id, atoms in coverage.items()
    }
    eligible_mask = result["passes_initial_anomaly_filter"].eq(True) & result["slice_depth"].gt(0)
    eligible = result[eligible_mask].copy()
    fatal_issues: List[str] = []
    coverage_by_id: Dict[str, frozenset[str]] = {}
    depth_by_id: Dict[str, int] = {}
    for _, row in eligible.iterrows():
        segment_id = str(row["segment_id"])
        segment_atoms = normalized_coverage.get(segment_id)
        if segment_atoms is None:
            fatal_issues.append(f"{segment_id} ({row.get('segment_key', '')}): отсутствует атомарное покрытие")
            continue
        if not segment_atoms:
            fatal_issues.append(f"{segment_id} ({row.get('segment_key', '')}): пустое атомарное покрытие")
            continue
        base_score = _safe_float(row.get("base_anomaly_score"), math.nan)
        if not math.isfinite(base_score):
            fatal_issues.append(f"{segment_id} ({row.get('segment_key', '')}): некорректный base_anomaly_score")
            continue
        coverage_by_id[segment_id] = segment_atoms
        depth_by_id[segment_id] = int(row["slice_depth"])
    if fatal_issues:
        raise ValueError(
            "Нельзя рассчитать локальный depth penalty для eligible-сегментов: "
            + "; ".join(fatal_issues[:10])
        )

    descendant_depths_by_id: Dict[str, List[int]] = {segment_id: [] for segment_id in coverage_by_id}
    for parent_id, parent_atoms in coverage_by_id.items():
        parent_depth = depth_by_id[parent_id]
        for child_id, child_atoms in coverage_by_id.items():
            if child_id == parent_id:
                continue
            child_depth = depth_by_id[child_id]
            if child_depth > parent_depth and child_atoms.issubset(parent_atoms):
                descendant_depths_by_id[parent_id].append(child_depth)

    index_by_id = {
        str(segment_id): index
        for index, segment_id in result["segment_id"].items()
    }
    for segment_id, descendant_depths in descendant_depths_by_id.items():
        index = index_by_id[segment_id]
        current_depth = depth_by_id[segment_id]
        local_max_depth = max([current_depth, *descendant_depths])
        local_gap = local_max_depth - current_depth
        result.at[index, "local_max_eligible_depth"] = int(local_max_depth)
        result.at[index, "local_depth_gap"] = int(local_gap)
        result.at[index, "eligible_descendant_count"] = int(len(descendant_depths))
        result.at[index, "depth_score_weight"] = float(depth_factor) ** int(local_gap)

    result["anomaly_score"] = result["base_anomaly_score"].astype(float) * result["depth_score_weight"].astype(float)
    return result


def _segment_feature_set_from_key(segment_key: object) -> frozenset[Tuple[str, str]]:
    """ADDED: Восстановить набор признаков сегмента из человекочитаемого ключа.

    Args:
        segment_key: Значение колонки `segment_key`.

    Returns:
        Набор пар `(dimension, value)`.

    Raises:
        ValueError: Не выбрасывается.

    Examples:
        >>> _segment_feature_set_from_key('geo=RF × product=QR') == frozenset({('geo', 'RF'), ('product', 'QR')})
        True
    """

    return frozenset(_segment_key_parts(segment_key))


def _build_coverage_from_segment_keys(candidates: pd.DataFrame) -> Dict[str, frozenset[str]]:
    """ADDED: Построить атомарное покрытие по `segment_key` без внешнего `dim_cols`.

    Args:
        candidates: Таблица кандидатов с `segment_id`, `segment_key`, `slice_depth`.

    Returns:
        Словарь `segment_id -> frozenset(atomic_segment_id)`.

    Raises:
        ValueError: Если таблица не содержит физический атомарный слой.

    Examples:
        >>> df = pd.DataFrame([
        ...     {'segment_id': 'p', 'segment_key': 'a=1', 'slice_depth': 1},
        ...     {'segment_id': 'c', 'segment_key': 'a=1 × b=2', 'slice_depth': 2},
        ... ])
        >>> _build_coverage_from_segment_keys(df)['p']
        frozenset({'c'})
    """

    if candidates.empty:
        return {}
    max_depth = int(candidates["slice_depth"].astype(int).max())
    atomic_df = candidates[candidates["slice_depth"].astype(int).eq(max_depth)].copy()
    if atomic_df.empty:
        raise ValueError("Физический атомарный слой search_anomal пуст")

    feature_sets = {
        str(row["segment_id"]): _segment_feature_set_from_key(row.get("segment_key", ""))
        for _, row in candidates.iterrows()
    }
    atomic_features = {
        str(row["segment_id"]): feature_sets[str(row["segment_id"])]
        for _, row in atomic_df.iterrows()
    }

    coverage: Dict[str, frozenset[str]] = {}
    for _, row in candidates.iterrows():
        segment_id = str(row["segment_id"])
        features = feature_sets.get(segment_id, frozenset())
        depth = int(row["slice_depth"])
        if not features and depth == max_depth:
            coverage[segment_id] = frozenset({segment_id})
            continue
        if not features:
            coverage[segment_id] = frozenset()
            continue
        coverage[segment_id] = frozenset(
            atom_id
            for atom_id, atom_features in atomic_features.items()
            if features.issubset(atom_features)
        )
    return coverage




def _set_packing_canonical_key(segment_id: str, lookup: Dict[str, pd.Series]) -> Tuple[str, int, str]:
    """ADDED: Build a deterministic segment key for solver-independent ordering.

    Args:
        segment_id: Segment identifier.
        lookup: Mapping `segment_id -> candidate row`.

    Returns:
        Tuple with human segment key, depth and technical identifier.

    Raises:
        ValueError: Not raised.

    Examples:
        >>> _set_packing_canonical_key('s', {'s': pd.Series({'segment_key': 'a=1', 'slice_depth': 1})})
        ('a=1', 1, 's')
    """

    row = lookup[segment_id]
    return (str(row.get("segment_key", "")), int(row.get("slice_depth", 0)), str(segment_id))


def _validate_set_packing_duplicates(candidates: pd.DataFrame) -> None:
    """ADDED: Validate that candidates do not contain duplicate segment identifiers or keys.

    Args:
        candidates: Candidate table.

    Returns:
        None.

    Raises:
        ValueError: If duplicate segment ids or duplicate segment keys are found.

    Examples:
        >>> _validate_set_packing_duplicates(pd.DataFrame({'segment_id': ['a'], 'segment_key': ['x']}))
    """

    duplicated_ids = sorted(candidates.loc[candidates["segment_id"].astype(str).duplicated(), "segment_id"].astype(str).unique())
    if duplicated_ids:
        raise ValueError(f"Duplicate segment_id values in candidates: {duplicated_ids[:10]}")
    duplicated_keys = sorted(candidates.loc[candidates["segment_key"].astype(str).duplicated(), "segment_key"].astype(str).unique())
    if duplicated_keys:
        raise ValueError(f"Duplicate segment_key values in candidates: {duplicated_keys[:10]}")


def _validate_set_packing_segment_keys(
    candidates: pd.DataFrame,
    dim_cols: Optional[Sequence[str]],
) -> Dict[str, Tuple[bool, str]]:
    """ADDED: Validate that segment keys are parseable and use known dimensions.

    Args:
        candidates: Candidate table with `segment_id`, `segment_key` and `slice_depth`.
        dim_cols: Known dimension columns. If None, only syntactic validation is applied.

    Returns:
        Mapping `segment_id -> (is_valid, reason)`.

    Raises:
        ValueError: Not raised; invalid keys are returned as diagnostic statuses.

    Examples:
        >>> df = pd.DataFrame([{'segment_id': 's', 'segment_key': 'a=1', 'slice_depth': 1}])
        >>> _validate_set_packing_segment_keys(df, ['a'])['s'][0]
        True
    """

    known_dims = set(dim_cols or [])
    result: Dict[str, Tuple[bool, str]] = {}
    for _, row in candidates.iterrows():
        segment_id = str(row["segment_id"])
        depth = int(row["slice_depth"])
        if depth == 0:
            result[segment_id] = (True, "")
            continue
        parts = _segment_key_parts(row.get("segment_key", ""))
        part_dims = [dimension for dimension, _ in parts]
        if not parts:
            result[segment_id] = (False, "segment_key is not parseable")
            continue
        if len(set(part_dims)) != len(part_dims):
            result[segment_id] = (False, "segment_key contains duplicate dimensions")
            continue
        unknown_dims = sorted(set(part_dims) - known_dims) if known_dims else []
        if unknown_dims:
            result[segment_id] = (False, f"segment_key contains unknown dimensions: {unknown_dims}")
            continue
        if len(parts) != depth:
            result[segment_id] = (False, f"segment_key depth {len(parts)} does not match slice_depth {depth}")
            continue
        result[segment_id] = (True, "")
    return result


def _prepare_set_packing_coverage(
    candidates: pd.DataFrame,
    coverage: Optional[Dict[str, frozenset[str]]],
) -> Tuple[Dict[str, frozenset[str]], str, Dict[str, str]]:
    """ADDED: Prepare factual atomic coverage or explicitly marked segment-key fallback.

    Args:
        candidates: Candidate table.
        coverage: Optional factual mapping `segment_id -> atomic segment ids`.

    Returns:
        Tuple with normalized coverage, coverage source and per-segment validation issue.

    Raises:
        ValueError: If fallback coverage cannot be built from segment keys.

    Examples:
        >>> df = pd.DataFrame([{'segment_id': 'a', 'segment_key': 'x=1', 'slice_depth': 1}])
        >>> _prepare_set_packing_coverage(df, None)[1]
        'SEGMENT_KEY_FALLBACK'
    """

    if coverage is None:
        return _build_coverage_from_segment_keys(candidates), "SEGMENT_KEY_FALLBACK", {}

    normalized: Dict[str, frozenset[str]] = {}
    issues: Dict[str, str] = {}
    for raw_segment_id, raw_atoms in coverage.items():
        segment_id = str(raw_segment_id)
        atom_list = [str(atom_id).strip() for atom_id in list(raw_atoms or []) if str(atom_id).strip()]
        if len(atom_list) != len(set(atom_list)):
            issues[segment_id] = "duplicate atomic ids were removed from coverage"
        normalized[segment_id] = frozenset(atom_list)

    for segment_id in candidates["segment_id"].astype(str).tolist():
        if segment_id not in normalized:
            normalized[segment_id] = frozenset()
            issues[segment_id] = "coverage is missing for segment_id"
    return normalized, "FACTUAL_ATOMIC_COVERAGE", issues


def _build_set_packing_conflicts(
    segment_ids: Sequence[str],
    coverage: Dict[str, frozenset[str]],
    lookup: Dict[str, pd.Series],
) -> Tuple[Dict[str, List[str]], Dict[Tuple[str, str], frozenset[str]], Dict[str, int]]:
    """ADDED: Build conflict graph through an atom-to-segments inverted index.

    Args:
        segment_ids: Eligible segment ids.
        coverage: Mapping `segment_id -> atomic segment ids`.
        lookup: Mapping `segment_id -> candidate row`.

    Returns:
        Tuple with `atom_to_segments`, conflict pairs and conflict count by segment.

    Raises:
        ValueError: Not raised.

    Examples:
        >>> lookup = {'a': pd.Series({'segment_key': 'a', 'slice_depth': 1})}
        >>> _build_set_packing_conflicts(['a'], {'a': frozenset({'atom'})}, lookup)[2]['a']
        0
    """

    sorted_segment_ids = sorted(segment_ids, key=lambda segment_id: _set_packing_canonical_key(segment_id, lookup))
    atom_to_segments: Dict[str, List[str]] = {}
    for segment_id in sorted_segment_ids:
        for atom_id in sorted(coverage.get(segment_id, frozenset())):
            atom_to_segments.setdefault(atom_id, []).append(segment_id)

    for atom_id, atom_segment_ids in atom_to_segments.items():
        atom_to_segments[atom_id] = sorted(atom_segment_ids, key=lambda segment_id: _set_packing_canonical_key(segment_id, lookup))

    pair_atoms: Dict[Tuple[str, str], set[str]] = {}
    conflict_neighbors: Dict[str, set[str]] = {segment_id: set() for segment_id in sorted_segment_ids}
    for atom_id, atom_segment_ids in atom_to_segments.items():
        if len(atom_segment_ids) <= 1:
            continue
        for left_id, right_id in combinations(atom_segment_ids, 2):
            pair = tuple(sorted((left_id, right_id), key=lambda segment_id: _set_packing_canonical_key(segment_id, lookup)))
            pair_atoms.setdefault(pair, set()).add(atom_id)
            conflict_neighbors[left_id].add(right_id)
            conflict_neighbors[right_id].add(left_id)

    conflict_pair_atoms = {
        pair: frozenset(atoms)
        for pair, atoms in pair_atoms.items()
    }
    conflict_count_by_segment = {
        segment_id: len(conflict_neighbors.get(segment_id, set()))
        for segment_id in sorted_segment_ids
    }
    return atom_to_segments, conflict_pair_atoms, conflict_count_by_segment


def _build_set_packing_components(
    segment_ids: Sequence[str],
    conflict_pair_atoms: Dict[Tuple[str, str], frozenset[str]],
    coverage: Dict[str, frozenset[str]],
    lookup: Dict[str, pd.Series],
    scores: Dict[str, float],
) -> List[Dict[str, object]]:
    """ADDED: Split the conflict graph into exact independent components.

    Args:
        segment_ids: Eligible segment ids.
        conflict_pair_atoms: Mapping conflict pair -> shared atomic ids.
        coverage: Mapping `segment_id -> atomic segment ids`.
        lookup: Mapping `segment_id -> candidate row`.
        scores: Objective coefficient by segment.

    Returns:
        List of component dictionaries with deterministic component ids.

    Raises:
        ValueError: Not raised.

    Examples:
        >>> lookup = {'a': pd.Series({'segment_key': 'a', 'slice_depth': 1})}
        >>> _build_set_packing_components(['a'], {}, {'a': frozenset({'atom'})}, lookup, {'a': 1.0})[0]['component_id']
        'C001'
    """

    adjacency: Dict[str, set[str]] = {segment_id: set() for segment_id in segment_ids}
    for left_id, right_id in conflict_pair_atoms:
        adjacency[left_id].add(right_id)
        adjacency[right_id].add(left_id)

    raw_components: List[List[str]] = []
    remaining = set(segment_ids)
    while remaining:
        seed_id = min(remaining, key=lambda segment_id: _set_packing_canonical_key(segment_id, lookup))
        stack = [seed_id]
        remaining.remove(seed_id)
        component: List[str] = []
        while stack:
            segment_id = stack.pop()
            component.append(segment_id)
            for neighbor_id in sorted(adjacency.get(segment_id, set()), key=lambda sid: _set_packing_canonical_key(sid, lookup)):
                if neighbor_id in remaining:
                    remaining.remove(neighbor_id)
                    stack.append(neighbor_id)
        raw_components.append(sorted(component, key=lambda sid: _set_packing_canonical_key(sid, lookup)))

    raw_components.sort(key=lambda component: _set_packing_canonical_key(component[0], lookup))
    components: List[Dict[str, object]] = []
    for component_index, component_segment_ids in enumerate(raw_components, start=1):
        component_id = f"C{component_index:03d}"
        component_atom_ids = sorted(
            set().union(*(coverage.get(segment_id, frozenset()) for segment_id in component_segment_ids))
        )
        component_pair_count = sum(
            1
            for left_id, right_id in conflict_pair_atoms
            if left_id in component_segment_ids and right_id in component_segment_ids
        )
        depths = [int(lookup[segment_id]["slice_depth"]) for segment_id in component_segment_ids]
        components.append(
            {
                "component_id": component_id,
                "segment_ids": component_segment_ids,
                "atom_ids": component_atom_ids,
                "conflict_pair_count": component_pair_count,
                "segment_count": len(component_segment_ids),
                "atom_count": len(component_atom_ids),
                "min_depth": min(depths),
                "max_depth": max(depths),
                "score_sum": float(sum(scores.get(segment_id, 0.0) for segment_id in component_segment_ids)),
            }
        )
    return components


def _component_atom_to_segments(
    component_segment_ids: Sequence[str],
    coverage: Dict[str, frozenset[str]],
    lookup: Dict[str, pd.Series],
) -> Dict[str, List[str]]:
    """ADDED: Build atom constraints for one optimization component.

    Args:
        component_segment_ids: Segment ids inside one conflict component.
        coverage: Mapping `segment_id -> atomic segment ids`.
        lookup: Mapping `segment_id -> candidate row`.

    Returns:
        Mapping `atomic_segment_id -> segment ids covering this atom`.

    Raises:
        ValueError: Not raised.

    Examples:
        >>> lookup = {'a': pd.Series({'segment_key': 'a', 'slice_depth': 1})}
        >>> _component_atom_to_segments(['a'], {'a': frozenset({'atom'})}, lookup)['atom']
        ['a']
    """

    atom_to_segments: Dict[str, List[str]] = {}
    for segment_id in component_segment_ids:
        for atom_id in coverage.get(segment_id, frozenset()):
            atom_to_segments.setdefault(atom_id, []).append(segment_id)
    return {
        atom_id: sorted(atom_segment_ids, key=lambda sid: _set_packing_canonical_key(sid, lookup))
        for atom_id, atom_segment_ids in sorted(atom_to_segments.items())
    }


def _set_packing_solver_result(
    component: Dict[str, object],
    solver_name: str,
    solver_status: str,
    selected_ids: Sequence[str],
    objective_value: float,
    best_bound: float,
    absolute_gap: float,
    relative_gap: float,
    solve_time_sec: float,
    variable_count: int,
    constraint_count: int,
    message: str = "",
) -> Dict[str, object]:
    """ADDED: Normalize solver output for diagnostics.

    Args:
        component: Component metadata.
        solver_name: Solver label.
        solver_status: Normalized status.
        selected_ids: Selected segment ids.
        objective_value: Proven or incumbent objective value.
        best_bound: Solver bound in the same maximization scale.
        absolute_gap: Absolute MIP gap.
        relative_gap: Relative MIP gap.
        solve_time_sec: Solver runtime.
        variable_count: Number of binary variables.
        constraint_count: Number of atom constraints.
        message: Optional solver message.

    Returns:
        Component result dictionary.

    Raises:
        ValueError: Not raised.

    Examples:
        >>> _set_packing_solver_result({'component_id': 'C001'}, 'TRIVIAL', 'OPTIMAL', [], 0, 0, 0, 0, 0, 0, 0)['component_id']
        'C001'
    """

    return {
        **component,
        "solver_name": solver_name,
        "solver_status": solver_status,
        "selected_ids": list(selected_ids),
        "objective_value": float(objective_value),
        "best_bound": float(best_bound),
        "absolute_gap": float(absolute_gap),
        "relative_gap": float(relative_gap),
        "solve_time_sec": float(solve_time_sec),
        "variable_count": int(variable_count),
        "constraint_count": int(constraint_count),
        "message": message,
        "is_optimal": solver_status == "OPTIMAL" and abs(float(absolute_gap)) <= 1e-7,
    }


def _set_packing_result_is_proven_optimal(result: Dict[str, object], gap_tolerance: float) -> bool:
    """ADDED: Проверить, что результат компоненты доказанно оптимален.

    Args:
        result: Нормализованный результат solver-а.
        gap_tolerance: Допустимый gap для признания оптимума доказанным.

    Returns:
        True, если статус OPTIMAL и absolute или relative gap находится в допустимой точности.

    Raises:
        ValueError: Не выбрасывается.

    Examples:
        >>> _set_packing_result_is_proven_optimal({'solver_status': 'OPTIMAL', 'relative_gap': 0.0, 'absolute_gap': 0.0}, 1e-9)
        True
    """

    if str(result.get("solver_status", "")) != "OPTIMAL":
        return False
    relative_gap = _safe_float(result.get("relative_gap"), math.nan)
    absolute_gap = _safe_float(result.get("absolute_gap"), math.nan)
    tolerance = float(gap_tolerance) + 1e-12
    return relative_gap <= tolerance or absolute_gap <= tolerance


def _try_solve_component_with_gurobi(
    component: Dict[str, object],
    atom_to_segments: Dict[str, List[str]],
    lookup: Dict[str, pd.Series],
    scores: Dict[str, float],
    gap_tolerance: float,
) -> Optional[Dict[str, object]]:
    """ADDED: Try solving one component with Gurobi if it is installed and licensed.

    Args:
        component: Component metadata.
        atom_to_segments: Atom constraints for the component.
        lookup: Mapping `segment_id -> candidate row`.
        scores: Objective coefficient by segment.
        gap_tolerance: Allowed relative MIP gap.

    Returns:
        Solver result or None if Gurobi is unavailable.

    Raises:
        ValueError: Not raised; solver errors trigger fallback to the next solver.

    Examples:
        >>> # result = _try_solve_component_with_gurobi(component, atoms, lookup, scores, 1e-9)
    """

    try:
        import gurobipy as gp  # type: ignore[import-not-found]
        from gurobipy import GRB  # type: ignore[import-not-found]
    except Exception:
        return None

    start_time = time.perf_counter()
    component_segment_ids = list(component["segment_ids"])
    try:
        model = gp.Model("gmv_set_packing")
        model.Params.OutputFlag = 0
        model.Params.MIPGap = gap_tolerance
        variables = {
            segment_id: model.addVar(vtype=GRB.BINARY, name=f"x_{position}")
            for position, segment_id in enumerate(component_segment_ids)
        }
        model.setObjective(
            gp.quicksum(float(scores[segment_id]) * variables[segment_id] for segment_id in component_segment_ids),
            GRB.MAXIMIZE,
        )
        for atom_segment_ids in atom_to_segments.values():
            model.addConstr(gp.quicksum(variables[segment_id] for segment_id in atom_segment_ids) <= 1)
        model.optimize()
    except Exception as exc:
        return _set_packing_solver_result(
            component,
            "GUROBI",
            "ERROR",
            [],
            0.0,
            math.nan,
            math.nan,
            math.nan,
            time.perf_counter() - start_time,
            len(component_segment_ids),
            len(atom_to_segments),
            str(exc),
        )

    status = "OPTIMAL" if model.Status == GRB.OPTIMAL else str(model.Status)
    selected_ids = [
        segment_id
        for segment_id in component_segment_ids
        if variables[segment_id].X >= 0.5
    ]
    objective_value = float(model.ObjVal) if model.SolCount else 0.0
    best_bound = float(model.ObjBound) if model.SolCount else math.nan
    relative_gap = float(model.MIPGap) if model.SolCount else math.nan
    absolute_gap = abs(best_bound - objective_value) if not math.isnan(best_bound) else math.nan
    return _set_packing_solver_result(
        component,
        "GUROBI",
        status,
        selected_ids,
        objective_value,
        best_bound,
        absolute_gap,
        relative_gap,
        time.perf_counter() - start_time,
        len(component_segment_ids),
        len(atom_to_segments),
        "",
    )


def _try_solve_component_with_scipy(
    component: Dict[str, object],
    atom_to_segments: Dict[str, List[str]],
    lookup: Dict[str, pd.Series],
    scores: Dict[str, float],
    gap_tolerance: float,
) -> Optional[Dict[str, object]]:
    """ADDED: Solve one component through scipy.optimize.milp and HiGHS.

    Args:
        component: Component metadata.
        atom_to_segments: Atom constraints for the component.
        lookup: Mapping `segment_id -> candidate row`.
        scores: Objective coefficient by segment.
        gap_tolerance: Allowed relative MIP gap.

    Returns:
        Solver result or None if SciPy MILP is unavailable.

    Raises:
        ValueError: Not raised; solver errors trigger fallback.

    Examples:
        >>> # result = _try_solve_component_with_scipy(component, atoms, lookup, scores, 1e-9)
    """

    try:
        import numpy as np
        from scipy.optimize import Bounds, LinearConstraint, milp
        from scipy.sparse import coo_array
    except Exception:
        return None

    start_time = time.perf_counter()
    component_segment_ids = sorted(
        list(component["segment_ids"]),
        key=lambda segment_id: _set_packing_canonical_key(segment_id, lookup),
    )
    variable_index = {segment_id: position for position, segment_id in enumerate(component_segment_ids)}
    row_indices: List[int] = []
    col_indices: List[int] = []
    data_values: List[float] = []
    for row_index, atom_id in enumerate(sorted(atom_to_segments)):
        for segment_id in atom_to_segments[atom_id]:
            row_indices.append(row_index)
            col_indices.append(variable_index[segment_id])
            data_values.append(1.0)

    try:
        constraint_matrix = coo_array(
            (data_values, (row_indices, col_indices)),
            shape=(len(atom_to_segments), len(component_segment_ids)),
        ).tocsr()
        constraints = LinearConstraint(
            constraint_matrix,
            lb=-np.inf * np.ones(len(atom_to_segments)),
            ub=np.ones(len(atom_to_segments)),
        )
        result = milp(
            c=-np.array([float(scores[segment_id]) for segment_id in component_segment_ids], dtype=float),
            integrality=np.ones(len(component_segment_ids), dtype=int),
            bounds=Bounds(np.zeros(len(component_segment_ids)), np.ones(len(component_segment_ids))),
            constraints=constraints,
            options={"mip_rel_gap": gap_tolerance},
        )
    except Exception as exc:
        return _set_packing_solver_result(
            component,
            "SCIPY_HIGHS",
            "ERROR",
            [],
            0.0,
            math.nan,
            math.nan,
            math.nan,
            time.perf_counter() - start_time,
            len(component_segment_ids),
            len(atom_to_segments),
            str(exc),
        )

    selected_ids = []
    if getattr(result, "x", None) is not None:
        selected_ids = [
            segment_id
            for segment_id, value in zip(component_segment_ids, result.x)
            if float(value) >= 0.5
        ]
    objective_value = -float(result.fun) if getattr(result, "fun", None) is not None else 0.0
    raw_dual_bound = getattr(result, "mip_dual_bound", math.nan)
    best_bound = -float(raw_dual_bound) if raw_dual_bound is not None and not math.isnan(float(raw_dual_bound)) else objective_value
    raw_gap = getattr(result, "mip_gap", 0.0 if int(getattr(result, "status", -1)) == 0 else math.nan)
    relative_gap = float(raw_gap) if raw_gap is not None else math.nan
    absolute_gap = abs(best_bound - objective_value) if not math.isnan(best_bound) else math.nan
    status = "OPTIMAL" if int(getattr(result, "status", -1)) == 0 and relative_gap <= gap_tolerance + 1e-12 else str(getattr(result, "status", "UNKNOWN"))
    return _set_packing_solver_result(
        component,
        "SCIPY_HIGHS",
        status,
        selected_ids,
        objective_value,
        best_bound,
        absolute_gap,
        relative_gap,
        time.perf_counter() - start_time,
        len(component_segment_ids),
        len(atom_to_segments),
        str(getattr(result, "message", "")),
    )


def _solve_component_exact_branch_and_bound(
    component: Dict[str, object],
    atom_to_segments: Dict[str, List[str]],
    lookup: Dict[str, pd.Series],
    coverage: Dict[str, frozenset[str]],
    scores: Dict[str, float],
) -> Dict[str, object]:
    """ADDED: Solve set packing exactly without external MILP dependencies.

    Args:
        component: Component metadata.
        atom_to_segments: Atom constraints for the component.
        lookup: Mapping `segment_id -> candidate row`.
        coverage: Mapping `segment_id -> atomic segment ids`.
        scores: Objective coefficient by segment.

    Returns:
        Proven optimal component result.

    Raises:
        RecursionError: If Python recursion depth is exceeded on an unusually large component.

    Examples:
        >>> lookup = {'a': pd.Series({'segment_key': 'a', 'slice_depth': 1})}
        >>> component = {'component_id': 'C001', 'segment_ids': ['a'], 'atom_ids': ['atom']}
        >>> _solve_component_exact_branch_and_bound(component, {'atom': ['a']}, lookup, {'a': frozenset({'atom'})}, {'a': 1.0})['solver_status']
        'OPTIMAL'
    """

    start_time = time.perf_counter()
    ordered_ids = sorted(
        list(component["segment_ids"]),
        key=lambda sid: (-float(scores.get(sid, 0.0)), _set_packing_canonical_key(sid, lookup)),
    )
    suffix_positive_score = [0.0] * (len(ordered_ids) + 1)
    for index in range(len(ordered_ids) - 1, -1, -1):
        suffix_positive_score[index] = suffix_positive_score[index + 1] + max(0.0, float(scores.get(ordered_ids[index], 0.0)))

    best_score = 0.0
    best_selected: List[str] = []
    best_signature: Tuple[int, Tuple[Tuple[str, int, str], ...]] = (0, tuple())

    def solution_signature(selected_ids: Sequence[str]) -> Tuple[int, Tuple[Tuple[str, int, str], ...]]:
        """ADDED: Deterministic tie-break signature among equal-score optima.

        Args:
            selected_ids: Selected segment ids.

        Returns:
            Signature preferring fewer rows, then lexical segment order.

        Raises:
            ValueError: Not raised.

        Examples:
            >>> solution_signature([])
            (0, ())
        """

        return (
            len(selected_ids),
            tuple(_set_packing_canonical_key(segment_id, lookup) for segment_id in sorted(selected_ids, key=lambda sid: _set_packing_canonical_key(sid, lookup))),
        )

    def update_best(selected_ids: Sequence[str], score_value: float) -> None:
        """ADDED: Update incumbent solution with deterministic tie handling.

        Args:
            selected_ids: Candidate selected segment ids.
            score_value: Candidate objective value.

        Returns:
            None.

        Raises:
            ValueError: Not raised.

        Examples:
            >>> # update_best(['a'], 1.0)
        """

        nonlocal best_score, best_selected, best_signature
        signature = solution_signature(selected_ids)
        if score_value > best_score + 1e-12 or (abs(score_value - best_score) <= 1e-12 and signature < best_signature):
            best_score = float(score_value)
            best_selected = sorted(selected_ids, key=lambda sid: _set_packing_canonical_key(sid, lookup))
            best_signature = signature

    def branch(index: int, used_atoms: frozenset[str], selected_ids: Tuple[str, ...], score_value: float) -> None:
        """ADDED: Recursive exact branch-and-bound search.

        Args:
            index: Current position in ordered ids.
            used_atoms: Atoms already covered by selected segments.
            selected_ids: Current selected segment ids.
            score_value: Current objective value.

        Returns:
            None.

        Raises:
            RecursionError: If recursion depth is exceeded.

        Examples:
            >>> # branch(0, frozenset(), tuple(), 0.0)
        """

        if score_value + suffix_positive_score[index] < best_score - 1e-12:
            return
        if index >= len(ordered_ids):
            update_best(selected_ids, score_value)
            return
        segment_id = ordered_ids[index]
        segment_score = float(scores.get(segment_id, 0.0))
        segment_atoms = coverage.get(segment_id, frozenset())
        if segment_score > 0.0 and not (used_atoms & segment_atoms):
            branch(index + 1, used_atoms | segment_atoms, (*selected_ids, segment_id), score_value + segment_score)
        branch(index + 1, used_atoms, selected_ids, score_value)

    branch(0, frozenset(), tuple(), 0.0)
    return _set_packing_solver_result(
        component,
        "EXACT_BRANCH_AND_BOUND",
        "OPTIMAL",
        best_selected,
        best_score,
        best_score,
        0.0,
        0.0,
        time.perf_counter() - start_time,
        len(ordered_ids),
        len(atom_to_segments),
        "",
    )


def _solve_set_packing_component(
    component: Dict[str, object],
    coverage: Dict[str, frozenset[str]],
    lookup: Dict[str, pd.Series],
    scores: Dict[str, float],
    gap_tolerance: float,
    max_exact_fallback_size: int,
) -> Dict[str, object]:
    """ADDED: Solve one independent set-packing component with the best available exact solver.

    Args:
        component: Component metadata.
        coverage: Mapping `segment_id -> atomic segment ids`.
        lookup: Mapping `segment_id -> candidate row`.
        scores: Objective coefficient by segment.
        gap_tolerance: Allowed relative MIP gap.
        max_exact_fallback_size: Largest component allowed for internal exact fallback.

    Returns:
        Component solver result.

    Raises:
        RuntimeError: If MILP solvers do not prove optimum and the component is too large for fallback.
        RecursionError: If exact fallback is allowed but exceeds recursion depth.

    Examples:
        >>> lookup = {'a': pd.Series({'segment_key': 'a', 'slice_depth': 1})}
        >>> component = {'component_id': 'C001', 'segment_ids': ['a'], 'atom_ids': ['atom']}
        >>> _solve_set_packing_component(component, {'a': frozenset({'atom'})}, lookup, {'a': 1.0}, 1e-9, 25)['solver_status']
        'OPTIMAL'
    """

    component_segment_ids = list(component["segment_ids"])
    atom_to_segments = _component_atom_to_segments(component_segment_ids, coverage, lookup)
    if all(len(segment_ids) <= 1 for segment_ids in atom_to_segments.values()):
        selected_ids = [
            segment_id
            for segment_id in component_segment_ids
            if float(scores.get(segment_id, 0.0)) > 0.0
        ]
        objective_value = float(sum(scores.get(segment_id, 0.0) for segment_id in selected_ids))
        return _set_packing_solver_result(
            component,
            "TRIVIAL",
            "OPTIMAL",
            sorted(selected_ids, key=lambda sid: _set_packing_canonical_key(sid, lookup)),
            objective_value,
            objective_value,
            0.0,
            0.0,
            0.0,
            len(component_segment_ids),
            len(atom_to_segments),
            "",
        )

    last_result: Optional[Dict[str, object]] = None
    for solver in (_try_solve_component_with_gurobi, _try_solve_component_with_scipy):
        result = solver(component, atom_to_segments, lookup, scores, gap_tolerance)
        if result is None:
            continue
        last_result = result
        if _set_packing_result_is_proven_optimal(result, gap_tolerance):
            return result

    if len(component_segment_ids) > int(max_exact_fallback_size):
        last_solver = str(last_result["solver_name"]) if last_result is not None else "NONE"
        last_status = str(last_result["solver_status"]) if last_result is not None else "UNAVAILABLE"
        raise RuntimeError(
            "MILP solver не доказал optimum, exact fallback слишком велик: "
            f"component={component.get('component_id')}, "
            f"segment_count={len(component_segment_ids)}, "
            f"limit={int(max_exact_fallback_size)}, "
            f"last_solver={last_solver}, last_status={last_status}. "
            "Для production-расчёта нужен рабочий точный MILP solver."
        )

    exact_result = _solve_component_exact_branch_and_bound(component, atom_to_segments, lookup, coverage, scores)
    if last_result is not None and exact_result["message"] == "":
        exact_result["message"] = f"Fallback after {last_result['solver_name']} status {last_result['solver_status']}"
    return exact_result


def _build_set_packing_decision_log(
    component_results: Sequence[Dict[str, object]],
    atom_to_segments: Dict[str, List[str]],
    conflict_pair_atoms: Dict[Tuple[str, str], frozenset[str]],
    lookup: Dict[str, pd.Series],
    scores: Dict[str, float],
    global_status: str,
) -> pd.DataFrame:
    """ADDED: Build a transparent journal for set-packing optimization.

    Args:
        component_results: Solver results by independent component.
        atom_to_segments: Inverted index `atomic_segment_id -> segment ids`.
        conflict_pair_atoms: Mapping conflict pair -> shared atomic ids.
        lookup: Mapping `segment_id -> candidate row`.
        scores: Objective coefficient by segment.
        global_status: Overall optimization status.

    Returns:
        Decision log DataFrame.

    Raises:
        ValueError: Not raised.

    Examples:
        >>> _build_set_packing_decision_log([], {}, {}, {}, {}, 'EMPTY').empty
        False
    """

    rows: List[Dict[str, object]] = [
        {
            "event_type": "GLOBAL_OPTIMIZATION_SUMMARY",
            "global_status": global_status,
            "component_count": len(component_results),
            "selected_count": sum(len(result.get("selected_ids", [])) for result in component_results),
            "objective_value": sum(float(result.get("objective_value", 0.0)) for result in component_results),
            "best_bound": sum(float(result.get("best_bound", 0.0)) for result in component_results),
            "absolute_gap": sum(float(result.get("absolute_gap", 0.0)) for result in component_results),
            "relative_gap": max((float(result.get("relative_gap", 0.0)) for result in component_results), default=0.0),
            "solver_status": global_status,
        }
    ]

    for atom_id, segment_ids in sorted(atom_to_segments.items()):
        rows.append(
            {
                "event_type": "ATOM_TO_SEGMENTS",
                "atomic_segment_id": atom_id,
                "segment_count": len(segment_ids),
                "segment_ids": " || ".join(segment_ids),
                "segment_keys": " || ".join(str(lookup[segment_id]["segment_key"]) for segment_id in segment_ids),
            }
        )

    for (left_id, right_id), shared_atoms in sorted(
        conflict_pair_atoms.items(),
        key=lambda item: (
            _set_packing_canonical_key(item[0][0], lookup),
            _set_packing_canonical_key(item[0][1], lookup),
        ),
    ):
        rows.append(
            {
                "event_type": "CONFLICT_PAIR",
                "left_segment_id": left_id,
                "left_segment_key": str(lookup[left_id]["segment_key"]),
                "right_segment_id": right_id,
                "right_segment_key": str(lookup[right_id]["segment_key"]),
                "shared_atomic_count": len(shared_atoms),
                "shared_atomic_ids": " || ".join(sorted(shared_atoms)),
            }
        )

    for result in component_results:
        selected_ids = set(result.get("selected_ids", []))
        rows.append(
            {
                "event_type": "COMPONENT_SOLVE",
                "component_id": result["component_id"],
                "solver_name": result["solver_name"],
                "solver_status": result["solver_status"],
                "objective_value": result["objective_value"],
                "best_bound": result["best_bound"],
                "absolute_gap": result["absolute_gap"],
                "relative_gap": result["relative_gap"],
                "solve_time_sec": result["solve_time_sec"],
                "variable_count": result["variable_count"],
                "constraint_count": result["constraint_count"],
                "segment_count": result["segment_count"],
                "atom_count": result["atom_count"],
                "conflict_pair_count": result["conflict_pair_count"],
                "min_depth": result["min_depth"],
                "max_depth": result["max_depth"],
                "score_sum": result["score_sum"],
                "selected_ids": " || ".join(sorted(selected_ids, key=lambda sid: _set_packing_canonical_key(sid, lookup))),
                "message": result.get("message", ""),
            }
        )
        for segment_id in result["segment_ids"]:
            rows.append(
                {
                    "event_type": "SEGMENT_DECISION",
                    "component_id": result["component_id"],
                    "segment_id": segment_id,
                    "segment_key": str(lookup[segment_id]["segment_key"]),
                    "slice_depth": int(lookup[segment_id]["slice_depth"]),
                    "anomaly_score": float(scores.get(segment_id, 0.0)),
                    "selected": segment_id in selected_ids,
                    "solver_name": result["solver_name"],
                    "solver_status": result["solver_status"],
                    "global_status": global_status,
                }
            )
    return pd.DataFrame(rows)


def validate_set_packing_solution(
    final_df: pd.DataFrame,
    diagnostics: pd.DataFrame,
    coverage: Dict[str, frozenset[str]],
    component_results: Sequence[Dict[str, object]],
    scores: Dict[str, float],
    global_status: str,
    gap_tolerance: float,
) -> None:
    """ADDED: Обязательно проверить корректность найденного Set Packing решения.

    Args:
        final_df: Итоговые выбранные сегменты.
        diagnostics: Диагностика всех кандидатов после оптимизации.
        coverage: Покрытие `segment_id -> atomic_segment_id`.
        component_results: Результаты решения независимых компонент.
        scores: Вес `anomaly_score` по каждому оптимизируемому сегменту.
        global_status: Итоговый статус оптимизации.
        gap_tolerance: Допустимый solver gap.

    Returns:
        None.

    Raises:
        RuntimeError: Если решение не доказано оптимальным, нарушает атомарные ограничения
            или objective не совпадает с суммой score выбранных сегментов.

    Examples:
        >>> validate_set_packing_solution(pd.DataFrame(), pd.DataFrame(), {}, [], {}, "OPTIMAL", 1e-9)
    """

    objective_tolerance = max(1e-7, float(gap_tolerance) + 1e-12)
    if str(global_status) != "OPTIMAL":
        raise RuntimeError(f"Set Packing solution is not globally OPTIMAL: global_status={global_status}")

    if not diagnostics.empty and {"passes_initial_anomaly_filter", "slice_depth", "set_packing_status"}.issubset(diagnostics.columns):
        passed_mask = diagnostics["passes_initial_anomaly_filter"].eq(True) & diagnostics["slice_depth"].astype(int).gt(0)
        valid_final_statuses = {"SET_PACKING_SELECTED", "SET_PACKING_NOT_SELECTED"}
        unresolved = diagnostics.loc[passed_mask & ~diagnostics["set_packing_status"].isin(valid_final_statuses)]
        if not unresolved.empty:
            examples = [
                f"{row.segment_id}: {row.set_packing_status}"
                for row in unresolved.head(10).itertuples(index=False)
            ]
            raise RuntimeError(
                "После оптимизации остались прошедшие первичный фильтр сегменты вне доказанного Set Packing: "
                + "; ".join(examples)
            )

    selected_ids_from_components: List[str] = []
    component_objective_sum = 0.0
    for result in component_results:
        if not _set_packing_result_is_proven_optimal(result, gap_tolerance):
            raise RuntimeError(
                "Компонента Set Packing не доказала OPTIMAL: "
                f"component={result.get('component_id')}, "
                f"solver={result.get('solver_name')}, "
                f"status={result.get('solver_status')}, "
                f"abs_gap={result.get('absolute_gap')}, rel_gap={result.get('relative_gap')}"
            )
        component_selected_ids = [str(segment_id) for segment_id in result.get("selected_ids", [])]
        if len(component_selected_ids) != len(set(component_selected_ids)):
            raise RuntimeError(f"Компонента {result.get('component_id')} содержит дубли selected_ids")
        missing_scores = sorted(set(component_selected_ids) - set(scores))
        if missing_scores:
            raise RuntimeError(
                f"Для выбранных сегментов компоненты {result.get('component_id')} отсутствует score: {missing_scores[:10]}"
            )
        selected_score_sum = float(sum(float(scores[segment_id]) for segment_id in component_selected_ids))
        objective_value = _safe_float(result.get("objective_value"), math.nan)
        if not math.isclose(selected_score_sum, objective_value, rel_tol=1e-9, abs_tol=objective_tolerance):
            raise RuntimeError(
                "Objective компоненты не равен сумме score выбранных сегментов: "
                f"component={result.get('component_id')}, objective={objective_value}, selected_score_sum={selected_score_sum}"
            )
        component_objective_sum += objective_value
        selected_ids_from_components.extend(component_selected_ids)

    if len(selected_ids_from_components) != len(set(selected_ids_from_components)):
        raise RuntimeError("Один и тот же сегмент выбран более чем в одной компоненте Set Packing")

    final_ids = final_df["segment_id"].astype(str).tolist() if not final_df.empty and "segment_id" in final_df.columns else []
    if set(final_ids) != set(selected_ids_from_components):
        raise RuntimeError(
            "final_df не совпадает с selected_ids solver-а: "
            f"final_only={sorted(set(final_ids) - set(selected_ids_from_components))[:10]}, "
            f"solver_only={sorted(set(selected_ids_from_components) - set(final_ids))[:10]}"
        )
    final_score_sum = (
        float(pd.to_numeric(final_df["selection_score"], errors="coerce").sum())
        if not final_df.empty and "selection_score" in final_df.columns
        else 0.0
    )
    if not math.isclose(final_score_sum, component_objective_sum, rel_tol=1e-9, abs_tol=objective_tolerance):
        raise RuntimeError(
            "Глобальный objective не равен сумме objective компонент или score итоговых сегментов: "
            f"final_score_sum={final_score_sum}, component_objective_sum={component_objective_sum}"
        )

    atom_to_selected_segments: Dict[str, List[str]] = {}
    for segment_id in selected_ids_from_components:
        segment_atoms = coverage.get(segment_id, frozenset())
        if not segment_atoms:
            raise RuntimeError(f"Выбранный сегмент {segment_id} не имеет атомарного покрытия")
        for atom_id in segment_atoms:
            atom_to_selected_segments.setdefault(str(atom_id), []).append(segment_id)
    atom_violations = {
        atom_id: segment_ids
        for atom_id, segment_ids in atom_to_selected_segments.items()
        if len(segment_ids) > 1
    }
    if atom_violations:
        examples = [
            f"{atom_id}: {' || '.join(segment_ids)}"
            for atom_id, segment_ids in list(atom_violations.items())[:10]
        ]
        raise RuntimeError(
            "Нарушено ограничение Set Packing sum(x_i covering atom) <= 1: "
            + "; ".join(examples)
        )


def search_anomal(
    candidates: pd.DataFrame,
    thresholds: AnomalyThresholds,
    coverage: Optional[Dict[str, frozenset[str]]] = None,
    dim_cols: Optional[Sequence[str]] = None,
) -> Tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    """FIXED: Select anomalies via exact Maximum Weighted Set Packing.

    Args:
        candidates: Candidate table after `build_anomaly_candidates`.
        thresholds: Algorithm thresholds; only anomaly filters and set-packing gap tolerance affect selection.
        coverage: Preferred factual mapping `segment_id -> observed atomic segment ids`.
        dim_cols: Known dimension columns used to validate `segment_key`.

    Returns:
        Tuple with final selected anomalies, candidate diagnostics and optimization decision log.

    Raises:
        ValueError: If required columns are missing, duplicate candidates exist, or atomic coverage is invalid.
        RecursionError: If all MILP solvers are unavailable and exact fallback exceeds recursion depth.

    Examples:
        >>> # final_df, diagnostics, log = search_anomal(candidates, AnomalyThresholds(), coverage, dims)
    """

    required_columns = {
        "segment_id",
        "segment_key",
        "slice_depth",
        "passes_initial_anomaly_filter",
        "robust_z_capped",
        "wow_delta_gmv",
        "anomaly_score",
    }
    missing_columns = sorted(required_columns - set(candidates.columns))
    if missing_columns:
        raise ValueError(f"Для search_anomal не хватает колонок: {missing_columns}")

    diagnostics = candidates.copy()
    if diagnostics.empty:
        return diagnostics.copy(), diagnostics, _build_set_packing_decision_log([], {}, {}, {}, {}, "EMPTY")

    diagnostics["segment_id"] = diagnostics["segment_id"].astype(str)
    diagnostics["segment_key"] = diagnostics["segment_key"].astype(str)
    diagnostics["slice_depth"] = diagnostics["slice_depth"].astype(int)
    _validate_set_packing_duplicates(diagnostics)
    normalized_coverage, coverage_source, coverage_issues = _prepare_set_packing_coverage(diagnostics, coverage)
    key_validation = _validate_set_packing_segment_keys(diagnostics, dim_cols)

    string_columns = [
        "action",
        "output_block",
        "reason",
        "original_atomic_descendants",
        "selected_atomic_descendants",
        "set_packing_status",
        "set_packing_global_status",
        "set_packing_component_id",
        "set_packing_solver",
        "set_packing_solver_status",
        "set_packing_reason",
        "atomic_coverage_source",
        "atomic_coverage_validation_status",
        "conflict_segment_ids",
        "conflict_segment_keys",
    ]
    for column in string_columns:
        diagnostics[column] = ""

    diagnostics["is_eligible"] = False
    # FIXED: Не подменяем нечисловой score нулём; прошедший первичный фильтр кандидат
    # с некорректным score должен остановить расчёт, а не исчезнуть из оптимизации.
    diagnostics["selection_score"] = pd.to_numeric(diagnostics["anomaly_score"], errors="coerce")
    diagnostics["covered_atomic_count"] = diagnostics["segment_id"].map(lambda sid: len(normalized_coverage.get(str(sid), frozenset())))
    diagnostics["original_atomic_count"] = 0
    diagnostics["selected_atomic_count"] = 0
    diagnostics["is_resolved"] = False
    diagnostics["selected"] = False
    diagnostics["selection_exclusion_reason"] = ""
    diagnostics["conflict_count"] = 0
    diagnostics["set_packing_objective_value"] = math.nan
    diagnostics["set_packing_best_bound"] = math.nan
    diagnostics["set_packing_abs_gap"] = math.nan
    diagnostics["set_packing_rel_gap"] = math.nan
    diagnostics["set_packing_solve_time_sec"] = math.nan
    diagnostics["set_packing_variable_count"] = 0
    diagnostics["set_packing_constraint_count"] = 0
    diagnostics["set_packing_component_segment_count"] = 0
    diagnostics["set_packing_component_atom_count"] = 0
    diagnostics["set_packing_component_conflict_pair_count"] = 0
    diagnostics["set_packing_component_score_sum"] = 0.0

    lookup = {
        str(row["segment_id"]): row.copy()
        for _, row in diagnostics.iterrows()
    }
    index_by_id = {
        str(segment_id): index
        for index, segment_id in diagnostics["segment_id"].items()
    }

    eligible_ids: List[str] = []
    fatal_input_issues: List[str] = []
    for _, row in diagnostics.iterrows():
        segment_id = str(row["segment_id"])
        index = index_by_id[segment_id]
        depth = int(row["slice_depth"])
        passed_initial_filter = bool(row.get("passes_initial_anomaly_filter", False))
        atoms = sorted(normalized_coverage.get(segment_id, frozenset()))
        key_is_valid, key_reason = key_validation.get(segment_id, (False, "segment_key validation missing"))
        coverage_issue = coverage_issues.get(segment_id, "")
        score_value = _safe_float(row.get("selection_score"), math.nan)
        diagnostics.at[index, "atomic_coverage_source"] = coverage_source
        diagnostics.at[index, "atomic_coverage_validation_status"] = coverage_issue or "OK"
        diagnostics.at[index, "original_atomic_descendants"] = " || ".join(atoms)
        diagnostics.at[index, "original_atomic_count"] = len(atoms)
        diagnostics.at[index, "covered_atomic_count"] = len(atoms)

        if depth == 0:
            reason = "total-слой исключён из оптимизационного отбора аномалий"
            status = "NOT_IN_SET_PACKING_GRAPH"
        elif not passed_initial_filter:
            reason = "сегмент не прошёл первичный фильтр аномальности"
            status = "NOT_IN_SET_PACKING_GRAPH"
        elif not key_is_valid:
            reason = key_reason
            status = "INVALID_SEGMENT_KEY"
        elif coverage_issue:
            reason = coverage_issue
            status = "INVALID_ATOMIC_COVERAGE"
        elif not atoms:
            reason = "у сегмента пустое фактическое атомарное покрытие"
            status = "EMPTY_ATOMIC_COVERAGE"
        elif not math.isfinite(score_value):
            reason = "anomaly_score не является конечным числом"
            status = "INVALID_SCORE"
        elif score_value <= 0.0:
            reason = "anomaly_score неположительный"
            status = "NONPOSITIVE_SCORE"
        else:
            reason = "сегмент участвует в точной задаче Maximum Weighted Set Packing"
            status = "SET_PACKING_CANDIDATE"
            eligible_ids.append(segment_id)
            diagnostics.at[index, "is_eligible"] = True

        diagnostics.at[index, "action"] = status
        diagnostics.at[index, "output_block"] = "кандидат Set Packing" if status == "SET_PACKING_CANDIDATE" else "исключён"
        diagnostics.at[index, "reason"] = reason
        diagnostics.at[index, "set_packing_status"] = status
        diagnostics.at[index, "set_packing_reason"] = reason
        diagnostics.at[index, "selection_exclusion_reason"] = "" if status == "SET_PACKING_CANDIDATE" else reason
        if passed_initial_filter and depth > 0 and status != "SET_PACKING_CANDIDATE":
            fatal_input_issues.append(
                f"{segment_id} ({row.get('segment_key', '')}): {status}: {reason}"
            )

    if fatal_input_issues:
        raise ValueError(
            "Нельзя доказать глобальный optimum: часть сегментов, прошедших первичный фильтр, "
            "не может быть корректно включена в Set Packing. "
            + "; ".join(fatal_input_issues[:10])
        )

    if not eligible_ids:
        diagnostics["set_packing_global_status"] = "EMPTY"
        decision_log = _build_set_packing_decision_log([], {}, {}, lookup, {}, "EMPTY")
        return diagnostics.iloc[0:0].copy(), diagnostics, decision_log

    scores = {
        segment_id: float(diagnostics.at[index_by_id[segment_id], "selection_score"])
        for segment_id in eligible_ids
    }
    atom_to_segments, conflict_pair_atoms, conflict_count_by_segment = _build_set_packing_conflicts(
        eligible_ids,
        normalized_coverage,
        lookup,
    )
    components = _build_set_packing_components(
        eligible_ids,
        conflict_pair_atoms,
        normalized_coverage,
        lookup,
        scores,
    )
    gap_tolerance = float(getattr(thresholds, "set_packing_gap_tolerance", 1e-9))
    max_exact_fallback_size = int(getattr(thresholds, "max_exact_fallback_size", 25))
    component_results = [
        _solve_set_packing_component(component, normalized_coverage, lookup, scores, gap_tolerance, max_exact_fallback_size)
        for component in components
    ]
    global_status = (
        "OPTIMAL"
        if all(_set_packing_result_is_proven_optimal(result, gap_tolerance) for result in component_results)
        else "NOT_OPTIMAL"
    )
    selected_ids = {
        segment_id
        for result in component_results
        for segment_id in result.get("selected_ids", [])
    }
    result_by_segment: Dict[str, Dict[str, object]] = {}
    for result in component_results:
        for segment_id in result["segment_ids"]:
            result_by_segment[segment_id] = result

    conflict_neighbors: Dict[str, set[str]] = {segment_id: set() for segment_id in eligible_ids}
    for left_id, right_id in conflict_pair_atoms:
        conflict_neighbors[left_id].add(right_id)
        conflict_neighbors[right_id].add(left_id)

    for segment_id in eligible_ids:
        index = index_by_id[segment_id]
        result = result_by_segment[segment_id]
        component_id = str(result["component_id"])
        atoms = sorted(normalized_coverage.get(segment_id, frozenset()))
        selected = segment_id in selected_ids
        segment_conflicts = sorted(conflict_neighbors.get(segment_id, set()), key=lambda sid: _set_packing_canonical_key(sid, lookup))
        conflict_keys = [str(lookup[conflict_id]["segment_key"]) for conflict_id in segment_conflicts]
        if selected:
            status = "SET_PACKING_SELECTED"
            reason = (
                f"выбран точной оптимизацией Maximum Weighted Set Packing; component={component_id}; "
                f"solver={result['solver_name']}; status={result['solver_status']}"
            )
        elif result["solver_status"] == "OPTIMAL":
            status = "SET_PACKING_NOT_SELECTED"
            reason = (
                f"не выбран глобальным оптимумом компоненты {component_id}; "
                "выбранный непересекающийся набор даёт большую или равную сумму anomaly_score"
            )
        else:
            status = "SET_PACKING_NOT_PROVEN"
            reason = f"компонента {component_id} не вернула доказанный статус OPTIMAL"

        diagnostics.at[index, "selected"] = selected
        diagnostics.at[index, "is_resolved"] = result["solver_status"] == "OPTIMAL"
        diagnostics.at[index, "action"] = status
        diagnostics.at[index, "output_block"] = "оптимальная аномалия Set Packing" if selected else "исключён Set Packing"
        diagnostics.at[index, "reason"] = reason
        diagnostics.at[index, "set_packing_status"] = status
        diagnostics.at[index, "conflict_count"] = int(conflict_count_by_segment.get(segment_id, 0))
        diagnostics.at[index, "conflict_segment_ids"] = " || ".join(segment_conflicts)
        diagnostics.at[index, "conflict_segment_keys"] = " || ".join(conflict_keys)
        diagnostics.at[index, "set_packing_global_status"] = global_status
        diagnostics.at[index, "set_packing_component_id"] = component_id
        diagnostics.at[index, "set_packing_solver"] = str(result["solver_name"])
        diagnostics.at[index, "set_packing_solver_status"] = str(result["solver_status"])
        diagnostics.at[index, "set_packing_reason"] = reason
        diagnostics.at[index, "set_packing_objective_value"] = float(result["objective_value"])
        diagnostics.at[index, "set_packing_best_bound"] = float(result["best_bound"])
        diagnostics.at[index, "set_packing_abs_gap"] = float(result["absolute_gap"])
        diagnostics.at[index, "set_packing_rel_gap"] = float(result["relative_gap"])
        diagnostics.at[index, "set_packing_solve_time_sec"] = float(result["solve_time_sec"])
        diagnostics.at[index, "set_packing_variable_count"] = int(result["variable_count"])
        diagnostics.at[index, "set_packing_constraint_count"] = int(result["constraint_count"])
        diagnostics.at[index, "set_packing_component_segment_count"] = int(result["segment_count"])
        diagnostics.at[index, "set_packing_component_atom_count"] = int(result["atom_count"])
        diagnostics.at[index, "set_packing_component_conflict_pair_count"] = int(result["conflict_pair_count"])
        diagnostics.at[index, "set_packing_component_score_sum"] = float(result["score_sum"])
        diagnostics.at[index, "selected_atomic_descendants"] = " || ".join(atoms) if selected else ""
        diagnostics.at[index, "selected_atomic_count"] = len(atoms) if selected else 0
        diagnostics.at[index, "selection_exclusion_reason"] = "" if selected else reason

    diagnostics.loc[diagnostics["set_packing_global_status"].eq(""), "set_packing_global_status"] = global_status
    final_df = diagnostics[diagnostics["selected"].astype(bool)].copy()
    final_df = final_df.sort_values(
        by=[
            "selection_score",
            "abs_z_capped",
            "materiality_share",
            "abs_abnormal_gmv",
            "reliability_factor",
            "segment_key",
        ],
        ascending=[False, False, False, False, False, True],
        kind="stable",
    ).reset_index(drop=True)
    final_df.insert(0, "rank", range(1, len(final_df) + 1))
    validate_set_packing_solution(
        final_df,
        diagnostics,
        normalized_coverage,
        component_results,
        scores,
        global_status,
        gap_tolerance,
    )
    decision_log = _build_set_packing_decision_log(
        component_results,
        atom_to_segments,
        conflict_pair_atoms,
        lookup,
        scores,
        global_status,
    )
    return final_df, diagnostics, decision_log




def build_manager_summary(
    final_df: pd.DataFrame,
    thresholds: AnomalyThresholds,
    current_total_gmv: float,
    candidates: Optional[pd.DataFrame] = None,
) -> pd.DataFrame:
    """Сформировать менеджерский вывод по необычным сегментам.

    Args:
        final_df: Итоговые выбранные аномалии.
        thresholds: Пороги алгоритма.
        current_total_gmv: Total GMV текущей недели.
        candidates: Диагностика всех кандидатов для добавления структурных изменений.

    Returns:
        Таблица менеджерского вывода.

    Raises:
        ValueError: Не выбрасывается.

    Examples:
        >>> # manager_df = build_manager_summary(final_df, AnomalyThresholds(), 1000)
    """

    def metric_pct_output(row: Optional[pd.Series] = None) -> Dict[str, object]:
        """Вернуть отформатированные WoW-проценты для менеджерской строки.

        Args:
            row: Строка выбранного сегмента или None для служебных строк.

        Returns:
            Словарь с колонками процентов по GMV и операционным метрикам.

        Raises:
            ValueError: Не выбрасывается.

        Examples:
            >>> metric_pct_output(None)["GMV WoW %"]
            ''
        """

        if row is None:
            return {display_col: "" for display_col in MANAGER_METRIC_PCT_COLUMNS.values()}
        return {
            display_col: _format_pct(row.get(source_col))
            for source_col, display_col in MANAGER_METRIC_PCT_COLUMNS.items()
        }

    structure_state_labels = {
        "новый сегмент": "новый",
        "возобновившийся сегмент": "возобновившийся",
        "исчезнувший сегмент": "исчезнувший",
    }

    def structure_change_label(row: pd.Series) -> str:
        """ADDED: Вернуть короткий тип структурного изменения.

        Args:
            row: Строка кандидата или выбранной аномалии.

        Returns:
            Одно из значений `новый`, `возобновившийся`, `исчезнувший` или пустая строка.

        Raises:
            ValueError: Не выбрасывается.

        Examples:
            >>> structure_change_label(pd.Series({'state': 'новый сегмент'}))
            'новый'
        """

        return structure_state_labels.get(str(row.get("state", "")).strip(), "")

    def structure_change_interpretation(row: pd.Series) -> str:
        """ADDED: Сформировать интерпретацию структурного изменения для менеджера.

        Args:
            row: Строка кандидата или выбранной аномалии.

        Returns:
            Текст с типом изменения и GMV текущей/предыдущей недели.

        Raises:
            ValueError: Не выбрасывается.

        Examples:
            >>> structure_change_interpretation(pd.Series({'state': 'новый сегмент', 'gmv_current': 10, 'gmv_previous': 0, 'wow_delta_gmv': 10}))
            'Тип изменения: новый. Сегмент появился впервые на последней неделе; GMV текущей недели: +10 ₽, GMV предыдущей недели: 0 ₽, Delta GMV: +10 ₽.'
        """

        change_label = structure_change_label(row)
        current_gmv_text = _format_rub(float(row.get("gmv_current", 0.0)))
        previous_gmv_text = _format_rub(float(row.get("gmv_previous", 0.0)))
        delta_text = _format_rub(float(row.get("wow_delta_gmv", 0.0)))
        if change_label == "новый":
            detail = "Сегмент появился впервые на последней неделе"
        elif change_label == "возобновившийся":
            detail = "Сегмент вернулся на последней неделе после нулевой предыдущей недели"
        elif change_label == "исчезнувший":
            detail = "Сегмент исчез на последней неделе после ненулевой предыдущей недели"
        else:
            detail = "Зафиксировано структурное изменение сегмента"
        return (
            f"Тип изменения: {change_label}. "
            f"{detail}; GMV текущей недели: {current_gmv_text}, "
            f"GMV предыдущей недели: {previous_gmv_text}, Delta GMV: {delta_text}."
        )

    def structure_change_rows() -> List[Dict[str, object]]:
        """ADDED: Собрать дополнительные строки менеджерского вывода по структурным изменениям.

        Args:
            Нет аргументов.

        Returns:
            Список строк для вкладки `01_Менеджерский_вывод`.

        Raises:
            ValueError: Не выбрасывается.

        Examples:
            >>> structure_change_rows()
            []
        """

        if candidates is None or candidates.empty or "state" not in candidates.columns:
            return []
        state_order = {"новый сегмент": 0, "возобновившийся сегмент": 1, "исчезнувший сегмент": 2}
        structure_df = candidates[
            candidates["state"].astype(str).isin(structure_state_labels)
            & candidates["slice_depth"].astype(int).gt(0)
        ].copy()
        if structure_df.empty:
            return []
        structure_df["state_order"] = structure_df["state"].astype(str).map(state_order).fillna(99).astype(int)
        structure_df["abs_delta_sort"] = structure_df["wow_delta_gmv"].astype(float).abs()
        structure_df = structure_df.sort_values(
            ["state_order", "abs_delta_sort", "slice_depth", "segment_key"],
            ascending=[True, False, False, True],
            kind="stable",
        )
        rows_out: List[Dict[str, object]] = []
        for _, row in structure_df.iterrows():
            rows_out.append(
                {
                    "раздел": "Изменение структуры",
                    "тип": "Изменение структуры",
                    "сегмент": str(row["segment_key"]),
                    "Delta GMV": _format_rub(float(row["wow_delta_gmv"])),
                    **metric_pct_output(row),
                    "z_score": round(float(row["robust_z"]), 2),
                    "интерпретация": structure_change_interpretation(row),
                }
            )
        return rows_out

    rows: List[Dict[str, object]] = [
        {
            "раздел": "Заголовок",
            "тип": "",
            "сегмент": "Менеджерский вывод по необычным сегментам GMV",
            "Delta GMV": "",
            **metric_pct_output(None),
            "z_score": "",
            "интерпретация": f"Total GMV текущей недели: {_format_rub(current_total_gmv)}.",
        }
    ]

    if final_df.empty:
        rows.append(
            {
                "раздел": "Краткий вывод",
                "тип": "",
                "сегмент": "",
                "Delta GMV": "",
                **metric_pct_output(None),
                "z_score": "",
                "интерпретация": "Материальные необычные сегменты по заданным порогам не найдены.",
            }
        )
    else:
        top = final_df.head(thresholds.max_manager_facts)
        main = top.iloc[0]
        rows.append(
            {
                "раздел": "Краткий вывод",
                "тип": str(main["output_block"]),
                "сегмент": str(main["segment_key"]),
                "Delta GMV": _format_rub(float(main["wow_delta_gmv"])),
                **metric_pct_output(main),
                "z_score": round(float(main["robust_z"]), 2),
                "интерпретация": "Самый сильный выбранный сегмент из глобально оптимального непересекающегося набора по anomaly_score.",
            }
        )

        for _, row in top.iterrows():
            direction = "выше" if float(row["abnormal_gmv"]) > 0 else "ниже"
            interpretation = (
                f"Фактический GMV сегмента {direction} ожидаемого уровня. "
                f"Причина отбора: {row['reason']}."
            )
            rows.append(
                {
                    "раздел": "Таблица факторов",
                    "тип": str(row["output_block"]),
                    "сегмент": str(row["segment_key"]),
                    "Delta GMV": _format_rub(float(row["wow_delta_gmv"])),
                    **metric_pct_output(row),
                    "z_score": round(float(row["robust_z"]), 2),
                    "интерпретация": interpretation,
                }
            )

    rows.extend(structure_change_rows())

    return pd.DataFrame(rows)


def build_history_for_selected(panel_df: pd.DataFrame, final_df: pd.DataFrame, total_by_date: pd.Series) -> pd.DataFrame:
    """Подготовить историю недель для выбранных аномалий.

    Args:
        panel_df: Полная недельная панель.
        final_df: Итоговые выбранные аномалии.
        total_by_date: Total GMV по неделям.

    Returns:
        Таблица истории выбранных сегментов.

    Raises:
        ValueError: Не выбрасывается.

    Examples:
        >>> # history = build_history_for_selected(panel, final_df, total)
    """

    if final_df.empty:
        return pd.DataFrame()

    selected_ids = final_df["segment_id"].astype(str).tolist()
    history = panel_df[panel_df["segment_id"].astype(str).isin(selected_ids)].copy()
    history["total_gmv"] = history["cal_date"].map(total_by_date.to_dict()).astype(float)
    history["share_in_total"] = history["gmv"] / history["total_gmv"]
    history = history.sort_values(["segment_key", "cal_date"])
    history["share_delta"] = history.groupby("segment_id")["share_in_total"].diff()
    return history[
        [
            "segment_id",
            "segment_key",
            "slice_depth",
            "cal_date",
            "gmv",
            "total_gmv",
            "share_in_total",
            "share_delta",
            "row_missing_in_source",
        ]
    ].reset_index(drop=True)


def build_missing_zero_report(panel_df: pd.DataFrame, dates: Sequence[int], current_cal_date: int) -> pd.DataFrame:
    """Сформировать отчёт по пропускам и нулевому GMV.

    Args:
        panel_df: Полная недельная панель.
        dates: Список недель.
        current_cal_date: Текущая неделя.

    Returns:
        Таблица сегментов с пропусками или нулевыми неделями.

    Raises:
        ValueError: Не выбрасывается.

    Examples:
        >>> # missing_df = build_missing_zero_report(panel, dates, dates[-1])
    """

    previous_cal_date = int(dates[list(dates).index(int(current_cal_date)) - 1])
    grouped = panel_df.groupby(["segment_id", "segment_key", "segment_level", "slice_depth"], as_index=False).agg(
        missing_week_count=("row_missing_in_source", "sum"),
        zero_week_count=("gmv", lambda s: int((s == 0).sum())),
        nonzero_week_count=("gmv", lambda s: int((s > 0).sum())),
    )
    current = panel_df[panel_df["cal_date"].astype(int) == int(current_cal_date)][["segment_id", "gmv", "row_missing_in_source"]]
    previous = panel_df[panel_df["cal_date"].astype(int) == previous_cal_date][["segment_id", "gmv", "row_missing_in_source"]]
    current = current.rename(columns={"gmv": "gmv_current", "row_missing_in_source": "current_row_missing"})
    previous = previous.rename(columns={"gmv": "gmv_previous", "row_missing_in_source": "previous_row_missing"})
    report = grouped.merge(previous, how="left", on="segment_id").merge(current, how="left", on="segment_id")
    report = report[(report["missing_week_count"] > 0) | (report["zero_week_count"] > 0)].copy()
    return report.sort_values(["missing_week_count", "zero_week_count", "segment_key"], ascending=[False, False, True]).reset_index(drop=True)


def build_control_table(
    history_df: pd.DataFrame,
    panel_df: pd.DataFrame,
    candidates: pd.DataFrame,
    final_df: pd.DataFrame,
    coverage: Dict[str, frozenset[str]],
    dates: Sequence[int],
    current_cal_date: int,
    total_by_date: pd.Series,
) -> pd.DataFrame:
    """Сформировать контрольные показатели результата.

    Args:
        history_df: Исходная очищенная таблица.
        panel_df: Полная недельная панель.
        candidates: Диагностика кандидатов.
        final_df: Итоговые выбранные аномалии.
        coverage: Покрытие кандидатов атомами.
        dates: Список недель.
        current_cal_date: Текущая неделя.
        total_by_date: Total GMV по неделям.

    Returns:
        Таблица контрольных показателей.

    Raises:
        ValueError: Не выбрасывается.

    Examples:
        >>> # control = build_control_table(history_df, panel, candidates, final_df, coverage, dates, dates[-1], total)
    """

    selected_atoms: List[str] = []
    if not final_df.empty and "selected_atomic_descendants" in final_df.columns:
        for value in final_df["selected_atomic_descendants"].fillna("").astype(str):
            selected_atoms.extend(
                atom_id.strip()
                for atom_id in value.split(" || ")
                if atom_id.strip()
            )
    else:
        for segment_id in final_df["segment_id"].astype(str).tolist() if not final_df.empty else []:
            selected_atoms.extend(list(coverage.get(segment_id, frozenset())))
    selected_atom_unique_count = len(set(selected_atoms))
    double_count_violation_count = len(selected_atoms) - selected_atom_unique_count
    max_depth = int(candidates["slice_depth"].max()) if not candidates.empty else 0
    atomic_count = int((candidates["slice_depth"].astype(int) == max_depth).sum()) if not candidates.empty else 0

    rows = [
        ("source_rows", len(history_df)),
        ("panel_rows", len(panel_df)),
        ("week_count", len(dates)),
        ("previous_cal_date", int(dates[list(dates).index(int(current_cal_date)) - 1])),
        ("current_cal_date", int(current_cal_date)),
        ("current_total_gmv", float(total_by_date.loc[current_cal_date])),
        ("candidate_count", len(candidates)),
        ("eligible_candidate_count", int(candidates["is_eligible"].astype(bool).sum()) if "is_eligible" in candidates else 0),
        ("selected_count", len(final_df)),
        (
            "set_packing_unresolved_count",
            int((candidates["is_eligible"].astype(bool) & ~candidates["is_resolved"].astype(bool)).sum())
            if {"is_eligible", "is_resolved"}.issubset(candidates.columns)
            else 0,
        ),
        (
            "set_packing_not_proven_count",
            int(candidates.get("set_packing_status", pd.Series("", index=candidates.index)).astype(str).eq("SET_PACKING_NOT_PROVEN").sum())
            if not candidates.empty and "set_packing_status" in candidates
            else 0,
        ),
        ("atomic_count", atomic_count),
        ("selected_atomic_unique_count", selected_atom_unique_count),
        ("double_count_violation_count", double_count_violation_count),
        (
            "set_packing_global_status",
            str(candidates.get("set_packing_global_status", pd.Series([""], index=[0])).dropna().astype(str).replace("", pd.NA).dropna().iloc[0])
            if "set_packing_global_status" in candidates and not candidates["set_packing_global_status"].dropna().astype(str).replace("", pd.NA).dropna().empty
            else "",
        ),
        (
            "set_packing_component_count",
            int(candidates.get("set_packing_component_id", pd.Series("", index=candidates.index)).astype(str).replace("", pd.NA).dropna().nunique())
            if not candidates.empty and "set_packing_component_id" in candidates
            else 0,
        ),
        (
            "set_packing_non_optimal_component_count",
            int(
                candidates.loc[
                    candidates.get("set_packing_component_id", pd.Series("", index=candidates.index)).astype(str).ne("")
                    & candidates.get("set_packing_solver_status", pd.Series("", index=candidates.index)).astype(str).ne("OPTIMAL"),
                    "set_packing_component_id",
                ].astype(str).nunique()
            )
            if not candidates.empty and {"set_packing_component_id", "set_packing_solver_status"}.issubset(candidates.columns)
            else 0,
        ),
        (
            "set_packing_objective_value",
            float(final_df["selection_score"].astype(float).sum()) if not final_df.empty and "selection_score" in final_df else 0.0,
        ),
        (
            "set_packing_conflict_pair_count",
            int(candidates["conflict_count"].astype(int).sum() // 2) if "conflict_count" in candidates else 0,
        ),
        ("filled_missing_rows", int(panel_df["row_missing_in_source"].sum())),
    ]
    return pd.DataFrame(rows, columns=["показатель", "значение"])


def build_anomaly_analysis_sheet(candidates: pd.DataFrame, final_df: pd.DataFrame, thresholds: AnomalyThresholds) -> pd.DataFrame:
    """Сформировать лист анализа аномалий, прошедших первичный фильтр.

    Args:
        candidates: Диагностика кандидатов после оптимизационного отбора.
        final_df: Итоговые выбранные аномалии.
        thresholds: Пороги алгоритма, включая лимит фактов менеджерского вывода.

    Returns:
        Таблица аномалий, прошедших первичный фильтр, с причиной отсутствия в менеджерском выводе.

    Raises:
        ValueError: Не выбрасывается.

    Examples:
        >>> # analysis_df = build_anomaly_analysis_sheet(candidates, final_df, AnomalyThresholds())
    """

    columns = [
        "сегмент",
        "глубина",
        "z_scope",
        "base_anomaly_score",
        "depth_score_weight",
        "local_max_eligible_depth",
        "local_depth_gap",
        "eligible_descendant_count",
        "anomaly_score",
        "выбран",
        "разрешён",
        "set_packing_status",
        "set_packing_reason",
        "set_packing_global_status",
        "set_packing_component_id",
        "set_packing_solver",
        "set_packing_solver_status",
        "set_packing_abs_gap",
        "set_packing_rel_gap",
        "conflict_count",
        "conflict_segment_keys",
        "selected_atomic_count",
        "номер добавления в менеджерский вывод",
        "причина не попадания в менеджерский вывод",
        "Delta GMV",
    ]
    if candidates.empty or "passes_initial_anomaly_filter" not in candidates.columns:
        return pd.DataFrame(columns=columns)

    initial_anomaly_mask = (
        candidates["passes_initial_anomaly_filter"].eq(True)
        & candidates["abs_abnormal_gmv"].astype(float).ge(thresholds.min_anomaly_abs)
    )
    analysis_mask = initial_anomaly_mask
    analysis = candidates[analysis_mask].copy()
    if analysis.empty:
        return pd.DataFrame(columns=columns)

    selected_ids = set(final_df["segment_id"].astype(str).tolist()) if not final_df.empty else set()
    manager_ids = (
        set(final_df.head(thresholds.max_manager_facts)["segment_id"].astype(str).tolist())
        if not final_df.empty and "segment_id" in final_df.columns
        else set()
    )
    manager_order_by_id = (
        {
            str(segment_id): order
            for order, segment_id in enumerate(
                final_df.head(thresholds.max_manager_facts)["segment_id"].astype(str).tolist(),
                start=1,
            )
        }
        if not final_df.empty and "segment_id" in final_df.columns
        else {}
    )

    def manager_exclusion_reason(row: pd.Series) -> str:
        """Вернуть причину отсутствия кандидата в менеджерском выводе.

        Args:
            row: Строка кандидата из диагностики.

        Returns:
            Текст причины для листа анализа аномалий.

        Raises:
            ValueError: Не выбрасывается.

        Examples:
            >>> manager_exclusion_reason(pd.Series({'segment_id': 'x'}))
            'не выбран в итоговый набор аномалий'
        """

        segment_id = str(row.get("segment_id", ""))
        if segment_id in manager_ids:
            return "попал в итоговый менеджерский вывод"
        if segment_id in selected_ids:
            return f"выбран в итоговый набор, но не попал в топ-{thresholds.max_manager_facts} менеджерского вывода"
        selection_reason = str(row.get("selection_exclusion_reason", "")).strip()
        if selection_reason:
            return selection_reason
        classifier_reason = str(row.get("reason", "")).strip()
        return classifier_reason if classifier_reason else "не выбран в итоговый набор аномалий"

    analysis["причина не попадания в менеджерский вывод"] = analysis.apply(manager_exclusion_reason, axis=1)
    export = pd.DataFrame(
        {
            "сегмент": analysis["segment_key"].astype(str),
            "глубина": analysis["slice_depth"].astype(int),
            "z_scope": analysis["robust_z_capped"].astype(float).round(2),
            "base_anomaly_score": pd.to_numeric(
                analysis.get("base_anomaly_score", pd.Series(math.nan, index=analysis.index)),
                errors="coerce",
            ),
            "depth_score_weight": pd.to_numeric(
                analysis.get("depth_score_weight", pd.Series(math.nan, index=analysis.index)),
                errors="coerce",
            ),
            "local_max_eligible_depth": pd.to_numeric(
                analysis.get("local_max_eligible_depth", pd.Series(pd.NA, index=analysis.index)),
                errors="coerce",
            ).astype("Int64"),
            "local_depth_gap": pd.to_numeric(
                analysis.get("local_depth_gap", pd.Series(pd.NA, index=analysis.index)),
                errors="coerce",
            ).astype("Int64"),
            "eligible_descendant_count": pd.to_numeric(
                analysis.get("eligible_descendant_count", pd.Series(pd.NA, index=analysis.index)),
                errors="coerce",
            ).astype("Int64"),
            "anomaly_score": analysis["anomaly_score"].astype(float),
            "выбран": analysis.get(
                "selected", pd.Series(False, index=analysis.index)
            ).astype(bool),
            "разрешён": analysis.get(
                "is_resolved", pd.Series(False, index=analysis.index)
            ).astype(bool),
            "set_packing_status": analysis.get(
                "set_packing_status", pd.Series("", index=analysis.index)
            ).astype(str),
            "set_packing_reason": analysis.get(
                "set_packing_reason", pd.Series("", index=analysis.index)
            ).astype(str),
            "set_packing_global_status": analysis.get(
                "set_packing_global_status", pd.Series("", index=analysis.index)
            ).astype(str),
            "set_packing_component_id": analysis.get(
                "set_packing_component_id", pd.Series("", index=analysis.index)
            ).astype(str),
            "set_packing_solver": analysis.get(
                "set_packing_solver", pd.Series("", index=analysis.index)
            ).astype(str),
            "set_packing_solver_status": analysis.get(
                "set_packing_solver_status", pd.Series("", index=analysis.index)
            ).astype(str),
            "set_packing_abs_gap": pd.to_numeric(
                analysis.get("set_packing_abs_gap", pd.Series(math.nan, index=analysis.index)),
                errors="coerce",
            ),
            "set_packing_rel_gap": pd.to_numeric(
                analysis.get("set_packing_rel_gap", pd.Series(math.nan, index=analysis.index)),
                errors="coerce",
            ),
            "conflict_count": pd.to_numeric(
                analysis.get("conflict_count", pd.Series(0, index=analysis.index)),
                errors="coerce",
            ).fillna(0).astype(int),
            "conflict_segment_keys": analysis.get(
                "conflict_segment_keys", pd.Series("", index=analysis.index)
            ).astype(str),
            "selected_atomic_count": pd.to_numeric(
                analysis.get("selected_atomic_count", pd.Series(0, index=analysis.index)),
                errors="coerce",
            ).fillna(0).astype(int),
            "номер добавления в менеджерский вывод": analysis["segment_id"].astype(str).map(manager_order_by_id).astype("Int64"),
            "причина не попадания в менеджерский вывод": analysis["причина не попадания в менеджерский вывод"],
            "Delta GMV": analysis["wow_delta_gmv"].astype(float).round(0).astype("Int64"),
        }
    )
    return export.sort_values(
        by=["глубина", "anomaly_score", "Delta GMV", "сегмент"],
        ascending=[False, False, False, True],
    ).reset_index(drop=True)


def highlight_manager_rows_on_anomaly_analysis(worksheet) -> None:
    """Выделить зелёным строки листа анализа аномалий, попавшие в менеджерский вывод.

    Args:
        worksheet: Лист openpyxl `Анализ аномалий`.

    Returns:
        None.

    Raises:
        AttributeError: Если объект не похож на лист openpyxl.

    Examples:
        >>> # highlight_manager_rows_on_anomaly_analysis(worksheet)
    """

    if worksheet.max_row < 2:
        return

    target_header = "причина не попадания в менеджерский вывод"
    target_value = "попал в итоговый менеджерский вывод"
    green_fill = PatternFill(fill_type="solid", fgColor="C6EFCE")
    header_to_col = {
        str(cell.value): cell.column
        for cell in worksheet[1]
        if cell.value is not None
    }
    reason_col = header_to_col.get(target_header)
    if reason_col is None:
        return

    for row_idx in range(2, worksheet.max_row + 1):
        if worksheet.cell(row=row_idx, column=reason_col).value != target_value:
            continue
        for col_idx in range(1, worksheet.max_column + 1):
            worksheet.cell(row=row_idx, column=col_idx).fill = green_fill


def _order_tree_layers_by_shared_children(
    node_ids_by_depth: Dict[int, List[str]],
    edges: Sequence[Tuple[str, str]],
    records: Dict[str, Dict[str, object]],
    sweep_count: int = 5,
) -> Dict[int, List[str]]:
    """Упорядочить уровни дерева с группировкой родителей общих детей.

    Сначала выполняется top-down barycenter-проход для сближения детей с
    родителями. Затем bottom-up проход объединяет в непрерывные компоненты
    родителей, связанных хотя бы с одним общим ребёнком. Последний проход
    всегда bottom-up, поэтому требование близости общих родителей имеет
    приоритет над косметическим выравниванием детей.

    Args:
        node_ids_by_depth: Узлы, сгруппированные по глубине.
        edges: Рёбра ``(parent_id, child_id)`` соседних глубин.
        records: Атрибуты узлов, включая человекочитаемый сегмент.
        sweep_count: Число двунаправленных проходов минимизации пересечений.

    Returns:
        Упорядоченные идентификаторы узлов для каждой глубины.

    Raises:
        ValueError: Если sweep_count меньше единицы.

    Examples:
        >>> records = {'p1': {'сегмент': 'p1'}, 'p2': {'сегмент': 'p2'}, 'c': {'сегмент': 'c'}}
        >>> _order_tree_layers_by_shared_children({1: ['p2', 'p1'], 2: ['c']}, [('p1', 'c'), ('p2', 'c')], records)[1]
        ['p1', 'p2']
    """

    if sweep_count < 1:
        raise ValueError("sweep_count должен быть не меньше единицы")

    depth_values = sorted(node_ids_by_depth)
    ordered = {
        depth: sorted(
            node_ids,
            key=lambda node_id: str(records[node_id]["сегмент"]),
        )
        for depth, node_ids in node_ids_by_depth.items()
    }
    children_by_parent: Dict[str, List[str]] = {
        node_id: []
        for node_ids in node_ids_by_depth.values()
        for node_id in node_ids
    }
    parents_by_child: Dict[str, List[str]] = {
        node_id: []
        for node_ids in node_ids_by_depth.values()
        for node_id in node_ids
    }
    for parent_id, child_id in edges:
        children_by_parent[parent_id].append(child_id)
        parents_by_child[child_id].append(parent_id)

    def top_down_sweep() -> None:
        """Сблизить детей со средними позициями их родителей.

        Args:
            Аргументы отсутствуют.

        Returns:
            None.

        Raises:
            ValueError: Не выбрасывается.

        Examples:
            >>> # top_down_sweep()
        """

        for depth in depth_values[1:]:
            parent_depth = int(depth) - 1
            if parent_depth not in ordered:
                continue
            parent_positions = {
                node_id: position
                for position, node_id in enumerate(ordered[parent_depth])
            }
            sort_keys: Dict[str, Tuple[float, Tuple[int, ...], str]] = {}
            for node_id in ordered[int(depth)]:
                positions = sorted(
                    parent_positions[parent_id]
                    for parent_id in parents_by_child[node_id]
                    if parent_id in parent_positions
                )
                barycenter = (
                    sum(positions) / len(positions)
                    if positions
                    else float("inf")
                )
                sort_keys[node_id] = (
                    barycenter,
                    tuple(positions),
                    str(records[node_id]["сегмент"]),
                )
            ordered[int(depth)] = sorted(
                ordered[int(depth)],
                key=lambda node_id: sort_keys[node_id],
            )

    def bottom_up_shared_child_sweep() -> None:
        """Сделать компоненты родителей общих детей непрерывными.

        Args:
            Аргументы отсутствуют.

        Returns:
            None.

        Raises:
            ValueError: Не выбрасывается.

        Examples:
            >>> # bottom_up_shared_child_sweep()
        """

        for depth in reversed(depth_values[:-1]):
            child_depth = int(depth) + 1
            if child_depth not in ordered:
                continue
            child_positions = {
                node_id: position
                for position, node_id in enumerate(ordered[child_depth])
            }
            parent_ids = list(ordered[int(depth)])
            parent_id_set = set(parent_ids)
            shared_child_neighbors = {
                parent_id: set()
                for parent_id in parent_ids
            }
            for child_id in ordered[child_depth]:
                linked_parents = [
                    parent_id
                    for parent_id in parents_by_child[child_id]
                    if parent_id in parent_id_set
                ]
                for parent_id in linked_parents:
                    shared_child_neighbors[parent_id].update(
                        other_parent_id
                        for other_parent_id in linked_parents
                        if other_parent_id != parent_id
                    )

            components: List[List[str]] = []
            remaining = set(parent_ids)
            for seed_id in parent_ids:
                if seed_id not in remaining:
                    continue
                stack = [seed_id]
                component: List[str] = []
                remaining.remove(seed_id)
                while stack:
                    parent_id = stack.pop()
                    component.append(parent_id)
                    for neighbor_id in sorted(shared_child_neighbors[parent_id]):
                        if neighbor_id in remaining:
                            remaining.remove(neighbor_id)
                            stack.append(neighbor_id)
                components.append(component)

            parent_sort_keys: Dict[str, Tuple[float, Tuple[int, ...], str]] = {}
            for parent_id in parent_ids:
                positions = sorted(
                    child_positions[child_id]
                    for child_id in children_by_parent[parent_id]
                    if child_id in child_positions
                )
                barycenter = (
                    sum(positions) / len(positions)
                    if positions
                    else float("inf")
                )
                parent_sort_keys[parent_id] = (
                    barycenter,
                    tuple(positions),
                    str(records[parent_id]["сегмент"]),
                )

            component_rows: List[Tuple[Tuple[float, float, str], List[str]]] = []
            for component in components:
                component_order = sorted(
                    component,
                    key=lambda node_id: parent_sort_keys[node_id],
                )
                finite_barycenters = [
                    parent_sort_keys[node_id][0]
                    for node_id in component_order
                    if not math.isinf(parent_sort_keys[node_id][0])
                ]
                component_key = (
                    min(finite_barycenters) if finite_barycenters else float("inf"),
                    (
                        sum(finite_barycenters) / len(finite_barycenters)
                        if finite_barycenters
                        else float("inf")
                    ),
                    min(str(records[node_id]["сегмент"]) for node_id in component_order),
                )
                component_rows.append((component_key, component_order))

            component_rows.sort(key=lambda row: row[0])
            ordered[int(depth)] = [
                node_id
                for _, component in component_rows
                for node_id in component
            ]

    def refine_shared_child_adjacency() -> None:
        """Локально минимизировать разрывы между родителями общего ребёнка.

        Перестановки оцениваются лексикографически: сначала уменьшается число
        разорванных родительских групп, затем суммарный и максимальный разрыв.

        Args:
            Аргументы отсутствуют.

        Returns:
            None.

        Raises:
            ValueError: Не выбрасывается.

        Examples:
            >>> # refine_shared_child_adjacency()
        """

        for depth in depth_values[:-1]:
            child_depth = int(depth) + 1
            if child_depth not in ordered:
                continue
            parent_groups = [
                [
                    parent_id
                    for parent_id in parents_by_child[child_id]
                    if parent_id in ordered[int(depth)]
                ]
                for child_id in ordered[child_depth]
            ]
            parent_groups = [
                parent_ids
                for parent_ids in parent_groups
                if len(parent_ids) > 1
            ]
            if not parent_groups:
                continue

            current_order = list(ordered[int(depth)])

            def gap_objective(candidate_order: Sequence[str]) -> Tuple[int, int, int]:
                """Оценить разрывы родительских групп для одного уровня.

                Args:
                    candidate_order: Проверяемый порядок родителей.

                Returns:
                    Число разорванных групп, сумма разрывов и максимальный разрыв.

                Raises:
                    KeyError: Если родитель отсутствует в порядке.

                Examples:
                    >>> # gap_objective(['p1', 'p2'])
                """

                positions = {
                    node_id: position
                    for position, node_id in enumerate(candidate_order)
                }
                gaps = [
                    max(positions[parent_id] for parent_id in parent_ids)
                    - min(positions[parent_id] for parent_id in parent_ids)
                    + 1
                    - len(parent_ids)
                    for parent_ids in parent_groups
                ]
                return (
                    sum(gap > 0 for gap in gaps),
                    sum(gaps),
                    max(gaps, default=0),
                )

            while True:
                best_objective = gap_objective(current_order)
                best_order: Optional[List[str]] = None
                for left_index in range(len(current_order)):
                    for right_index in range(left_index + 1, len(current_order)):
                        candidate_order = list(current_order)
                        candidate_order[left_index], candidate_order[right_index] = (
                            candidate_order[right_index],
                            candidate_order[left_index],
                        )
                        candidate_objective = gap_objective(candidate_order)
                        if candidate_objective < best_objective:
                            best_objective = candidate_objective
                            best_order = candidate_order
                if best_order is None:
                    break
                current_order = best_order
            ordered[int(depth)] = current_order

    for _ in range(sweep_count):
        top_down_sweep()
        bottom_up_shared_child_sweep()
    bottom_up_shared_child_sweep()
    refine_shared_child_adjacency()
    return ordered


def _tree_edge_port_offset(edge_index: int, edge_count: int, node_width: float) -> float:
    """Рассчитать отдельный горизонтальный порт ребра на границе узла.

    Args:
        edge_index: Индекс ребра среди рёбер узла, начиная с нуля.
        edge_count: Общее число рёбер на соответствующей стороне узла.
        node_width: Ширина узла в координатах графика.

    Returns:
        Смещение порта относительно центра узла.

    Raises:
        ValueError: Если индекс или количество рёбер некорректны.

    Examples:
        >>> round(_tree_edge_port_offset(0, 2, 3.0), 2)
        -0.54
        >>> round(_tree_edge_port_offset(1, 2, 3.0), 2)
        0.54
    """

    if edge_count < 1 or edge_index < 0 or edge_index >= edge_count:
        raise ValueError("Некорректный индекс порта ребра")
    usable_width = node_width * 0.72
    return usable_width * ((edge_index + 0.5) / edge_count - 0.5)


def build_anomaly_tree_from_excel(
    report_path: str | Path,
    output_path: str | Path | None = None,
    sheet_name: str = "Анализ аномалий",
    dpi: int = 160,
) -> Path:
    """Построить дерево аномальных сегментов по листу итогового Excel-файла.

    Узлы группируются по глубине. Ребро проводится только между соседними
    глубинами ``d`` и ``d + 1``, если признаки родителя являются строгим
    подмножеством признаков ребёнка. Один ребёнок может иметь несколько
    родителей, поэтому визуализация фактически является ориентированным DAG.
    Родители общих детей образуют непрерывные группы. Каждому ребру выделяются
    отдельные порты и отдельная трасса; цвет ребра кодирует дочерний сегмент.

    Args:
        report_path: Путь к Excel-отчёту с листом анализа аномалий.
        output_path: Путь к PNG, SVG или PDF. Если None, рядом с отчётом
            создаётся файл ``<имя_отчёта>_tree.png``.
        sheet_name: Имя листа-источника.
        dpi: Разрешение растрового изображения.

    Returns:
        Путь к созданному файлу визуализации.

    Raises:
        FileNotFoundError: Если Excel-отчёт отсутствует.
        ImportError: Если matplotlib недоступен.
        ValueError: Если лист пуст, не содержит обязательных колонок или
            указан неподдерживаемый формат результата.
        OSError: Если изображение невозможно записать.

    Examples:
        >>> # build_anomaly_tree_from_excel('gmv_anomaly_report_1.xlsx', 'gmv_tree.png')
    """

    try:
        import matplotlib

        matplotlib.use("Agg")
        import matplotlib.pyplot as plt
        from matplotlib.lines import Line2D
        from matplotlib.path import Path as MatplotlibPath
        from matplotlib.patches import FancyArrowPatch, FancyBboxPatch, Patch, Rectangle
    except ImportError as exc:
        raise ImportError("Для построения дерева требуется matplotlib") from exc

    report = Path(report_path)
    if not report.exists():
        raise FileNotFoundError(f"Excel-отчёт не найден: {report}")

    tree_path = (
        Path(output_path)
        if output_path is not None
        else report.with_name(f"{report.stem}_tree.png")
    )
    if not tree_path.suffix:
        tree_path = tree_path.with_suffix(".png")
    if tree_path.suffix.lower() not in {".png", ".svg", ".pdf"}:
        raise ValueError("Дерево можно сохранить только в PNG, SVG или PDF")
    if dpi <= 0:
        raise ValueError("dpi должен быть положительным целым числом")

    analysis = pd.read_excel(report, sheet_name=sheet_name)
    required_columns = {"сегмент", "глубина", "z_scope", "anomaly_score", "Delta GMV"}
    missing_columns = sorted(required_columns - set(analysis.columns))
    if missing_columns:
        raise ValueError(
            f"На листе {sheet_name!r} отсутствуют обязательные колонки: {missing_columns}"
        )

    nodes = analysis.copy()
    optional_tree_columns = [
        "set_packing_status",
        "set_packing_reason",
        "set_packing_component_id",
        "set_packing_solver_status",
        "conflict_count",
        "номер добавления в менеджерский вывод",
    ]
    for column in optional_tree_columns:
        if column not in nodes.columns:
            nodes[column] = ""
    nodes["сегмент"] = nodes["сегмент"].astype(str).str.strip()
    nodes["глубина"] = pd.to_numeric(nodes["глубина"], errors="coerce")
    nodes["z_scope"] = pd.to_numeric(nodes["z_scope"], errors="coerce")
    nodes["anomaly_score"] = pd.to_numeric(nodes["anomaly_score"], errors="coerce")
    nodes["Delta GMV"] = pd.to_numeric(nodes["Delta GMV"], errors="coerce")
    nodes["номер добавления в менеджерский вывод"] = pd.to_numeric(
        nodes["номер добавления в менеджерский вывод"],
        errors="coerce",
    )
    nodes["conflict_count"] = pd.to_numeric(nodes["conflict_count"], errors="coerce").fillna(0).astype(int)
    nodes = nodes[
        nodes["глубина"].notna()
        & nodes["глубина"].gt(0)
        & nodes["сегмент"].ne("")
        & nodes["сегмент"].ne("nan")
    ].copy()
    if nodes.empty:
        raise ValueError(f"Лист {sheet_name!r} не содержит сегментов глубины выше нуля")

    nodes["глубина"] = nodes["глубина"].astype(int)
    nodes = nodes.drop_duplicates(subset=["глубина", "сегмент"], keep="first")
    nodes = nodes.sort_values(["глубина", "сегмент"], kind="stable").reset_index(drop=True)
    nodes["node_id"] = [
        f"depth={int(depth)}|segment={segment}"
        for depth, segment in zip(nodes["глубина"], nodes["сегмент"])
    ]
    nodes["parts"] = nodes["сегмент"].map(_segment_key_parts)
    nodes["feature_set"] = nodes["parts"].map(frozenset)

    def clean_tree_value(value: object) -> str:
        """Вернуть безопасный текст для подписи узла дерева.

        Args:
            value: Исходное значение из Excel.

        Returns:
            Строка без `nan` и лишних пробелов.

        Raises:
            ValueError: Не выбрасывается.

        Examples:
            >>> clean_tree_value(float('nan'))
            ''
        """

        if value is None or pd.isna(value):
            return ""
        text = str(value).strip()
        return "" if text.lower() in {"", "nan", "<na>", "none"} else text

    def build_tree_detail_lines(row: pd.Series) -> List[str]:
        """Собрать дополнительные строки для не-менеджерского узла.

        Args:
            row: Строка узла из листа анализа аномалий.

        Returns:
            Список строк со статусом оптимизационного выбора.

        Raises:
            ValueError: Не выбрасывается.

        Examples:
            >>> build_tree_detail_lines(pd.Series({'номер добавления в менеджерский вывод': 1}))
            []
        """

        if not pd.isna(row.get("номер добавления в менеджерский вывод")):
            return []
        lines: List[str] = []
        status = clean_tree_value(row.get("set_packing_status"))
        reason = clean_tree_value(row.get("set_packing_reason"))
        component_id = clean_tree_value(row.get("set_packing_component_id"))
        solver_status = clean_tree_value(row.get("set_packing_solver_status"))
        conflict_count = int(row.get("conflict_count", 0) or 0)
        if status:
            lines.append(status)
        if component_id:
            lines.append(f"component={component_id}")
        if solver_status:
            lines.append(f"solver={solver_status}")
        if conflict_count:
            lines.append(f"conflicts={conflict_count}")
        if reason:
            lines.append(reason[:60].rstrip() + ("…" if len(reason) > 60 else ""))
        return lines[:5]

    nodes["is_manager_output"] = nodes["номер добавления в менеджерский вывод"].notna()
    nodes["tree_detail_lines"] = nodes.apply(build_tree_detail_lines, axis=1)

    records = {
        str(row["node_id"]): row.to_dict()
        for _, row in nodes.iterrows()
    }
    depth_values = sorted(nodes["глубина"].unique().tolist())
    node_ids_by_depth = {
        int(depth): nodes.loc[nodes["глубина"].eq(depth), "node_id"].astype(str).tolist()
        for depth in depth_values
    }

    edges: List[Tuple[str, str]] = []
    children_by_parent: Dict[str, List[str]] = {
        node_id: [] for node_id in records
    }
    parents_by_child: Dict[str, List[str]] = {
        node_id: [] for node_id in records
    }
    for depth in depth_values:
        parent_ids = node_ids_by_depth[int(depth)]
        child_ids = node_ids_by_depth.get(int(depth) + 1, [])
        for parent_id in parent_ids:
            parent_features = records[parent_id]["feature_set"]
            for child_id in child_ids:
                if parent_features < records[child_id]["feature_set"]:
                    edges.append((parent_id, child_id))
                    children_by_parent[parent_id].append(child_id)
                    parents_by_child[child_id].append(parent_id)

    ordered_by_depth = _order_tree_layers_by_shared_children(
        node_ids_by_depth,
        edges,
        records,
    )

    node_width = 3.15
    horizontal_gap = 0.22
    cluster_padding_x = 0.28
    cluster_padding_y = 0.24
    max_nodes_in_layer = max(len(node_ids) for node_ids in ordered_by_depth.values())
    total_width = max_nodes_in_layer * node_width + max(0, max_nodes_in_layer - 1) * horizontal_gap
    node_heights = {
        depth: 1.24
        + 0.24
        * max(
            len(records[node_id]["parts"])
            for node_id in node_ids
        )
        + 0.18
        * max(
            len(records[node_id].get("tree_detail_lines", []))
            for node_id in node_ids
        )
        for depth, node_ids in ordered_by_depth.items()
    }
    edges_by_parent_depth = {
        int(depth): [
            edge
            for edge in edges
            if int(records[edge[0]]["глубина"]) == int(depth)
        ]
        for depth in depth_values
    }
    gap_after_depth = {
        int(depth): max(
            1.05,
            min(3.0, 0.07 * len(edges_by_parent_depth[int(depth)]) + 0.72),
        )
        for depth in depth_values
    }
    total_height = sum(
        node_heights[depth] + 2 * cluster_padding_y
        for depth in depth_values
    ) + sum(
        gap_after_depth[int(depth)]
        for depth in depth_values[:-1]
    )

    positions: Dict[str, Tuple[float, float]] = {}
    layer_bounds: Dict[int, Tuple[float, float, float, float]] = {}
    current_top = total_height
    for depth_index, depth in enumerate(depth_values):
        node_ids = ordered_by_depth[int(depth)]
        node_height = node_heights[int(depth)]
        cluster_height = node_height + 2 * cluster_padding_y
        center_y = current_top - cluster_height / 2
        row_width = len(node_ids) * node_width + max(0, len(node_ids) - 1) * horizontal_gap
        row_left = (total_width - row_width) / 2
        for position, node_id in enumerate(node_ids):
            center_x = row_left + node_width / 2 + position * (node_width + horizontal_gap)
            positions[node_id] = (center_x, center_y)
        layer_bounds[int(depth)] = (
            row_left - cluster_padding_x,
            center_y - cluster_height / 2,
            row_width + 2 * cluster_padding_x,
            cluster_height,
        )
        current_top -= cluster_height
        if depth_index < len(depth_values) - 1:
            current_top -= gap_after_depth[int(depth)]

    figure_width = max(14.0, min(42.0, total_width * 0.78))
    figure_height = max(8.0, total_height * 1.15 + 2.0)
    figure, axis = plt.subplots(figsize=(figure_width, figure_height))
    axis.set_xlim(-0.55, total_width + 0.55)
    axis.set_ylim(-1.15, total_height + 1.15)
    axis.axis("off")

    for depth in depth_values:
        left, bottom, width, height = layer_bounds[int(depth)]
        axis.add_patch(
            FancyBboxPatch(
                (left, bottom),
                width,
                height,
                boxstyle="round,pad=0.04,rounding_size=0.22",
                linewidth=0.9,
                edgecolor="#c7daf7",
                facecolor="#f8fbff",
                zorder=0,
            )
        )
        axis.text(
            left + 0.14,
            bottom + height - 0.12,
            f"Сегменты глубины {int(depth)}",
            color="#315a9f",
            fontsize=9,
            ha="left",
            va="top",
            zorder=4,
        )

    sorted_children_by_parent = {
        parent_id: sorted(
            child_ids,
            key=lambda child_id: (positions[child_id][0], child_id),
        )
        for parent_id, child_ids in children_by_parent.items()
    }
    sorted_parents_by_child = {
        child_id: sorted(
            parent_ids,
            key=lambda parent_id: (positions[parent_id][0], parent_id),
        )
        for child_id, parent_ids in parents_by_child.items()
    }
    edge_colors = [
        "#2563eb",
        "#7c3aed",
        "#0891b2",
        "#db2777",
        "#4f46e5",
        "#0f766e",
        "#c2410c",
        "#0369a1",
    ]
    ordered_child_ids = sorted(
        {child_id for _, child_id in edges},
        key=lambda child_id: (
            int(records[child_id]["глубина"]),
            positions[child_id][0],
            child_id,
        ),
    )
    edge_color_by_child = {
        child_id: edge_colors[index % len(edge_colors)]
        for index, child_id in enumerate(ordered_child_ids)
    }

    for parent_depth in depth_values[:-1]:
        layer_edges = sorted(
            edges_by_parent_depth[int(parent_depth)],
            key=lambda edge: (
                positions[edge[0]][0],
                positions[edge[1]][0],
                edge[0],
                edge[1],
            ),
        )
        if not layer_edges:
            continue
        lane_count = len(layer_edges)
        for lane_index, (parent_id, child_id) in enumerate(layer_edges):
            parent_x, parent_y = positions[parent_id]
            child_x, child_y = positions[child_id]
            outgoing_ids = sorted_children_by_parent[parent_id]
            incoming_ids = sorted_parents_by_child[child_id]
            start_x = parent_x + _tree_edge_port_offset(
                outgoing_ids.index(child_id),
                len(outgoing_ids),
                node_width,
            )
            end_x = child_x + _tree_edge_port_offset(
                incoming_ids.index(parent_id),
                len(incoming_ids),
                node_width,
            )
            start_y = parent_y - node_heights[int(records[parent_id]["глубина"])] / 2
            end_y = child_y + node_heights[int(records[child_id]["глубина"])] / 2
            lane_top = start_y - 0.13
            lane_bottom = end_y + 0.13
            lane_y = lane_top - (lane_top - lane_bottom) * (lane_index + 1) / (lane_count + 1)
            path = MatplotlibPath(
                [
                    (start_x, start_y),
                    (start_x, lane_y),
                    (end_x, lane_y),
                    (end_x, end_y),
                ],
                [
                    MatplotlibPath.MOVETO,
                    MatplotlibPath.LINETO,
                    MatplotlibPath.LINETO,
                    MatplotlibPath.LINETO,
                ],
            )
            axis.add_patch(
                FancyArrowPatch(
                    path=path,
                    arrowstyle="-",
                    linewidth=2.2,
                    color="white",
                    alpha=0.96,
                    zorder=1,
                )
            )
            axis.add_patch(
                FancyArrowPatch(
                    path=path,
                    arrowstyle="-|>",
                    mutation_scale=6.5,
                    linewidth=0.82,
                    color=edge_color_by_child[child_id],
                    alpha=0.88,
                    zorder=1.1,
                )
            )

    connected_node_ids = {
        node_id
        for edge in edges
        for node_id in edge
    }
    sequence = 1
    for depth in depth_values:
        for node_id in ordered_by_depth[int(depth)]:
            node = records[node_id]
            center_x, center_y = positions[node_id]
            node_height = node_heights[int(depth)]
            delta_gmv = _safe_float(node["Delta GMV"])
            positive = delta_gmv >= 0
            facecolor = "#eff9f1" if positive else "#fff0f0"
            edgecolor = "#afd8b8" if positive else "#f0b2b2"
            accent = "#49a15d" if positive else "#df4d4d"
            isolated = node_id not in connected_node_ids
            if isolated:
                edgecolor = "#8b9bb2"
            is_manager_output = bool(node.get("is_manager_output", False))
            node_edgecolor = "#172554" if is_manager_output else edgecolor
            node_linewidth = 2.85 if is_manager_output else 1.0

            left = center_x - node_width / 2
            bottom = center_y - node_height / 2
            axis.add_patch(
                Rectangle(
                    (left, bottom),
                    node_width,
                    node_height,
                    linewidth=node_linewidth,
                    linestyle="-" if is_manager_output else ("--" if isolated else "-"),
                    edgecolor=node_edgecolor,
                    facecolor=facecolor,
                    zorder=2,
                )
            )
            badge_width = 0.52
            badge_height = 0.27
            badge_left = left + 0.13
            badge_bottom = bottom + node_height - 0.39
            axis.add_patch(
                Rectangle(
                    (badge_left, badge_bottom),
                    badge_width,
                    badge_height,
                    linewidth=0,
                    facecolor=accent,
                    zorder=3,
                )
            )
            axis.text(
                badge_left + badge_width / 2,
                badge_bottom + badge_height / 2,
                str(sequence),
                color="white",
                fontsize=6.5,
                fontweight="bold",
                ha="center",
                va="center",
                zorder=4,
            )
            axis.text(
                left + 0.83,
                badge_bottom + badge_height / 2,
                f"Глубина {int(depth)}",
                color="#263342",
                fontsize=6.2,
                fontweight="bold",
                ha="left",
                va="center",
                zorder=4,
            )

            feature_lines = [
                f"{dimension}={value}"
                for dimension, value in node["parts"]
            ]
            body_top = badge_bottom - 0.11
            axis.text(
                left + 0.14,
                body_top,
                "\n".join(feature_lines),
                color="#222222",
                fontsize=5.8,
                linespacing=1.35,
                ha="left",
                va="top",
                zorder=4,
            )
            detail_lines = [
                str(line)
                for line in node.get("tree_detail_lines", [])
                if str(line).strip()
            ]
            if detail_lines:
                detail_top = body_top - 0.215 * max(1, len(feature_lines)) - 0.07
                axis.text(
                    left + 0.14,
                    detail_top,
                    "\n".join(detail_lines),
                    color="#334155",
                    fontsize=4.75,
                    linespacing=1.22,
                    ha="left",
                    va="top",
                    zorder=4,
                )
            delta_y = bottom + 0.35
            axis.text(
                left + 0.14,
                delta_y,
                f"ΔGMV: {'+' if delta_gmv > 0 else ''}{_format_rub(delta_gmv)}",
                color="#171717",
                fontsize=6.3,
                fontweight="bold",
                ha="left",
                va="bottom",
                zorder=4,
            )
            z_value = _safe_float(node["z_scope"], math.nan)
            score_value = _safe_float(node["anomaly_score"], math.nan)
            z_text = "nan" if math.isnan(z_value) else f"{z_value:.2f}"
            score_text = "nan" if math.isnan(score_value) else f"{score_value:.3f}"
            axis.text(
                left + 0.14,
                bottom + 0.12,
                f"z={z_text} · score={score_text}",
                color="#7f8b99",
                fontsize=5.3,
                ha="left",
                va="bottom",
                zorder=4,
            )
            sequence += 1

    figure.suptitle(
        "Граф",
        fontsize=15,
        y=0.985,
    )
    legend_handles = [
        Patch(facecolor="#eff9f1", edgecolor="#afd8b8", label="Положительный ΔGMV"),
        Patch(facecolor="#fff0f0", edgecolor="#f0b2b2", label="Отрицательный ΔGMV"),
        Patch(facecolor="white", edgecolor="#172554", linewidth=2.4, label="Попал в менеджерский вывод"),
        Patch(facecolor="white", edgecolor="#8b9bb2", linestyle="--", label="Изолированный сегмент"),
        Line2D([0], [0], color="#2563eb", linewidth=1.4, label="Одинаковый цвет рёбер — один ребёнок"),
    ]
    figure.legend(
        handles=legend_handles,
        loc="lower left",
        bbox_to_anchor=(0.015, 0.01),
        frameon=False,
        ncol=5,
        fontsize=8,
    )
    tree_path.parent.mkdir(parents=True, exist_ok=True)
    try:
        figure.savefig(tree_path, dpi=dpi, bbox_inches="tight", facecolor="white")
    finally:
        plt.close(figure)
    return tree_path.resolve()


def write_anomaly_excel(
    output_path: str | Path,
    thresholds: AnomalyThresholds,
    dim_cols: Sequence[str],
    history_df: pd.DataFrame,
    panel_df: pd.DataFrame,
    candidates: pd.DataFrame,
    final_df: pd.DataFrame,
    total_by_date: pd.Series,
    dates: Sequence[int],
    current_cal_date: int,
    coverage: Dict[str, frozenset[str]],
    optimization_decision_log: pd.DataFrame,
) -> None:
    """Записать результат поиска аномалий в Excel.

    Args:
        output_path: Путь к итоговому Excel-файлу.
        thresholds: Пороги алгоритма.
        dim_cols: Список признаков.
        history_df: Исходная очищенная таблица.
        panel_df: Полная недельная панель.
        candidates: Диагностика кандидатов.
        final_df: Итоговые выбранные аномалии.
        total_by_date: Total GMV по неделям.
        dates: Список недель.
        current_cal_date: Текущая неделя.
        coverage: Покрытие кандидатов атомами.
        optimization_decision_log: Журнал построения и решения Set Packing.

    Returns:
        None.

    Raises:
        OSError: Если Excel-файл невозможно записать.

    Examples:
        >>> # write_anomaly_excel('gmv_anomaly_report.xlsx', thresholds, dims, history_df, panel, candidates, final_df, total, dates, dates[-1], coverage, log)
    """

    output_path = Path(output_path)
    params = pd.DataFrame(
        [
            ("признаки", " × ".join(dim_cols)),
            ("min_anomaly_abs", thresholds.min_anomaly_abs),
            ("min_z_score", thresholds.min_z_score),
            ("min_materiality_share", thresholds.min_materiality_share),
            ("sigma_floor", thresholds.sigma_floor),
            ("z_cap", thresholds.z_cap),
            ("set_packing_gap_tolerance", thresholds.set_packing_gap_tolerance),
            ("max_exact_fallback_size", thresholds.max_exact_fallback_size),
            ("current_cal_date", int(current_cal_date)),
        ],
        columns=["показатель", "значение"],
    )
    control = build_control_table(history_df, panel_df, candidates, final_df, coverage, dates, current_cal_date, total_by_date)
    params.insert(0, "раздел", "Параметры")
    control.insert(0, "раздел", "Контроль")
    params_and_control = pd.concat([params, control], ignore_index=True)

    manager_df = build_manager_summary(final_df, thresholds, float(total_by_date.loc[current_cal_date]), candidates)
    anomaly_analysis = build_anomaly_analysis_sheet(candidates, final_df, thresholds)
    history_top = build_history_for_selected(panel_df, final_df, total_by_date)
    missing_zero = build_missing_zero_report(panel_df, dates, current_cal_date)

    final_cols = [
        "rank",
        "output_block",
        "segment_key",
        "slice_depth",
        "gmv_current",
        "gmv_previous",
        "wow_delta_gmv",
        "relative_wow",
        "share_current",
        "share_delta_current",
        "baseline_relative_growth",
        "robust_z",
        "robust_z_capped",
        "materiality_share",
        "gross_atomic_movement",
        "base_anomaly_score",
        "depth_score_weight",
        "local_max_eligible_depth",
        "local_depth_gap",
        "eligible_descendant_count",
        "anomaly_score",
        "abnormal_gmv",
        "abs_abnormal_gmv",
        "selection_score",
        "reliability_factor",
        "history_nonzero_weeks",
        "state",
        "reason",
        "covered_atomic_count",
        "selected",
        "set_packing_global_status",
        "set_packing_component_id",
        "set_packing_solver",
        "set_packing_solver_status",
        "set_packing_reason",
        "set_packing_objective_value",
        "set_packing_best_bound",
        "set_packing_abs_gap",
        "set_packing_rel_gap",
        "set_packing_solve_time_sec",
        "set_packing_variable_count",
        "set_packing_constraint_count",
        "conflict_count",
        "conflict_segment_ids",
        "conflict_segment_keys",
        "atomic_coverage_source",
        "atomic_coverage_validation_status",
        "is_resolved",
        "original_atomic_count",
        "selected_atomic_count",
        "original_atomic_descendants",
        "selected_atomic_descendants",
    ]
    final_export = final_df[[col for col in final_cols if col in final_df.columns]].copy()

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        params_and_control.to_excel(writer, sheet_name="00_Параметры_и_контроль", index=False)
        manager_df.to_excel(writer, sheet_name="01_Менеджерский_вывод", index=False)
        final_export.to_excel(writer, sheet_name="02_Итог_аномалий", index=False)
        anomaly_analysis.to_excel(writer, sheet_name="Анализ аномалий", index=False)
        history_top.to_excel(writer, sheet_name="03_История_top", index=False)
        candidates.to_excel(writer, sheet_name="04_Диагностика_кандидатов", index=False)
        missing_zero.to_excel(writer, sheet_name="05_Пропуски_и_нули", index=False)
        control.to_excel(writer, sheet_name="06_Контроль", index=False)
        optimization_decision_log.to_excel(writer, sheet_name="07_Журнал_set_packing", index=False)

        highlight_manager_rows_on_anomaly_analysis(writer.sheets["Анализ аномалий"])

        for worksheet in writer.sheets.values():
            worksheet.freeze_panes = "A2"
            worksheet.auto_filter.ref = worksheet.dimensions
            for col_cells in worksheet.columns:
                max_length = 0
                col_letter = col_cells[0].column_letter
                for cell in col_cells[:200]:
                    value = "" if cell.value is None else str(cell.value)
                    max_length = max(max_length, len(value))
                worksheet.column_dimensions[col_letter].width = min(max(max_length + 2, 10), 48)


def run_anomaly_analysis(
    input_path: str | Path,
    output_path: str | Path,
    sheet_name: int | str = 0,
    period: Optional[str] = "1W",
    dim_cols: Optional[Sequence[str]] = None,
    current_cal_date: Optional[int] = None,
    thresholds: Optional[AnomalyThresholds] = None,
    tree_output_path: str | Path | None = None,
) -> Dict[str, pd.DataFrame]:
    """Запустить полный анализ необычных сегментов.

    Args:
        input_path: Путь к входному файлу.
        output_path: Путь к итоговому Excel-файлу.
        sheet_name: Имя или номер листа Excel.
        period: Период для фильтрации.
        dim_cols: Явно заданные признаки.
        current_cal_date: Анализируемая неделя. Если None, берётся последняя.
        thresholds: Пороги алгоритма. Если None, используются значения по умолчанию.
        tree_output_path: Необязательный путь к PNG, SVG или PDF с деревом
            листа «Анализ аномалий».

    Returns:
        Словарь таблиц результата.

    Raises:
        ValueError: Если входные данные некорректны.
        ImportError: Если запрошено дерево, но matplotlib недоступен.
        OSError: Если отчёт или дерево невозможно записать.

    Examples:
        >>> # result = run_anomaly_analysis('input.xlsx', 'output.xlsx')
    """

    thresholds = thresholds or AnomalyThresholds()
    history_df, dims, dates = load_history_table(input_path, sheet_name=sheet_name, period=period, dim_cols=dim_cols)
    current = int(current_cal_date) if current_cal_date is not None else int(dates[-1])
    panel_df = build_full_week_grid(history_df, dims, dates)
    candidates, total_by_date = build_anomaly_candidates(panel_df, dims, dates, thresholds, current)
    coverage = build_atomic_coverage(candidates, dims)
    candidates = apply_local_depth_penalty(candidates, coverage)
    final_df, diagnostics, optimization_decision_log = search_anomal(candidates, thresholds, coverage=coverage, dim_cols=dims)
    write_anomaly_excel(
        output_path,
        thresholds,
        dims,
        history_df,
        panel_df,
        diagnostics,
        final_df,
        total_by_date,
        dates,
        current,
        coverage,
        optimization_decision_log,
    )
    if tree_output_path is not None:
        build_anomaly_tree_from_excel(output_path, tree_output_path)

    return {
        "history": history_df,
        "panel": panel_df,
        "candidates": diagnostics,
        "final": final_df,
        "optimization_decision_log": optimization_decision_log,
        "control": build_control_table(history_df, panel_df, diagnostics, final_df, coverage, dates, current, total_by_date),
    }


def parse_args() -> argparse.Namespace:
    """Разобрать аргументы командной строки.

    Args:
        Нет аргументов.

    Returns:
        Объект argparse.Namespace.

    Raises:
        SystemExit: Если аргументы некорректны.

    Examples:
        >>> # args = parse_args()
    """

    parser = argparse.ArgumentParser(description="Поиск необычных сегментов GMV на истории недель.")
    parser.add_argument("--input", default=str(DEFAULT_INPUT_PATH), help="Путь к входному .xlsx/.xls/.csv файлу.")
    parser.add_argument("--output", default=str(DEFAULT_OUTPUT_PATH), help="Путь к итоговому Excel-файлу.")
    parser.add_argument(
        "--tree-output",
        default=str(DEFAULT_TREE_OUTPUT_PATH),
        help="Путь к PNG, SVG или PDF с деревом листа «Анализ аномалий». По умолчанию создаётся «Граф.png».",
    )
    parser.add_argument("--sheet-name", default=0, help="Имя или номер листа Excel.")
    parser.add_argument("--period", default="1W", help="Период для фильтрации.")
    parser.add_argument("--dims", nargs="*", default=None, help="Явный список признаков.")
    parser.add_argument("--current-cal-date", type=int, default=None, help="Текущая анализируемая неделя cal_date.")
    parser.add_argument("--min-anomaly-abs", type=float, default=AnomalyThresholds.min_anomaly_abs)
    parser.add_argument("--min-z-score", type=float, default=AnomalyThresholds.min_z_score)
    parser.add_argument("--min-materiality-share", type=float, default=AnomalyThresholds.min_materiality_share)
    parser.add_argument("--sigma-floor", type=float, default=AnomalyThresholds.sigma_floor)
    parser.add_argument("--z-cap", type=float, default=AnomalyThresholds.z_cap)
    parser.add_argument(
        "--set-packing-gap-tolerance",
        type=float,
        default=AnomalyThresholds.set_packing_gap_tolerance,
        help="Relative MIP gap tolerance required to mark set-packing optimization as OPTIMAL.",
    )
    parser.add_argument(
        "--max-exact-fallback-size",
        type=int,
        default=AnomalyThresholds.max_exact_fallback_size,
        help="Maximum conflicting component size allowed for internal exact branch-and-bound fallback.",
    )
    parser.add_argument("--max-manager-facts", type=int, default=AnomalyThresholds.max_manager_facts)
    return parser.parse_args()


def main() -> None:
    """Точка входа CLI.

    Args:
        Нет аргументов.

    Returns:
        None.

    Raises:
        ValueError: Если входные данные некорректны.

    Examples:
        >>> # python gmv_anomaly_analysis.py --input payoffline_pulse_hier_4_13w.xlsx
    """

    args = parse_args()
    try:
        sheet_name: int | str = int(args.sheet_name)
    except (TypeError, ValueError):
        sheet_name = args.sheet_name

    thresholds = AnomalyThresholds(
        min_anomaly_abs=args.min_anomaly_abs,
        min_z_score=args.min_z_score,
        min_materiality_share=args.min_materiality_share,
        sigma_floor=args.sigma_floor,
        z_cap=args.z_cap,
        set_packing_gap_tolerance=args.set_packing_gap_tolerance,
        max_exact_fallback_size=args.max_exact_fallback_size,
        max_manager_facts=args.max_manager_facts,
    )
    result = run_anomaly_analysis(
        input_path=args.input,
        output_path=args.output,
        sheet_name=sheet_name,
        period=args.period,
        dim_cols=args.dims,
        current_cal_date=args.current_cal_date,
        thresholds=thresholds,
        tree_output_path=args.tree_output,
    )

    control = result["control"].set_index("показатель")["значение"].to_dict()
    print("Готово.")
    print(f"Итоговый файл: {args.output}")
    if args.tree_output:
        tree_path = Path(args.tree_output)
        if not tree_path.suffix:
            tree_path = tree_path.with_suffix(".png")
        print(f"Дерево аномалий: {tree_path.resolve()}")
    print(f"Кандидатов: {control.get('candidate_count')}")
    print(f"Выбрано аномалий: {control.get('selected_count')}")
    print(f"Нарушения пересечения атомов: {control.get('double_count_violation_count')}")


if __name__ == "__main__":
    main()
